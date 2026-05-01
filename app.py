# ═══════════════════════════════════════════════════════════════════════════════
# Australian Disaster Data Explorer
# Samuel Marcus · Monash University · PhD: Compound Hazards & EM Capacity
# ═══════════════════════════════════════════════════════════════════════════════
import re
import json
import datetime as _dt
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from pathlib import Path

# ── Global chart readability ──────────────────────────────────────────────────
pio.templates["readable"] = go.layout.Template(
    layout=go.Layout(
        font=dict(size=14),
        title=dict(font=dict(size=16)),
        xaxis=dict(title=dict(font=dict(size=14)), tickfont=dict(size=13)),
        yaxis=dict(title=dict(font=dict(size=14)), tickfont=dict(size=13)),
        legend=dict(font=dict(size=13)),
        hoverlabel=dict(font=dict(size=13)),
        annotationdefaults=dict(font=dict(size=13)),
    )
)
pio.templates.default = "plotly+readable"

st.set_page_config(
    page_title="Australian Disaster Data Explorer",
    page_icon="🌪",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── constants ─────────────────────────────────────────────────────────────────

DATA_DIR = Path(__file__).parent
MAP_SAMPLE_THRESHOLD = 5_000   # points before sampling prompt

# Australian state/territory geographic centroids (lat, lon)
_STATE_CENTROIDS = {
    "Queensland":                  (-22.6, 143.9),
    "New South Wales":             (-31.9, 146.6),
    "Victoria":                    (-37.0, 144.5),
    "South Australia":             (-30.0, 135.8),
    "Western Australia":           (-25.5, 121.6),
    "Tasmania":                    (-42.0, 147.0),
    "Northern Territory":          (-19.4, 133.5),
    "Australian Capital Territory": (-35.3, 149.1),
}

# Simplified hazard groupings for DRFA primary hazard field
_DRFA_HAZARD_SIMPLE = {
    "Flood": "Flood", "Rainfall": "Flood",
    "Trough/monsoonal trough": "Flood", "Low/tropical low": "Flood/Cyclone",
    "Cyclone": "Cyclone",
    "Storm": "Storm", "Thunderstorm": "Storm", "Weather event": "Storm",
    "Storm surge": "Storm", "Tornado": "Storm",
    "Bushfire": "Bushfire",
    "Hailstorm": "Hail",
    "Earthquake": "Earthquake",
}

# ── AFAC National Capability constants ───────────────────────────────────────
_AFAC_STATES = ["TAS", "VIC", "NSW", "ACT", "QLD", "NT", "WA", "SA"]
_AFAC_STATE_FULL = {
    "TAS": "Tasmania",                      "VIC": "Victoria",
    "NSW": "New South Wales",               "ACT": "Australian Capital Territory",
    "QLD": "Queensland",                    "NT":  "Northern Territory",
    "WA":  "Western Australia",             "SA":  "South Australia",
}
_DRFA_TO_AFAC = {  # DRFA hazard_group → AFAC Excel sheet name
    "Flood":         "Severe Weather Response",
    "Flood/Cyclone": "Severe Weather Response",
    "Cyclone":       "Severe Weather Response",
    "Storm":         "Severe Weather Response",
    "Hail":          "Severe Weather Response",  # SES tarping/debris ops, not structural FF
    "Bushfire":      "Firefighting (bushfire)",
    "Earthquake":    "Search and Rescue",
    "Landslide":     "Search and Rescue",
}

# ABS 2022–23 state/territory population shares (used as benchmark in capability profiles)
# Source: ABS National, State and Territory Population, Sep 2023 (cat. 3101.0)
_AUS_STATE_POP_SHARE = {
    "NSW": 32.1, "VIC": 25.8, "QLD": 20.4, "WA": 10.9,
    "SA":   7.0, "TAS":  2.2, "ACT":  1.8, "NT":  1.0,
}
_AFAC_DOM_SHORT = {
    "Firefighting (bushfire)":      "FF Bushfire",
    "Firefighting (other)":         "FF Other",
    "Search and Rescue":            "SAR",
    "Severe Weather Response":      "Severe Wx",
    "HAZMAT":                       "HAZMAT",
    "Damage and Impact Assessment": "DIA",
    "Incident Management":          "IMT",
    "Aviation":                     "Aviation",
}


# Precompiled state inference rules — checked against ALL patterns to support
# multi-state events. Order: most specific first within each state.
_STATE_RULES: list[tuple[re.Pattern, str]] = [
    (re.compile(
        r"Queensland|South.?East Queensland|Southern Queensland|Northern.*Queensland"
        r"|North.*Queensland|Western Queensland|\bQLD\b|\bQld\b|\bSEQ\b"
        r"|Far North Qld|West QLD"
        r"|Tropical Cyclone (Jasper|Kirrily|Alfred|Koji|Debbie|Marcia|Ita|Yasi)"
        r"|North and Far North(?! Tropical Low)|Queensland Monsoon"
        r"|Queensland Bushfire|Queensland Flood|Queensland Storms|Queensland Rainfall"
        r"|Rockhampton|Mackay|Cairns|Townsville|Bundaberg|Toowoomba|Ipswich|Fernvale"
        r"|Emerald|Lockyer Valley|Somerset|Moreton|Sunshine Coast|Gold Coast"
        r"|Christmas Storms|Christmas and New Year Storms",
        re.IGNORECASE), "Queensland"),
    (re.compile(
        r"\bNSW\b|New South Wales|NSW North Coast|NSW East Coast|Far West Region"
        r"|Hunter Valley|Northern Rivers|Newcastle|Sydney|Blue Mountains"
        r"|Warrumbungle|Wagga|Lismore|Blacktown|East Coast Low|Northern NSW"
        r"|Shoalhaven|Coffs Harbour|Grafton|Armidale|Tamworth|Goulburn"
        r"|Mid North Coast|Clarence Valley|Kempsey|Port Macquarie|Singleton"
        r"|Cessnock|Muswellbrook|Oberon|Maitland|Snowy|Penrith|Richmond Valley"
        r"|Greater Taree|Bourke|Glen Innes|Tenterfield|Cooma|Hilltops|Carrathool",
        re.IGNORECASE), "New South Wales"),
    (re.compile(
        r"Victorian|Western Victoria|Victoria(?!\s*Bushfire.*QLD)|\bVIC\b"
        r"|Melbourne|Gippsland|Black Saturday|Great Ocean Rd|Great Ocean Road|Wye River"
        r"|Bunyip|Latrobe Valley|Bendigo|Ballarat|Mornington Peninsula"
        r"|Yarra Ranges|Pyrenees|Thomson Catchment|Rosedale|Timbarra"
        r"|Corangamite|Grampians|Colac|Surf Coast|Gannawarra|Loddon"
        r"|East Gippsland|West Gippsland|South Gippsland|Central Victoria",
        re.IGNORECASE), "Victoria"),
    (re.compile(
        r"South Australian|South Australia(?! Flood.*QLD)|\bSA\b"
        r"|\bAdelaide\b|Sampson.?s? Flat|Pinery|Kangaroo Island|Coober Pedy"
        r"|Mount Lofty|Eyre Peninsula|Fleurieu|Cherryville|Kingston|Tulka|Coomunga"
        r"|Bundaleer|Clare Valley|Yorke Peninsula|Barossa",
        re.IGNORECASE), "South Australia"),
    (re.compile(
        r"Tasmanian|Tasmania|\bTAS\b|Hobart|Tasman Peninsula|Dunalley|Launceston"
        r"|Devonport|Molesworth|Circular Head|St Helens|Bicheno|Boomer Bay"
        r"|North West Tasmania|Southern Tasmania|East Coast Tasmania",
        re.IGNORECASE), "Tasmania"),
    (re.compile(
        r"Australian Capital Territory|\bACT\b|Canberra",
        re.IGNORECASE), "Australian Capital Territory"),
    (re.compile(
        r"Western Australia|\bWA\b|Wooroloo|Perth|Margaret River|Leschenault"
        r"|Pilbara|Kimberley|Broome|Busselton|Gascoyne|Wheatbelt|Esperance"
        r"|Carnarvon|Geraldton|Kalgoorlie|Collie|Augusta|Mandurah|Rockingham"
        r"|Tropical Cyclone (George|Hilda|Lua|Rusty|Olwyn|Christine|Heidi|Kelvin|Blake)"
        r"|Cyclone (George|Jacob)|South West Land Division|Great Southern Region"
        r"|Shire of Wandering|Shire of Kellerberrin",
        re.IGNORECASE), "Western Australia"),
    (re.compile(
        r"Northern Territory|\bNT\b|Northern Region"
        r"|Ex.?TC Ellie|Ex.?Tropical Cyclone Ellie|North and Far North Tropical Low"
        r"|Darwin|Daly River|Kakadu|Top End|Alice Springs|Nhulunbuy"
        r"|Tropical Cyclone (Marcus|Carlos|Lam|Trevor|Nathan|Heatlie)"
        r"|Cyclone (Marcus|Grant|Carlos)|Central Australian|Arnhem",
        re.IGNORECASE), "Northern Territory"),
]

DATASET_SOURCES = {
    "Knowledge Hub": dict(
        name="AEMI/AIDR Emergency Management Knowledge Hub — Merged",
        provider="Australian Institute for Disaster Resilience (AIDR) / Attorney-General's Department",
        url="https://knowledge.aidr.org.au/resources/disaster-mapper/",
        description="Unified dataset merging the AIDR Disaster Mapper (Excel, 2023) with the AGD geocoded "
                    "export (data.gov.au CSV, 2014). AIDR fields take precedence; AGD contributes "
                    "lat/lon and structural damage columns (homes, buildings, livestock, etc.). "
                    "Events present in both sources are flagged 'Both'; events unique to one source are "
                    "flagged 'AIDR only' or 'AGD only'.",
        coverage="1727–2023",
        notes="Includes events outside Australia affecting Australians (zone = 'Offshore'). "
              "AIDR last updated March 2023; AGD export covers up to 2014 only.",
    ),
    "DRFA Activations": dict(
        name="DRFA Activation History by LGA",
        provider="National Emergency Management Agency (NEMA)",
        url="https://data.gov.au/data/dataset/drfa-activation-history-by-lga",
        description="Disaster Recovery Funding Arrangements (DRFA) activation history by Local Government Area, "
                    "including hazard type, DRFA categories (A–D), and payment eligibility flags.",
        coverage="2006–March 2026",
        notes="Updated March 2026. '<20' values in claim counts are a privacy sentinel for small numbers.",
    ),
    "DRFA Payments": dict(
        name="Disaster History Payments",
        provider="Services Australia / National Emergency Management Agency (NEMA)",
        url="https://www.nema.gov.au/our-work/disaster-recovery/disaster-recovery-payments",
        description="Australian Government Disaster Recovery Payment (AGDRP) and Disaster Recovery Allowance (DRA) "
                    "claim statistics by disaster, location, and payment type.",
        coverage="2009–present (updated March 2026)",
        notes="'<20' values are a privacy sentinel for small claim counts. "
              "Dollars reflect total granted/paid at the recorded date snapshot.",
    ),
    "DRFA Merged": dict(
        name="DRFA Activations joined with DRFA Payment Aggregates",
        provider="NEMA / Services Australia (merged)",
        url="https://data.gov.au/data/dataset/drfa-activation-history-by-lga",
        description="Each row is one LGA–event activation record (from the activations dataset), enriched "
                    "with event-level payment totals aggregated from the payments dataset. Payment columns "
                    "are present only for the 44 events where payment data is available (joined on AGRN).",
        coverage="Activations: 2006–March 2026 | Payments: 2009–present",
        notes="Payment figures are summed across all payment types and locations for each AGRN. "
              "Activations without a matching AGRN in the payments data will show NaN payment columns.",
    ),
    "EMDAT": dict(
        name="EM-DAT International Disaster Database — Australia",
        provider="Centre for Research on the Epidemiology of Disasters (CRED), UCLouvain",
        url="https://public.emdat.be/",
        description="Global database of natural and technological disasters from 1900 to present. "
                    "Records deaths, injuries, affected persons, and economic losses for each event.",
        coverage="1939–present (Australia subset)",
        notes="Free registration required for full download. Damage figures in '000 USD. "
              "'Adjusted' columns normalised to 2022 USD. Access via public.emdat.be.",
    ),
    "ICA Catastrophes": dict(
        name="ICA Historical Normalised Catastrophe Loss Data",
        provider="Insurance Council of Australia (ICA)",
        url="https://insurancecouncil.com.au/industry-members/data-hub/",
        description="Insurance catastrophe events with original and inflation-normalised insured losses "
                    "(2022 baseline). Covers declared catastrophes and significant events tracked by the ICA. "
                    "~94% of losses from weather-related perils (storms, floods, bushfires, cyclones).",
        coverage="1967–February 2026",
        notes="Normalised losses adjust for changes in property values, inflation, and building codes to "
              "2022 AUD. 'SE' prefix = Significant Event; 'CAT' prefix = declared Catastrophe.",
    ),
}

# ── utilities (pure functions, no Streamlit calls) ────────────────────────────

def read_csv_with_schema(
    path,
    date_cols: dict | None = None,
    num_cols: list | None = None,
    strip_col_names: bool = False,
) -> pd.DataFrame:
    """Read CSV with optional date/numeric parsing and column name stripping."""
    df = pd.read_csv(path, low_memory=False)
    if strip_col_names:
        df.columns = df.columns.str.strip()
    if date_cols:
        for col, fmt in date_cols.items():
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], format=fmt, errors="coerce")
    if num_cols:
        for col in num_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def validate_df(df: pd.DataFrame, required_cols: list[str], name: str = "dataset") -> None:
    """Raise ValueError listing any required columns missing from df."""
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"[{name}] Missing required columns: {missing}")


def to_num_dollars(series: pd.Series) -> pd.Series:
    """Strip $, commas, spaces; replace 'nan' with pd.NA; coerce to numeric."""
    return (
        series.astype(str)
        .str.replace(r"[\$,\s]", "", regex=True)
        .replace("nan", pd.NA)
        .pipe(pd.to_numeric, errors="coerce")
    )


def to_num_counts(series: pd.Series) -> pd.Series:
    """Strip commas; replace <20 with 10; replace 'nan' with pd.NA; coerce to numeric.

    '<20' is a privacy sentinel used in DRFA/Services Australia CSVs to suppress
    small claim counts.  It is replaced with 10 (midpoint imputation, lower bound 1,
    upper bound 19).  Sensitivity to this choice can be assessed with
    src.methods.sentinel_strategy.sentinel_sensitivity_table().

    Any aggregation (sum, mean) over columns derived from this function should be
    treated as approximate where sentinel values are present.
    """
    try:
        from src.methods.sentinel_strategy import apply_sentinel, SentinelStrategy
        return apply_sentinel(series, SentinelStrategy.MIDPOINT)
    except ImportError:
        # Fallback: replicate behaviour inline if src/ is not on path
        return (
            series.astype(str)
            .str.replace(",", "", regex=False)
            .str.replace("<20", "10", regex=False)
            .replace("nan", pd.NA)
            .pipe(pd.to_numeric, errors="coerce")
        )


def _infer_state_from_name(name: str) -> str:
    """
    Extract Australian state/territory from a disaster name string.

    Checks ALL rules and returns the first match (preserving original behaviour).
    Events matching multiple states are first-match assigned; the companion
    function _infer_state_confidence() exposes the full match list for auditing.

    Use src/validation/state_inference_audit.py to produce a full audit report.
    """
    n = str(name)
    for pattern, state in _STATE_RULES:
        if pattern.search(n):
            return state
    return "Unknown"


def _infer_state_confidence(name: str) -> tuple[str, list[str]]:
    """
    Return (confidence_class, all_matched_states) for a disaster name.

    confidence_class: "EXACT" (1 match) | "MULTI" (>1 match) | "UNKNOWN" (0 matches)
    all_matched_states: every state whose pattern matched (empty list if UNKNOWN)

    This is a diagnostic companion to _infer_state_from_name() — it does not
    change the primary state assignment used in downstream analyses.
    """
    n = str(name)
    matches = [state for pattern, state in _STATE_RULES if pattern.search(n)]
    if len(matches) == 0:
        return "UNKNOWN", []
    if len(matches) == 1:
        return "EXACT", matches
    return "MULTI", matches



def raw_cols(df: pd.DataFrame) -> list[str]:
    """Return column names not starting with '_'."""
    return [c for c in df.columns if not c.startswith("_")]


def fmt_dates(df: pd.DataFrame) -> pd.DataFrame:
    """Return a display copy with all date/datetime columns formatted as dd/mm/yyyy."""
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime("%d/%m/%Y").where(out[col].notna(), other="")
        elif out[col].dtype == object:
            # Python date objects stored as object dtype (e.g. after .dt.date)
            sample = out[col].dropna().head(5)
            if not sample.empty and all(isinstance(v, _dt.date) for v in sample):
                out[col] = out[col].apply(
                    lambda v: v.strftime("%d/%m/%Y") if isinstance(v, _dt.date) and pd.notna(v) else v
                )
    return out


# ── data loaders ──────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner="Loading Knowledge Hub (AIDR + AGD)…")
def load_knowledge_hub() -> pd.DataFrame:
    """
    Merge AIDR Disaster Mapper (primary) with AGD geocoded export.
    AIDR fields win on conflicts. AGD contributes lat/lon and structural damage columns.
    _source_flag: 'Both' | 'AIDR only' | 'AGD only'
    """
    # ── Load AIDR ────────────────────────────────────────────────────────────
    aidr = pd.read_excel(
        DATA_DIR / "AIDR_disaster_mapper_data.xlsx",
        sheet_name="Disaster Mapper Data",
    )
    aidr.columns = aidr.columns.str.strip()
    for col in ("Start Date", "End Date"):
        aidr[col] = pd.to_datetime(aidr[col], dayfirst=True, errors="coerce")
    aidr["Category"] = aidr["Category"].str.strip().str.title()
    aidr["_key"] = aidr["Event"].str.strip().str.lower()

    # ── Load AGD ─────────────────────────────────────────────────────────────
    agd = read_csv_with_schema(
        DATA_DIR / "au-govt-agd-disaster-events-impact-location-na.csv",
        date_cols={"startdate": "%m/%d/%Y", "enddate": "%m/%d/%Y"},
        num_cols=["lat", "lon", "deaths", "injuries", "homeless",
                  "homes_damaged", "homes_destroyed", "buildings_damaged",
                  "buildings_destroyed", "farms_damaged", "farms_destroyed",
                  "crops_destroyed", "livestock_destroyed"],
    )
    agd["_key"] = agd["title"].str.strip().str.lower()

    # AGD columns to carry — lat/lon only (damage columns are ≤7% populated, not usable)
    _AGD_CARRY = ["_key", "lat", "lon"]
    agd_slim = agd[[c for c in _AGD_CARRY if c in agd.columns]].copy()
    # De-duplicate AGD keys — keep first occurrence (titles are unique in practice)
    agd_slim = agd_slim.drop_duplicates(subset=["_key"])

    # ── Outer merge on title key ──────────────────────────────────────────────
    merged = aidr.merge(agd_slim, on="_key", how="outer", indicator=True)
    merged["_source_flag"] = merged["_merge"].map({
        "both":       "Both",
        "left_only":  "AIDR only",
        "right_only": "AGD only",
    })
    merged = merged.drop(columns=["_merge"])

    # ── For AGD-only rows: fill AIDR columns from AGD equivalents ─────────────
    agd_only = merged["_source_flag"] == "AGD only"
    # Map AGD full records back for AGD-only rows
    agd_full = agd.drop_duplicates(subset=["_key"]).set_index("_key")
    for idx in merged[agd_only].index:
        key = merged.at[idx, "_key"]
        if key not in agd_full.index:
            continue
        row = agd_full.loc[key]
        merged.at[idx, "Event"]      = row["title"]
        merged.at[idx, "Start Date"] = row["startdate"]
        merged.at[idx, "End Date"]   = row.get("enddate", pd.NaT)
        merged.at[idx, "Fatalities"] = row.get("deaths", pd.NA)
        merged.at[idx, "Injured"]    = row.get("injuries", pd.NA)
        merged.at[idx, "Zone"]       = row.get("regions", pd.NA)
        merged.at[idx, "Description"] = row.get("description", pd.NA)
        merged.at[idx, "URL"]        = row.get("url", pd.NA)
        # Derive category from title prefix
        raw = str(row["title"]).split("-")[0].strip().split("–")[0].strip()
        merged.at[idx, "Category"]   = "Cyclone" if raw.lower().startswith("cyclone") else raw.title()

    # ── Shared computed columns ───────────────────────────────────────────────
    def _parse_cost(v) -> float:
        if pd.isna(v):
            return float("nan")
        s = str(v).replace(",", "")
        m = re.search(r"[\d]+(?:\.\d+)?", s)
        return float(m.group()) if m else float("nan")

    merged["_num_insured_cost"] = merged["Insured Cost"].apply(_parse_cost)
    merged["_num_fatalities"]   = pd.to_numeric(merged["Fatalities"], errors="coerce")
    merged["_num_injured"]      = pd.to_numeric(merged["Injured"],    errors="coerce")
    # Ensure mixed-type columns are uniformly string-typed for Arrow serialisation.
    # Raw values include numerics, NaN, and sentinels like 'A ', 'Not Available'.
    for _col in ("Fatalities", "Injured", "Insured Cost"):
        if _col in merged.columns:
            merged[_col] = (
                merged[_col].astype(str).str.strip()
                .replace({"nan": None, "None": None, "": None})
            )
    merged["_num_year"]         = pd.to_datetime(merged["Start Date"], errors="coerce").dt.year

    _ZONE_NORM = {
        "New South Wales": "New South Wales", "Victoria": "Victoria",
        "Queensland": "Queensland", "South Australia": "South Australia",
        "Western Australia": "Western Australia", "Tasmania": "Tasmania",
        "Northern Territory": "Northern Territory",
        "Australian Capital Territory": "Australian Capital Territory",
        "National": "National", "Offshore": "Offshore",
        "Australia Wide": "National", "Outside Australia": "Offshore",
    }
    def _parse_zone(v) -> list[str]:
        parts = [p.strip() for p in str(v).split(",") if p.strip()]
        result = []
        for p in parts:
            norm = _ZONE_NORM.get(p)
            if norm and norm not in result:
                result.append(norm)
        return result if result else ["Unknown"]

    merged["_states"]     = merged["Zone"].apply(_parse_zone)
    merged["_states_str"] = merged["_states"].apply(lambda s: ", ".join(s))

    # Clean up internal key
    merged = merged.drop(columns=["_key"])

    validate_df(merged, ["Event", "Category", "Start Date"], name="Knowledge Hub")
    return merged


@st.cache_data(ttl=3600, show_spinner="Loading DRFA activations…")
def load_drfa_activations() -> pd.DataFrame:
    df = read_csv_with_schema(
        DATA_DIR / "drfa_activation_history_by_location_2026_march_19.csv",
        date_cols={"disaster_start_date": "%Y-%m-%d"},
        num_cols=["cat_A", "cat_B", "cat_C", "cat_D", "AGDRP", "DRA"],
    )
    validate_df(
        df,
        ["Location_Name", "STATE", "event_name", "agrn", "hazard_type", "disaster_start_date"],
        name="DRFA Activations",
    )
    df["_num_year"] = df["disaster_start_date"].dt.year
    return df


@st.cache_data(ttl=3600, show_spinner="Loading DRFA payments…")
def load_drfa_payments() -> pd.DataFrame:
    df = read_csv_with_schema(
        DATA_DIR / "disaster_history_payments_2026_march_19.csv",
        date_cols={"Date of Data": "%Y-%m-%d"},
    )
    validate_df(
        df,
        ["State Name", "Disaster Name", "Payment Type Name", "Dollars Paid ($)", "Dollars Granted ($)"],
        name="DRFA Payments",
    )
    unknown_mask = df["State Name"] == "Unknown"
    # Apply inference only to rows whose state is labelled "Unknown" in the source CSV
    if unknown_mask.any():
        inferred_names = df.loc[unknown_mask, "Disaster Name"]
        df.loc[unknown_mask, "State Name"] = inferred_names.apply(_infer_state_from_name)
        # Attach confidence class so downstream UI can warn on uncertain assignments
        conf_map = inferred_names.apply(lambda n: _infer_state_confidence(n)[0])
        df["_state_infer_conf"] = "LABELLED"
        df.loc[unknown_mask, "_state_infer_conf"] = conf_map.values
    else:
        df["_state_infer_conf"] = "LABELLED"

    df["_num_paid"] = to_num_dollars(df["Dollars Paid ($)"])
    df["_num_granted"] = to_num_dollars(df["Dollars Granted ($)"])
    df["_num_eligible"] = to_num_counts(df["Eligible Claims (No.)"])
    df["_num_total_received"] = to_num_counts(df["Total Recieved Claims (No.)"])  # sic — typo matches the source CSV column name exactly
    return df


@st.cache_data(ttl=3600, show_spinner="Loading EM-DAT…")
def load_emdat() -> pd.DataFrame:
    df = read_csv_with_schema(
        DATA_DIR / "EMDAT_Disaster_Aus.csv",
        num_cols=[
            "Start Year", "End Year", "Start Month", "Start Day", "End Month", "End Day",
            "Total Deaths", "No. Injured", "No. Affected", "No. Homeless", "Total Affected",
            "Total Damage ('000 US$)", "Total Damage, Adjusted ('000 US$)",
            "Insured Damage ('000 US$)", "Insured Damage, Adjusted ('000 US$)",
            "Reconstruction Costs ('000 US$)", "Reconstruction Costs, Adjusted ('000 US$)",
            "AID Contribution ('000 US$)", "Magnitude", "CPI", "Latitude", "Longitude",
        ],
        strip_col_names=True,
    )
    validate_df(df, ["DisNo.", "Disaster Type", "Start Year"], name="EM-DAT")
    # Compound flag: EM-DAT records associated hazard types when a second peril co-occurred.
    # This is CRED's own contemporaneous classification, not a derived proximity measure.
    df["_is_compound"] = (
        df["Associated Types"].notna() & (df["Associated Types"].astype(str).str.strip() != "")
    )
    return df


@st.cache_data(ttl=3600, show_spinner="Loading DRFA merged dataset…")
def load_drfa_merged() -> pd.DataFrame:
    act = load_drfa_activations()   # reuse cached loader — includes _num_year
    pay = load_drfa_payments()      # reuse cached loader — includes _num_paid etc.
    # Prepare payments for join.
    # The payments CSV has one row per payment type per (agrn, location) — e.g. AGDRP and DRA
    # are separate rows for the same event-location pair.  Joining directly causes fanout and
    # inflates the merged row count from ~5,967 to 8,000+.  Aggregate to one row per
    # (agrn, Location_Name) first, summing dollar/claim amounts and concatenating type labels.
    pay_join = pay.rename(columns={"Disaster AGRN": "agrn", "Location Name": "Location_Name"})
    pay_agg = (
        pay_join.groupby(["agrn", "Location_Name"], as_index=False)
        .agg(
            Payment_Types=(     "Payment Type Name",    lambda x: "; ".join(x.dropna().unique())),
            _num_paid=(         "_num_paid",            "sum"),
            _num_granted=(      "_num_granted",         "sum"),
            _num_eligible=(     "_num_eligible",        "sum"),
            _num_total_received=("_num_total_received", "sum"),
        )
    )
    merged = act.merge(pay_agg, on=["agrn", "Location_Name"], how="left")
    merged["Has_Payment_Data"] = merged["Payment_Types"].notna().map({True: "Yes", False: "No"})
    return merged


@st.cache_data(ttl=3600, show_spinner="Loading ICA catastrophes…")
def load_ica() -> pd.DataFrame:
    df = read_csv_with_schema(
        DATA_DIR / "ICA-Historical-Normalised-Catastrophe-Master-Updated-2026_02.csv",
        num_cols=["Year"],
        strip_col_names=True,
    )
    df["Event Start"] = pd.to_datetime(df["Event Start"], format="%d-%b-%y", errors="coerce")
    df["Event Finish"] = pd.to_datetime(df["Event Finish"], format="%d-%b-%y", errors="coerce")
    # Fix 2-digit year ambiguity
    mask = df["Event Start"].dt.year > df["Year"] + 1
    df.loc[mask, "Event Start"] -= pd.DateOffset(years=100)
    df.loc[mask, "Event Finish"] -= pd.DateOffset(years=100)
    validate_df(df, ["CAT Name", "Event Name", "Event Start", "Type", "Year"], name="ICA Catastrophes")
    df["_num_orig_loss"] = to_num_dollars(df["ORIGINAL LOSS VALUE"])
    df["_num_norm_loss"] = to_num_dollars(df["NORMALISED LOSS VALUE (2022)"])
    df["_num_total_claims"] = to_num_counts(df["TOTAL CLAIMS RECEIVED"])
    df["_num_dom_building"] = to_num_counts(df["Domestic Building Claims"])
    df["_num_dom_content"] = to_num_counts(df["Domestic Content Claims"])
    df["_num_dom_motor"] = to_num_counts(df["Domestic Motor Claims"])
    df["_num_com_property"] = to_num_counts(df["Commercial Property Claims"])
    return df


@st.cache_data(ttl=3600, show_spinner="Loading AIDR Disaster Mapper…")
def load_aidr() -> pd.DataFrame:
    df = pd.read_excel(
        DATA_DIR / "AIDR_disaster_mapper_data.xlsx",
        sheet_name="Disaster Mapper Data",
    )
    # Normalise column names
    df.columns = df.columns.str.strip()

    # Parse dates — Excel returns datetime for recent rows, strings for historical rows
    for col in ("Start Date", "End Date"):
        df[col] = pd.to_datetime(df[col], dayfirst=True, errors="coerce")

    # Normalise category (fix mixed case 'flood' / 'Flood')
    df["Category"] = df["Category"].str.strip().str.title()

    # Parse insured cost — extract first numeric value found, ignore text notes
    def _parse_cost(v) -> float:
        if pd.isna(v):
            return float("nan")
        s = str(v).replace(",", "")
        m = re.search(r"[\d]+(?:\.\d+)?", s)
        return float(m.group()) if m else float("nan")

    df["_num_insured_cost"] = df["Insured Cost"].apply(_parse_cost)

    # Parse numeric impact columns
    df["_num_fatalities"] = pd.to_numeric(df["Fatalities"], errors="coerce")
    df["_num_injured"]    = pd.to_numeric(df["Injured"],    errors="coerce")
    # Normalise mixed-type columns to uniform strings for Arrow serialisation.
    for _col in ("Fatalities", "Injured", "Insured Cost"):
        if _col in df.columns:
            df[_col] = (
                df[_col].astype(str).str.strip()
                .replace({"nan": None, "None": None, "": None})
            )

    # Year helper from Start Date
    df["_num_year"] = df["Start Date"].dt.year

    # Parse Zone into a list of state names (comma-separated; 'Offshore'/'National' kept as-is)
    _AIDR_STATE_NORM = {
        "New South Wales": "New South Wales",
        "Victoria": "Victoria",
        "Queensland": "Queensland",
        "South Australia": "South Australia",
        "Western Australia": "Western Australia",
        "Tasmania": "Tasmania",
        "Northern Territory": "Northern Territory",
        "Australian Capital Territory": "Australian Capital Territory",
        "National": "National",
        "Offshore": "Offshore",
        "Australia Wide": "National",
    }
    def _parse_zone(v) -> list[str]:
        parts = [p.strip() for p in str(v).split(",") if p.strip()]
        result = []
        for p in parts:
            norm = _AIDR_STATE_NORM.get(p)
            if norm and norm not in result:
                result.append(norm)
        return result if result else ["Unknown"]

    df["_states"] = df["Zone"].apply(_parse_zone)
    # Flat string for display/filtering
    df["_states_str"] = df["_states"].apply(lambda s: ", ".join(s))

    validate_df(df, ["Event", "Category", "Start Date"], name="AIDR Disaster Mapper")
    return df


@st.cache_data(ttl=3600, show_spinner="Loading DRFA event summary…")
def load_drfa_events() -> pd.DataFrame:
    """
    Event-level DRFA summary (one row per AGRN).
    Aggregates activation rows by event, adds simplified hazard group,
    and attaches state centroid coordinates for mapping.
    """
    act = read_csv_with_schema(
        DATA_DIR / "drfa_activation_history_by_location_2026_march_19.csv",
        date_cols={"disaster_start_date": "%Y-%m-%d"},
        num_cols=["cat_A", "cat_B", "cat_C", "cat_D", "AGDRP", "DRA"],
    )
    ev = act.groupby("agrn", as_index=False).agg(
        event_name=("event_name", "first"),
        start=("disaster_start_date", "min"),
        states_str=("STATE", lambda x: ", ".join(sorted(x.dropna().unique()))),
        state_primary=("STATE", lambda x: x.dropna().mode().iat[0] if not x.dropna().empty else "Unknown"),
        hazard_type=("hazard_type", "first"),
        lga_count=("Location_Name", "nunique"),
        cat_A_lgas=("cat_A", "sum"),
        cat_B_lgas=("cat_B", "sum"),
        cat_C_lgas=("cat_C", "sum"),
        cat_D_lgas=("cat_D", "sum"),
    )
    ev["primary_hazard"] = ev["hazard_type"].str.split(",").str[0].str.strip()
    ev["hazard_group"] = ev["primary_hazard"].map(_DRFA_HAZARD_SIMPLE).fillna("Other")
    ev["year"] = ev["start"].dt.year.astype("Int64")
    ev["year_month"] = ev["start"].dt.to_period("M").dt.to_timestamp()
    # Attach state centroid for mapping
    ev["lat"] = ev["state_primary"].map(lambda s: _STATE_CENTROIDS.get(s, (None, None))[0])
    ev["lon"] = ev["state_primary"].map(lambda s: _STATE_CENTROIDS.get(s, (None, None))[1])
    return ev


@st.cache_data(ttl=3600, show_spinner="Computing compound disaster clusters…")
def load_compound_disasters(nl_threshold_m: float = 100.0, window_days: int = 91):
    """
    Gissing et al. (2022) compound disaster methodology applied to the ICA dataset.

    Events ≥ nl_threshold_m million (2022 AUD normalised loss) are clustered within
    each Australian financial year (1 Jul – 30 Jun) using chain-link start-date proximity:
    if consecutive sorted events start within `window_days` of each other they belong to
    the same cluster.  A cluster of ≥ 2 events is a 'compound disaster'.

    Returns:
        events_df  — ICA events ≥ threshold with columns: _fy, _peril, _cluster_id,
                      _is_compound, _magnitude
        clusters_df — one row per cluster with: fy, cluster_start, n_events, total_nl,
                       perils, event_names, _is_compound, _magnitude
    """
    ica = load_ica()

    # Map ICA Type field to Gissing et al. (2022) peril categories
    _PERIL_MAP = {
        "Flood":      "Flood",
        "Storm":      "Storm",
        "Cyclone":    "Tropical Cyclone",
        "Bushfire":   "Bushfire",
        "Hail":       "Storm",          # grouped with Storm (hail is a storm subtype)
        "Heatwave":   "Heatwave",
        "Landslide":  "Landslide",
        "Earthquake": "Earthquake",
        "Man-made":   "Other",
    }

    # Filter to events with known normalised loss ≥ threshold and a valid start date
    df = ica[
        (ica["_num_norm_loss"] >= nl_threshold_m * 1e6) &
        ica["Event Start"].notna()
    ].copy()
    df["_peril"] = df["Type"].map(_PERIL_MAP).fillna(df["Type"].fillna("Unknown"))

    # Australian financial year: month ≥ 7 → same calendar year; else previous year
    df["_fy"] = df["Event Start"].apply(
        lambda d: d.year if d.month >= 7 else d.year - 1
    )

    # Sort within FY by start date for chain-link clustering
    df = df.sort_values(["_fy", "Event Start"]).reset_index(drop=True)

    # Chain-link clustering: new cluster if gap from previous event exceeds window
    df["_cluster_id"] = -1
    cluster_id = 0
    for _fy_val, grp in df.groupby("_fy"):
        idx    = grp.index.tolist()
        starts = grp["Event Start"].tolist()
        df.at[idx[0], "_cluster_id"] = cluster_id
        prev_start = starts[0]
        for i in range(1, len(idx)):
            if (starts[i] - prev_start).days > window_days:
                cluster_id += 1
            df.at[idx[i], "_cluster_id"] = cluster_id
            prev_start = starts[i]
        cluster_id += 1

    # Cluster-level summary
    clusters = df.groupby("_cluster_id").agg(
        fy=(             "_fy",             "first"),
        cluster_start=(  "Event Start",     "min"),
        cluster_end=(    "Event Finish",     lambda x: x.dropna().max() if x.notna().any() else pd.NaT),
        n_events=(       "CAT Name",         "nunique"),
        total_nl=(       "_num_norm_loss",   "sum"),
        min_event_nl=(   "_num_norm_loss",   "min"),
        perils=(         "_peril",           lambda x: "; ".join(sorted(x.dropna().unique()))),
        event_names=(    "Event Name",       lambda x: "; ".join(list(x.unique())[:5])),
    ).reset_index()
    clusters["_is_compound"] = clusters["n_events"] >= 2

    # Gissing et al. (2022) Compound Disaster Magnitude Scale (Table 3)
    # Classification is based on individual event loss threshold AND component count.
    # Each component event must individually exceed the threshold for that tier.
    def _magnitude(row):
        if not row["_is_compound"]:
            return None
        n        = row["n_events"]
        min_nl_m = row["min_event_nl"] / 1e6   # minimum individual event NL in $M
        if min_nl_m >= 20_000:   # each event ≥ $20B
            return "VIII" if n > 2 else "VII"
        elif min_nl_m >= 5_000:  # each event ≥ $5B
            return "VI" if n > 2 else "V"
        elif min_nl_m >= 1_000:  # each event ≥ $1B
            return "IV" if n > 2 else "III"
        elif min_nl_m >= 100:    # each event ≥ $100M
            return "II" if n > 3 else "I"
        return None

    clusters["_magnitude"] = clusters.apply(_magnitude, axis=1)

    # Annotate events with compound flag and magnitude from their cluster
    df = df.merge(
        clusters[["_cluster_id", "_is_compound", "_magnitude"]],
        on="_cluster_id", how="left",
    )

    return df, clusters


@st.cache_data(ttl=3600, show_spinner="Computing DRFA compound disaster clusters…")
def load_compound_disasters_drfa(window_days: int = 91):
    """
    Gissing et al. (2022) compound disaster methodology adapted for the DRFA dataset.

    Each unique AGRN = one declared disaster event. Events are clustered within each
    Australian financial year (1 Jul – 30 Jun) using chain-link onset-date proximity
    (identical to the ICA version). A cluster of ≥ 2 events is a 'compound disaster'.

    Severity proxy: total LGA activations within a cluster (replaces normalised loss,
    which is unavailable for most DRFA events).

    DRFA-adapted Magnitude Scale (DGMS):
        Rows: # component events (2, 3, ≥4)
        Cols: total cluster LGA activations (<5 | 5–20 | >20)

    Returns:
        events_df   — DRFA events (event-level, one row per agrn) with _fy, _peril,
                      _cluster_id, _is_compound, _magnitude, _lga_count, _states
        clusters_df — one row per cluster with fy, cluster_start, n_events, total_lgas,
                      perils, event_names, states, _is_compound, _magnitude
    """
    act = load_drfa_activations()

    # Aggregate to event level: one row per unique AGRN
    ev = act.groupby("agrn").agg(
        event_name=(            "event_name",                   "first"),
        hazard_type=(           "hazard_type",                  "first"),
        disaster_start_date=(   "disaster_start_date",          "first"),
        _lga_count=(            "Location_Name",                "nunique"),
        _state_count=(          "STATE",                        "nunique"),
        _states=(               "STATE",                        lambda x: "; ".join(sorted(x.unique()))),
        _highest_cat=(          "highest_drfa_category_group",  "first"),
    ).reset_index()

    # Map DRFA hazard_type to Gissing peril categories.
    # hazard_type can be a compound string like "Flood, Storm" — use the first token.
    _DRFA_PERIL_MAP = {
        "Flood":                    "Flood",
        "Storm":                    "Storm",
        "Cyclone":                  "Tropical Cyclone",
        "Bushfire":                 "Bushfire",
        "Hail":                     "Storm",
        "Hailstorm":                "Storm",
        "Thunderstorm":             "Storm",
        "Tornado":                  "Storm",
        "Low/tropical low":         "Tropical Cyclone",
        "Trough/monsoonal trough":  "Flood",
        "Rainfall":                 "Flood",
        "Storm surge":              "Storm",
        "Weather event":            "Other",
        "Earthquake":               "Earthquake",
        "Landslide":                "Landslide",
    }

    def _map_drfa_peril(hazard):
        if pd.isna(hazard):
            return "Unknown"
        primary = hazard.split(",")[0].strip()
        return _DRFA_PERIL_MAP.get(primary, "Other")

    ev["_peril"] = ev["hazard_type"].apply(_map_drfa_peril)
    ev = ev[ev["disaster_start_date"].notna()].copy()

    # Australian financial year
    ev["_fy"] = ev["disaster_start_date"].apply(
        lambda d: d.year if d.month >= 7 else d.year - 1
    )

    ev = ev.sort_values(["_fy", "disaster_start_date"]).reset_index(drop=True)

    # Chain-link clustering — identical algorithm to load_compound_disasters()
    ev["_cluster_id"] = -1
    cluster_id = 0
    for _fy_val, grp in ev.groupby("_fy"):
        idx    = grp.index.tolist()
        starts = grp["disaster_start_date"].tolist()
        ev.at[idx[0], "_cluster_id"] = cluster_id
        prev_start = starts[0]
        for i in range(1, len(idx)):
            if (starts[i] - prev_start).days > window_days:
                cluster_id += 1
            ev.at[idx[i], "_cluster_id"] = cluster_id
            prev_start = starts[i]
        cluster_id += 1

    # Cluster-level summary
    clusters = ev.groupby("_cluster_id").agg(
        fy=(            "_fy",                  "first"),
        cluster_start=( "disaster_start_date",  "min"),
        n_events=(      "agrn",                 "nunique"),
        total_lgas=(    "_lga_count",           "sum"),
        perils=(        "_peril",               lambda x: "; ".join(sorted(x.dropna().unique()))),
        event_names=(   "event_name",           lambda x: "; ".join(list(x.unique())[:5])),
        states=(        "_states",              lambda x: "; ".join(sorted({s for v in x for s in v.split("; ")}))),
    ).reset_index()
    clusters["_is_compound"] = clusters["n_events"] >= 2

    # DRFA Government Magnitude Scale (DGMS):
    # Rows = # component events (2, 3, ≥4); Cols = total cluster LGA activations
    _DGMS = {
        (2, "<5"):  "I",   (2, "5-20"): "II",  (2, ">20"): "III",
        (3, "<5"):  "II",  (3, "5-20"): "III", (3, ">20"): "IV",
        (4, "<5"):  "III", (4, "5-20"): "IV",  (4, ">20"): "V",
    }

    def _magnitude_drfa(row):
        if not row["_is_compound"]:
            return None
        n       = min(row["n_events"], 4)
        lgas    = row["total_lgas"]
        lga_bkt = "<5" if lgas < 5 else ("5-20" if lgas <= 20 else ">20")
        return _DGMS.get((n, lga_bkt))

    clusters["_magnitude"] = clusters.apply(_magnitude_drfa, axis=1)

    ev = ev.merge(
        clusters[["_cluster_id", "_is_compound", "_magnitude"]],
        on="_cluster_id", how="left",
    )
    return ev, clusters


@st.cache_data(ttl=3600, show_spinner="Computing FY climate phases…")
def load_climate_fy_phases() -> pd.DataFrame:
    """
    Aggregate ONI, SAM, and IOD monthly data to Australian financial year (FY) level.
    FY label = start calendar year (FY2000 = 1 Jul 2000 – 30 Jun 2001).

    Returns one row per FY with:
        fy          – int (start year)
        oni_mean    – float  (FY mean ONI)
        enso_phase  – "El Niño" / "La Niña" / "Neutral"
        sam_mean    – float  (FY mean SAM/AAO)
        sam_phase   – "Positive SAM" / "Negative SAM" / "Neutral"
        dmi_mean    – float  (May–Nov active-season mean DMI)
        iod_phase   – "Positive IOD" / "Negative IOD" / "Neutral"
    """
    def _assign_fy(s: pd.Series) -> pd.Series:
        return s.apply(lambda d: d.year if d.month >= 7 else d.year - 1)

    try:
        oni = fetch_oni_data()[["date", "month", "oni"]].copy()
        oni["fy"] = _assign_fy(oni["date"])
    except Exception:
        return pd.DataFrame()  # ONI required; bail without it

    try:
        sam = fetch_sam_data()[["date", "month", "sam"]].copy()
        sam["fy"] = _assign_fy(sam["date"])
    except Exception:
        sam = None

    try:
        iod = fetch_iod_data()[["date", "month", "dmi"]].copy()
        iod["fy"] = _assign_fy(iod["date"])
    except Exception:
        iod = None

    fy_min = int(oni["fy"].min())
    fy_max = int(oni["fy"].max())

    # ONI: FY mean (require ≥ 6 months)
    oni_fy = oni.groupby("fy")["oni"].agg(oni_mean="mean", _n="count").reset_index()
    oni_fy.loc[oni_fy["_n"] < 6, "oni_mean"] = float("nan")
    oni_fy["enso_phase"] = "Neutral"
    oni_fy.loc[oni_fy["oni_mean"] >=  0.5, "enso_phase"] = "El Niño"
    oni_fy.loc[oni_fy["oni_mean"] <= -0.5, "enso_phase"] = "La Niña"
    oni_fy.drop(columns="_n", inplace=True)

    fy_df = (
        pd.DataFrame({"fy": range(fy_min, fy_max + 1)})
        .merge(oni_fy, on="fy", how="left")
    )

    # SAM: FY mean (require ≥ 6 months)
    if sam is not None:
        sam_fy = sam.groupby("fy")["sam"].agg(sam_mean="mean", _n="count").reset_index()
        sam_fy.loc[sam_fy["_n"] < 6, "sam_mean"] = float("nan")
        sam_fy["sam_phase"] = "Neutral"
        sam_fy.loc[sam_fy["sam_mean"] >=  1.0, "sam_phase"] = "Positive SAM"
        sam_fy.loc[sam_fy["sam_mean"] <= -1.0, "sam_phase"] = "Negative SAM"
        sam_fy.drop(columns="_n", inplace=True)
        fy_df = fy_df.merge(sam_fy, on="fy", how="left")
    else:
        fy_df["sam_mean"] = float("nan")
        fy_df["sam_phase"] = None

    # IOD: sustained-event classification over May–Nov active season.
    # A FY is classified as Positive/Negative IOD if ≥ 3 of the 7 active-season months
    # exceed ±0.4 °C (BoM threshold). Averaging DMI over the season dilutes short events
    # to near-zero and produces near-universal Neutral classification — this approach
    # captures sustained episodes instead.
    if iod is not None:
        iod_act = iod[iod["month"].isin([5, 6, 7, 8, 9, 10, 11])].copy()
        iod_act["is_pos"] = (iod_act["dmi"] >=  0.4).astype(int)
        iod_act["is_neg"] = (iod_act["dmi"] <= -0.4).astype(int)
        iod_fy = iod_act.groupby("fy").agg(
            dmi_mean=("dmi",    "mean"),
            pos_months=("is_pos", "sum"),
            neg_months=("is_neg", "sum"),
            _n=("dmi",    "count"),
        ).reset_index()
        iod_fy.loc[iod_fy["_n"] < 4, "dmi_mean"] = float("nan")
        iod_fy["iod_phase"] = "Neutral"
        iod_fy.loc[iod_fy["pos_months"] >= 3, "iod_phase"] = "Positive IOD"
        iod_fy.loc[iod_fy["neg_months"] >= 3, "iod_phase"] = "Negative IOD"
        # Resolve the rare conflict where both thresholds are met: dominant phase wins
        conflict = (iod_fy["pos_months"] >= 3) & (iod_fy["neg_months"] >= 3)
        iod_fy.loc[conflict & (iod_fy["pos_months"] >= iod_fy["neg_months"]), "iod_phase"] = "Positive IOD"
        iod_fy.loc[conflict & (iod_fy["neg_months"] >  iod_fy["pos_months"]), "iod_phase"] = "Negative IOD"
        iod_fy = iod_fy[["fy", "dmi_mean", "iod_phase"]]
        fy_df = fy_df.merge(iod_fy, on="fy", how="left")
    else:
        fy_df["dmi_mean"] = float("nan")
        fy_df["iod_phase"] = None

    return fy_df


@st.cache_data(show_spinner=False)
def load_state_geojson() -> dict | None:
    """Load pre-generated Australian state boundary GeoJSON (dissolved from LGA shapefile)."""
    path = DATA_DIR / "aus_states.geojson"
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


@st.cache_data(ttl=3600, show_spinner="Loading AFAC 2023 capability data…")
def load_afac_capability() -> pd.DataFrame:
    """
    Parse the 2023 AFAC National Capability Statement Excel into a tidy DataFrame.
    Columns: domain, function, resource, national, TAS…SA, deployable (bool).
    deployable=True rows are interstate teams deployable within 48 hours.
    """
    import openpyxl

    xls = DATA_DIR / "2023-national-capability-statement-data.xlsx"
    if not xls.exists():
        return pd.DataFrame()

    def _int(v) -> int:
        if v is None:
            return 0
        try:
            s = str(v).strip()
            return 0 if s in ("-", "–", "", "None") else int(float(s))
        except (ValueError, TypeError):
            return 0

    records: list[dict] = []
    wb = openpyxl.load_workbook(xls, read_only=True, data_only=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            r = tuple(row[:12]) + (None,) * max(0, 12 - len(row[:12]))
            if any(c is not None for c in r):
                rows.append(r)

        is_deployable   = False
        cur_fn          = ""
        is_aviation     = (sheet == "Aviation")
        aviation_source = "National Contracted"  # tracks sub-table within Aviation sheet

        for r in rows:
            c = [str(v or "").strip() for v in r]

            # ── Skip header / section-title rows ──────────────────────────────
            # Deployable section marker must be checked BEFORE the c[3] skip
            # because its row also has c[3]="National Capability".
            if c[1] == "Deployable Capability":
                is_deployable = True
                continue
            if "Teams available to deploy" in c[2]:
                continue
            if c[1] == "Function statement":
                continue
            if c[3] in ("National capability", "National Capability", "Jurisdiction totals"):
                continue
            if c[2] in ("National Capability", "State contracted"):
                continue
            # state-abbreviation header row (TAS VIC NSW …)
            if any(r[i] == "TAS" and i + 1 < 12 and r[i + 1] == "VIC" for i in range(11)):
                continue
            # Aviation section headers — update sub-table tracker, then skip
            if c[0] == "National Contracted Aircraft":
                aviation_source = "National Contracted"
                continue
            if c[0] == "State Contracted and State Owned Aircraft":
                aviation_source = "State Contracted/Owned"
                continue
            if c[0] in ("Aircraft Type", "Total aircraft"):
                continue

            # ── Track current function label ──────────────────────────────────
            if c[0] and c[0] != "None":
                cur_fn = c[0]

            # ── Column layout differs for Aviation ────────────────────────────
            if is_aviation:
                resource = c[1]
                national = _int(r[2])
                sv       = [_int(r[3 + i]) for i in range(8)]
            else:
                resource = c[2] or c[1]
                national = _int(r[3])
                sv       = [_int(r[4 + i]) for i in range(8)]

            if not resource or resource in ("None", "Jurisdiction totals",
                                             "National capability", "National Capability"):
                continue
            if national == 0 and all(v == 0 for v in sv):
                continue

            records.append({
                "domain":     sheet,
                "function":   cur_fn,
                "resource":   resource,
                "national":   national,
                **{s: sv[i] for i, s in enumerate(_AFAC_STATES)},
                "deployable": is_deployable,
                "section":    aviation_source if is_aviation else "Inventory",
            })

    return pd.DataFrame(records)


_N_STATES = 8  # TAS, VIC, NSW, ACT, QLD, NT, WA, SA





def _monthly_cache_fresh(cache_path: "Path", lag_months: int = 2) -> "pd.DataFrame | None":
    """Return cached DataFrame if it has data up to (today − lag_months), else None."""
    if not cache_path.exists():
        return None
    try:
        df = pd.read_csv(cache_path, parse_dates=["date"])
        threshold = (pd.Timestamp.now() - pd.DateOffset(months=lag_months)).replace(day=1).normalize()
        if not df.empty and df["date"].max() >= threshold:
            return df
    except Exception:
        pass
    return None


def _daily_cache_fresh(cache_path: "Path", lag_days: int = 3) -> "pd.DataFrame | None":
    """Return cached DataFrame if it has data up to (today − lag_days), else None."""
    if not cache_path.exists():
        return None
    try:
        df = pd.read_csv(cache_path, parse_dates=["date"])
        threshold = pd.Timestamp.now().normalize() - pd.Timedelta(days=lag_days)
        if not df.empty and df["date"].max() >= threshold:
            return df
    except Exception:
        pass
    return None


@st.cache_data(ttl=3600, show_spinner="Checking ONI data from NOAA…")
def fetch_oni_data() -> pd.DataFrame:
    """
    Fetch and parse the Oceanic Niño Index (ONI) from the NOAA CPC HTML table.

    Returns a DataFrame with columns:
        date        – month-start datetime (e.g. 2010-01-01 = Jan 2010)
        year        – int
        month       – int (1–12)
        oni         – float  (3-month running mean SST anomaly, Niño 3.4 region)
        enso_phase  – str  ("El Niño" / "La Niña" / "Neutral")

    ONI is released monthly by NOAA CPC (~10th of the month for the previous
    3-month season). Cache is refreshed when the local copy is more than 2 months
    behind; otherwise the cached CSV is returned without a network call.
    Falls back to local CSV if NOAA is unreachable.
    """
    _ONI_URL = (
        "https://www.cpc.ncep.noaa.gov/products/analysis_monitoring/"
        "ensostuff/ONI_v5.php"
    )
    _ONI_CACHE = DATA_DIR / "oni_cache.csv"

    fresh = _monthly_cache_fresh(_ONI_CACHE, lag_months=2)
    if fresh is not None:
        return fresh

    # Each 3-month season code maps to the calendar month of its centre month.
    _SEASON_MONTH: dict[str, int] = {
        "DJF": 1, "JFM": 2, "FMA": 3, "MAM": 4,
        "AMJ": 5, "MJJ": 6, "JJA": 7, "JAS": 8,
        "ASO": 9, "SON": 10, "OND": 11, "NDJ": 12,
    }

    try:
        import requests as _req
        resp = _req.get(
            _ONI_URL,
            headers={"User-Agent": "Mozilla/5.0 (compatible; research-tool/1.0)"},
            timeout=15,
        )
        resp.raise_for_status()
        tables = pd.read_html(resp.text, header=0)

        oni_raw = None
        for t in tables:
            cols = [str(c).strip().upper() for c in t.columns]
            if "YEAR" in cols and "DJF" in cols:
                oni_raw = t.copy()
                break
        if oni_raw is None:
            raise ValueError("ONI table not found on NOAA page.")
    except Exception:
        if _ONI_CACHE.exists():
            return pd.read_csv(_ONI_CACHE, parse_dates=["date"])
        raise

    oni_raw.columns = [str(c).strip().upper() for c in oni_raw.columns]
    season_cols = [c for c in oni_raw.columns if c in _SEASON_MONTH]
    oni_raw = oni_raw[["YEAR"] + season_cols].copy()
    oni_raw["YEAR"] = pd.to_numeric(oni_raw["YEAR"], errors="coerce")
    oni_raw = oni_raw.dropna(subset=["YEAR"])
    oni_raw["YEAR"] = oni_raw["YEAR"].astype(int)

    long = oni_raw.melt(
        id_vars="YEAR", value_vars=season_cols, var_name="season", value_name="oni"
    )
    long["month"] = long["season"].map(_SEASON_MONTH)
    long["year"]  = long["YEAR"]
    long["oni"]   = pd.to_numeric(long["oni"], errors="coerce")
    long = long.dropna(subset=["oni"])
    long["date"]  = pd.to_datetime(dict(year=long["year"], month=long["month"], day=1))

    long = long[["date", "year", "month", "season", "oni"]].sort_values("date").reset_index(drop=True)

    long["enso_phase"] = "Neutral"
    long.loc[long["oni"] >= 0.5,  "enso_phase"] = "El Niño"
    long.loc[long["oni"] <= -0.5, "enso_phase"] = "La Niña"

    long.to_csv(_ONI_CACHE, index=False)
    return long


@st.cache_data(ttl=3600, show_spinner="Checking SAM data…")
def fetch_sam_data() -> pd.DataFrame:
    """
    Fetch the Southern Annular Mode (SAM) index from two complementary sources:

    1. BAS Marshall SAM (1957–2007, station-based, baseline 1971–2000):
       https://legacy.bas.ac.uk/met/gjma/newsam.1957.2007.txt
       — static legacy file; provides the long historical record.

    2. NOAA CPC AAO (1979–present, reanalysis-based, updated monthly):
       https://www.cpc.ncep.noaa.gov/products/precip/CWlink/daily_ao_index/
       aao/monthly.aao.index.b79.current.ascii.table
       — live source; fills the gap from 2008 to present.

    Strategy: BAS data is used for 1957–1978 (pre-reanalysis era); NOAA CPC AAO
    is used from 1979 onwards. The two series are not identical in variance (different
    methodologies) but are highly correlated (r > 0.95 for overlapping years).

    Returns a DataFrame with columns:
        date        – month-start datetime
        year        – int
        month       – int (1–12)
        sam         – float  (SAM/AAO index value)
        sam_phase   – str  ("Positive SAM" / "Negative SAM" / "Neutral")
        source      – str  ("BAS Marshall" / "NOAA CPC AAO")

    Cache is refreshed when local copy is more than 2 months behind the current
    date. Falls back to local CSV if both sources are unreachable.
    """
    _SAM_URL_BAS = "https://legacy.bas.ac.uk/met/gjma/newsam.1957.2007.txt"
    _SAM_URL_CPC = (
        "https://www.cpc.ncep.noaa.gov/products/precip/CWlink/"
        "daily_ao_index/aao/monthly.aao.index.b79.current.ascii.table"
    )
    _SAM_CACHE = DATA_DIR / "sam_cache.csv"
    _SAM_THRESHOLD = 1.0

    fresh = _monthly_cache_fresh(_SAM_CACHE, lag_months=2)
    if fresh is not None:
        return fresh

    import requests as _req
    from io import StringIO as _StringIO

    _MONTH_COLS = ["jan", "feb", "mar", "apr", "may", "jun",
                   "jul", "aug", "sep", "oct", "nov", "dec"]

    def _parse_wide_table(text: str, source_label: str) -> pd.DataFrame:
        raw = pd.read_csv(
            _StringIO(text),
            sep=r"\s+", header=0,
            names=["year"] + _MONTH_COLS,
        )
        raw = raw[pd.to_numeric(raw["year"], errors="coerce").notna()].copy()
        raw["year"] = raw["year"].astype(int)
        long = raw.melt(id_vars="year", value_vars=_MONTH_COLS,
                        var_name="month_name", value_name="sam")
        long["month"] = pd.Categorical(
            long["month_name"], categories=_MONTH_COLS, ordered=True
        ).codes + 1
        long["sam"] = pd.to_numeric(long["sam"], errors="coerce")
        long = long.dropna(subset=["sam"])
        long["date"] = pd.to_datetime(dict(year=long["year"], month=long["month"], day=1))
        long["source"] = source_label
        return long[["date", "year", "month", "sam", "source"]]

    bas_df = cpc_df = None

    try:
        resp = _req.get(
            _SAM_URL_BAS,
            headers={"User-Agent": "Mozilla/5.0 (compatible; research-tool/1.0)"},
            timeout=15, verify=False,
        )
        resp.raise_for_status()
        bas_df = _parse_wide_table(resp.text, "BAS Marshall")
    except Exception:
        pass

    try:
        resp = _req.get(
            _SAM_URL_CPC,
            headers={"User-Agent": "Mozilla/5.0 (compatible; research-tool/1.0)"},
            timeout=15,
        )
        resp.raise_for_status()
        cpc_df = _parse_wide_table(resp.text, "NOAA CPC AAO")
    except Exception:
        pass

    if bas_df is None and cpc_df is None:
        if _SAM_CACHE.exists():
            return pd.read_csv(_SAM_CACHE, parse_dates=["date"])
        raise ValueError("Both SAM sources (BAS and NOAA CPC) are unreachable.")

    if bas_df is not None and cpc_df is not None:
        # BAS for pre-1979 historical; NOAA CPC AAO from 1979 onwards
        parts = [bas_df[bas_df["year"] < 1979], cpc_df]
        long = pd.concat(parts, ignore_index=True)
    else:
        long = bas_df if bas_df is not None else cpc_df

    long = long.sort_values("date").drop_duplicates("date", keep="last").reset_index(drop=True)

    long["sam_phase"] = "Neutral"
    long.loc[long["sam"] >=  _SAM_THRESHOLD, "sam_phase"] = "Positive SAM"
    long.loc[long["sam"] <= -_SAM_THRESHOLD, "sam_phase"] = "Negative SAM"

    long.to_csv(_SAM_CACHE, index=False)
    return long


@st.cache_data(ttl=3600, show_spinner="Checking IOD (DMI) data…")
def fetch_iod_data() -> pd.DataFrame:
    """
    Fetch and parse the Dipole Mode Index (DMI / IOD).

    Primary source: NOAA PSL / HadISST1.1 monthly series (January 1870 – ~12 months ago).
    Gap-fill source: BoM GAMSSA weekly series (July 2008 – present), aggregated to monthly
    means, appended for any months not covered by the PSL series.

    Returns a DataFrame with columns:
        date        – month-start datetime
        year        – int
        month       – int (1–12)
        dmi         – float  (SST anomaly west minus east, °C)
        iod_phase   – str  ("Positive IOD" / "Negative IOD" / "Neutral")
        source      – str  ("HadISST1.1" / "BoM GAMSSA")

    Phase threshold: ±0.4 °C (BoM operational standard).
    HadISST1.1 lags ~2 months; BoM GAMSSA is weekly. Cache is refreshed when the
    local copy is more than 2 months behind. Falls back to local CSV if both sources
    are unreachable.
    """
    _IOD_URL_PSL = "https://psl.noaa.gov/data/timeseries/month/data/dmi.had.long.csv"
    _IOD_URL_BOM = "https://www.bom.gov.au/clim_data/IDCK000072/iod_1.txt"
    _IOD_CACHE   = DATA_DIR / "iod_cache.csv"
    _IOD_THRESHOLD = 0.4  # °C — BoM operational positive/negative IOD threshold

    fresh = _monthly_cache_fresh(_IOD_CACHE, lag_months=2)
    if fresh is not None:
        return fresh

    import requests as _req
    from io import StringIO as _StringIO

    def _fetch_psl() -> pd.DataFrame:
        resp = _req.get(
            _IOD_URL_PSL,
            headers={"User-Agent": "Mozilla/5.0 (compatible; research-tool/1.0)"},
            timeout=15,
        )
        resp.raise_for_status()
        raw = pd.read_csv(
            _StringIO(resp.text),
            skiprows=1, header=None, names=["date", "dmi"],
        )
        raw["date"] = pd.to_datetime(raw["date"], errors="coerce")
        raw["dmi"]  = pd.to_numeric(raw["dmi"],  errors="coerce")
        raw = raw.dropna(subset=["date"])
        raw = raw[raw["dmi"] > -9000].copy()
        raw["source"] = "HadISST1.1"
        return raw

    def _fetch_bom() -> pd.DataFrame:
        resp = _req.get(
            _IOD_URL_BOM,
            headers={"User-Agent": "Mozilla/5.0 (compatible; research-tool/1.0)"},
            timeout=15,
        )
        resp.raise_for_status()
        # Format: YYYYMMDD,YYYYMMDD,DMI_value  (no header, comma-separated)
        raw = pd.read_csv(
            _StringIO(resp.text),
            sep=",", header=None, names=["start", "end", "dmi"],
        )
        raw["date"] = pd.to_datetime(raw["start"].astype(str), format="%Y%m%d", errors="coerce")
        raw["dmi"]  = pd.to_numeric(raw["dmi"], errors="coerce")
        raw = raw.dropna(subset=["date", "dmi"])
        # Aggregate weekly → monthly mean, snap date to month-start
        raw["date"] = raw["date"].dt.to_period("M").dt.to_timestamp()
        monthly = raw.groupby("date")["dmi"].mean().reset_index()
        monthly["source"] = "BoM GAMSSA"
        return monthly

    try:
        psl = _fetch_psl()
        psl_last = psl["date"].max()

        try:
            bom = _fetch_bom()
            # Only keep BoM months strictly after the last PSL month
            bom_ext = bom[bom["date"] > psl_last].copy()
        except Exception:
            bom_ext = pd.DataFrame(columns=["date", "dmi", "source"])

        combined = pd.concat([psl, bom_ext], ignore_index=True)

    except Exception:
        if _IOD_CACHE.exists():
            return pd.read_csv(_IOD_CACHE, parse_dates=["date"])
        raise

    combined["date"]  = pd.to_datetime(combined["date"])
    combined["year"]  = combined["date"].dt.year
    combined["month"] = combined["date"].dt.month
    long = combined[["date", "year", "month", "dmi", "source"]].sort_values("date").reset_index(drop=True)

    long["iod_phase"] = "Neutral"
    long.loc[long["dmi"] >=  _IOD_THRESHOLD, "iod_phase"] = "Positive IOD"
    long.loc[long["dmi"] <= -_IOD_THRESHOLD, "iod_phase"] = "Negative IOD"

    long.to_csv(_IOD_CACHE, index=False)
    return long


@st.cache_data(ttl=3600, show_spinner="Checking MJO (RMM) data from BoM…")
def fetch_mjo_data() -> pd.DataFrame:
    """
    Fetch and parse the BoM Real-time Multivariate MJO (RMM) index.

    Daily data from 1 June 1974 to ~2 days before present (updated daily by BoM).
    Returns a DataFrame with columns:
        date        – daily datetime
        year, month, day – int
        rmm1        – float  PC1 of combined OLR+U850+U200 (positive → Indian Ocean active)
        rmm2        – float  PC2 (~90° phase shift, quadrature component)
        phase       – int    1–8 (geographic location of active convection); 0/9 = weak
        amplitude   – float  sqrt(rmm1²+rmm2²); > 1.0 = active MJO event

    Phase sectors in RMM1-RMM2 space (counterclockwise from 180°):
        1 (180°–225°): W. Hemisphere & Africa
        2 (225°–270°): Indian Ocean
        3 (270°–315°): Indian Ocean
        4 (315°–360°): Maritime Continent
        5 (  0°– 45°): Maritime Continent
        6 ( 45°– 90°): Western Pacific
        7 ( 90°–135°): W. Pacific & Pacific
        8 (135°–180°): W. Hemisphere & Africa

    Reference: Wheeler & Hendon (2004), Mon. Wea. Rev., 132, 1917–1932.
    Source: Australian Bureau of Meteorology (product IDCKGEM000).
    BoM updates ~daily with a 2-day lag. Cache is refreshed when the local copy
    is more than 3 days behind. Falls back to local CSV if BoM is unreachable.
    """
    _MJO_URL   = "http://www.bom.gov.au/clim_data/IDCKGEM000/rmm.74toRealtime.txt"
    _MJO_CACHE = DATA_DIR / "mjo_cache.csv"

    fresh = _daily_cache_fresh(_MJO_CACHE, lag_days=3)
    if fresh is not None:
        return fresh

    try:
        import requests as _req
        resp = _req.get(
            _MJO_URL,
            headers={"User-Agent": "Mozilla/5.0 (compatible; research-tool/1.0)"},
            timeout=20,
        )
        resp.raise_for_status()

        records = []
        for line in resp.text.strip().split("\n")[2:]:   # skip 2 header lines
            parts = line.split()
            if len(parts) < 7:
                continue
            try:
                yr, mo, dy = int(parts[0]), int(parts[1]), int(parts[2])
                r1, r2     = float(parts[3]), float(parts[4])
                ph, amp    = int(parts[5]), float(parts[6])
            except (ValueError, IndexError):
                continue
            if abs(r1) > 900 or abs(r2) > 900 or amp > 900:
                continue                         # drop missing-value sentinels
            records.append((yr, mo, dy, r1, r2, ph, amp))

        df = pd.DataFrame(records,
                          columns=["year", "month", "day", "rmm1", "rmm2", "phase", "amplitude"])
        df["date"] = pd.to_datetime(dict(year=df["year"], month=df["month"], day=df["day"]))
        df = df[["date", "year", "month", "day", "rmm1", "rmm2", "phase", "amplitude"]]
        df = df.sort_values("date").reset_index(drop=True)

    except Exception:
        if _MJO_CACHE.exists():
            return pd.read_csv(_MJO_CACHE, parse_dates=["date"])
        raise

    df.to_csv(_MJO_CACHE, index=False)
    return df


@st.cache_data(ttl=3600, show_spinner="Computing ICA simultaneity…")
def compute_ica_simultaneity() -> pd.DataFrame:
    """
    For each ICA event, compute how many other events have an overlapping active
    period.  Active period = [Event Start, Event Finish]; if no Finish, use
    Start + 30 days.

    The ICA CSV has one row per event (744 events total).  497 events share the
    CAT Name "Undeclared" (significant events that never received a formal CAT/SE
    number), so CAT Name cannot be used as a unique key — integer row index is used
    instead.

    Columns added:
        concurrent_count   — number of other ICA events active at the same time
        concurrent_events  — names of up to 3 concurrent events
        concurrent_states  — deduplicated states of concurrent events
        _p_start / _p_end  — active period datetimes (used for Gantt chart)
    """
    raw = load_ica()
    df = (
        raw[raw["Event Start"].notna()]
        .sort_values("Event Start")
        .reset_index(drop=True)
        .copy()
    )
    # Use integer row index as the unique event key (CAT Name is not unique —
    # 497 events share "Undeclared").
    df["_idx"] = df.index

    df["_p_start"] = pd.to_datetime(df["Event Start"])
    df["_p_end"]   = pd.to_datetime(df["Event Finish"])
    no_end = df["_p_end"].isna()
    df.loc[no_end, "_p_end"] = df.loc[no_end, "_p_start"] + pd.Timedelta(days=30)

    ev   = df[["_idx", "CAT Name", "Event Name", "State", "_p_start", "_p_end"]].copy()
    ev_b = ev.rename(columns={c: c + "_b" for c in ev.columns})
    ev["_k"] = 1;  ev_b["_k"] = 1
    pairs = ev.merge(ev_b, on="_k").drop(columns=["_k"])
    pairs = pairs[pairs["_idx"] != pairs["_idx_b"]]   # exclude self-pairs by row index
    pairs = pairs[
        (pairs["_p_start"]   <= pairs["_p_end_b"]) &
        (pairs["_p_start_b"] <= pairs["_p_end"])
    ]

    def _top3(s):
        vals = list(s.dropna().unique())
        return "; ".join(vals[:3]) + ("…" if len(vals) > 3 else "")

    conc = pairs.groupby("_idx").agg(
        concurrent_count=(  "_idx_b",      "nunique"),
        concurrent_events=( "Event Name_b", _top3),
        concurrent_states=( "State_b",      lambda x: ", ".join(sorted({
            s.strip() for v in x.dropna() for s in str(v).split(",") if s.strip()
        }))),
    ).reset_index()

    df = df.merge(conc, on="_idx", how="left")
    df["concurrent_count"]  = df["concurrent_count"].fillna(0).astype(int)
    df["concurrent_events"] = df["concurrent_events"].fillna("")
    df["concurrent_states"] = df["concurrent_states"].fillna("")
    return df


def compute_concurrent_events(events: pd.DataFrame, duration_days: int) -> pd.DataFrame:
    """
    Sweep-line algorithm: for each calendar day, count how many DRFA events
    were 'active' (defined as: disaster start ≤ day ≤ disaster start + duration_days).

    Complexity: O(n_events + n_days) — suitable for 800+ events over 20 years.

    Args:
        events:        event-level DRFA DataFrame with 'start' column (datetime).
        duration_days: assumed active window per event (user-configurable).

    Returns:
        DataFrame with columns ['date', 'active_events'].
    """
    import numpy as np

    starts = pd.to_datetime(events["start"].dropna())
    if starts.empty:
        return pd.DataFrame(columns=["date", "active_events"])

    ends = starts + pd.Timedelta(days=duration_days)

    # Build delta series: +1 at start, -1 one day after end
    deltas = pd.concat([
        pd.DataFrame({"date": starts, "delta": 1}),
        pd.DataFrame({"date": ends + pd.Timedelta(days=1), "delta": -1}),
    ]).groupby("date")["delta"].sum().sort_index()

    full_idx = pd.date_range(starts.min(), ends.max(), freq="D")
    result = (
        deltas.reindex(full_idx, fill_value=0)
        .cumsum()
        .reset_index()
    )
    result.columns = ["date", "active_events"]
    return result


@st.cache_data(ttl=3600, show_spinner="Computing state co-occurrence matrix…")
def compute_state_cooccurrence(duration_days: int = 180):
    """
    For each calendar day (2006–present) determine which Australian states have at
    least one active DRFA event (active = disaster_start ≤ day ≤ disaster_start +
    duration_days).  Then count the number of days each pair of states was
    simultaneously active.

    Args:
        duration_days: assumed active window per event (matches concurrency analysis).

    Returns:
        matrix_df  — symmetric DataFrame (state × state) of co-occurrence day counts.
                     Diagonal = days that state had any active event.
        daily_df   — daily DataFrame with one column per state (1/0 active flag),
                     plus 'date' and 'n_states_active'.
    """
    import numpy as np

    act = read_csv_with_schema(
        DATA_DIR / "drfa_activation_history_by_location_2026_march_19.csv",
        date_cols={"disaster_start_date": "%Y-%m-%d"},
        num_cols=[],
    )

    # One row per (STATE, agrn) — avoids LGA-level inflation
    state_events = (
        act[["STATE", "agrn", "disaster_start_date"]]
        .dropna(subset=["STATE", "disaster_start_date"])
        .drop_duplicates(subset=["STATE", "agrn"])
        .rename(columns={"disaster_start_date": "start"})
    )

    states = sorted(state_events["STATE"].dropna().unique())
    if state_events.empty or not states:
        empty = pd.DataFrame(dtype=int)
        return empty, pd.DataFrame(columns=["date", "n_states_active"])

    min_date = state_events["start"].min()
    max_date = state_events["start"].max() + pd.Timedelta(days=duration_days)
    all_days = pd.date_range(min_date, max_date, freq="D")
    n_days = len(all_days)

    # Build binary matrix: shape (n_days, n_states)
    state_arr = np.zeros((n_days, len(states)), dtype=np.int8)

    for col_idx, state in enumerate(states):
        st_ev = state_events[state_events["STATE"] == state]
        starts = st_ev["start"]
        ends = starts + pd.Timedelta(days=duration_days)

        deltas = pd.concat([
            pd.DataFrame({"date": starts, "delta": 1}),
            pd.DataFrame({"date": ends + pd.Timedelta(days=1), "delta": -1}),
        ]).groupby("date")["delta"].sum().sort_index()

        active_series = (
            deltas.reindex(all_days, fill_value=0)
            .cumsum()
            .clip(lower=0)
        )
        state_arr[:, col_idx] = (active_series.values > 0).astype(np.int8)

    # Symmetric co-occurrence matrix via dot product  (diagonal = solo-active days)
    cooc = state_arr.T.astype(int) @ state_arr.astype(int)  # (n_states, n_states)
    matrix_df = pd.DataFrame(cooc, index=states, columns=states)

    # Daily summary DataFrame
    daily_df = pd.DataFrame(state_arr, index=all_days, columns=states)
    daily_df["n_states_active"] = daily_df[states].sum(axis=1)
    daily_df = daily_df.reset_index().rename(columns={"index": "date"})

    return matrix_df, daily_df

# ── UI components ─────────────────────────────────────────────────────────────

def source_box(name, provider, url, description, coverage, notes=None):
    with st.expander("Data source", expanded=False):
        st.markdown(
            f"**{name}**  \n"
            f"*Provider:* {provider}  \n"
            f"*Access:* [{url}]({url})  \n"
            f"*Coverage:* {coverage}  \n"
            f"*Description:* {description}"
        )
        if notes:
            st.caption(notes)


def download_button(df: pd.DataFrame, label: str, filename: str):
    csv = fmt_dates(df[raw_cols(df)]).to_csv(index=False).encode("utf-8")
    st.download_button(f"⬇ Download {label} as CSV", data=csv,
                       file_name=filename, mime="text/csv")


def year_slider(year_series: pd.Series, key: str) -> tuple[int, int]:
    yr_min = int(year_series.min())
    yr_max = int(year_series.max())
    return st.slider("Year range", yr_min, yr_max, (yr_min, yr_max), key=key)


# ── page renderers ────────────────────────────────────────────────────────────

def render_map():
    st.title("Disaster Event Map")
    st.caption(
        "Plots geo-referenced events from the Knowledge Hub (AIDR+AGD merged; ~673 events with coordinates) "
        "and EM-DAT (23 events with coordinates). Use the controls below to filter."
    )

    kh = load_knowledge_hub()
    emdat = load_emdat()

    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        show_kh = st.checkbox("Knowledge Hub Events", value=True)
        show_emdat = st.checkbox("EM-DAT Events", value=True)
    with fc2:
        combined_years = pd.concat([
            kh["_num_year"].dropna(),
            emdat["Start Year"].dropna(),
        ])
        sel_years = year_slider(combined_years, key="map_years")
    with fc3:
        loc_filter = st.selectbox(
            "Location scope",
            ["All", "Australia only", "Outside Australia"],
        )

    frames = []

    if show_kh:
        kh_map = kh[
            (kh["_num_year"] >= sel_years[0]) &
            (kh["_num_year"] <= sel_years[1]) &
            kh["lat"].notna() & kh["lon"].notna() &
            ~((kh["lat"] == 0) & (kh["lon"] == 0)) &
            kh["lat"].between(-90, 90) &
            kh["lon"].between(-180, 180)
        ].copy()
        if loc_filter == "Australia only":
            kh_map = kh_map[~kh_map["_states_str"].str.contains("Offshore", na=False)]
        elif loc_filter == "Outside Australia":
            kh_map = kh_map[kh_map["_states_str"].str.contains("Offshore", na=False)]
        kh_map = kh_map.assign(
            _source="Knowledge Hub",
            _label=kh_map["Event"],
            _type=kh_map["Category"],
            _hover=(
                "Source: Knowledge Hub<br>"
                + "Type: " + kh_map["Category"].fillna("") + "<br>"
                + "Year: " + kh_map["_num_year"].astype("Int64").astype(str) + "<br>"
                + "Deaths: " + kh_map["_num_fatalities"].fillna(0).astype(int).astype(str)
            ),
        )[["lat", "lon", "_source", "_label", "_type", "_hover"]]
        frames.append(kh_map)

    if show_emdat:
        emdat_map = emdat.rename(columns={"Latitude": "lat", "Longitude": "lon"}).copy()
        emdat_map = emdat_map[
            (emdat_map["Start Year"] >= sel_years[0]) &
            (emdat_map["Start Year"] <= sel_years[1]) &
            emdat_map["lat"].notna() & emdat_map["lon"].notna()
        ]
        emdat_map = emdat_map.assign(
            _source="EM-DAT",
            _label=emdat_map["DisNo."],
            _type=emdat_map["Disaster Type"],
            _hover=(
                "Source: EM-DAT<br>"
                + "Type: " + emdat_map["Disaster Type"].fillna("") + "<br>"
                + "Year: " + emdat_map["Start Year"].astype("Int64").astype(str) + "<br>"
                + "Deaths: " + emdat_map["Total Deaths"].fillna(0).astype(int).astype(str)
            ),
        )[["lat", "lon", "_source", "_label", "_type", "_hover"]]
        frames.append(emdat_map)

    if frames:
        plot_df = pd.concat(frames, ignore_index=True)
        st.markdown(f"Showing **{len(plot_df):,} events** ({', '.join(plot_df['_source'].unique())})")

        # Map sampling
        if len(plot_df) > MAP_SAMPLE_THRESHOLD:
            show_all = st.checkbox(
                f"Show all {len(plot_df):,} points (may be slow)",
                value=False,
                key="map_show_all",
            )
            if not show_all:
                plot_df = plot_df.sample(MAP_SAMPLE_THRESHOLD, random_state=42)

        fig = px.scatter_map(
            plot_df,
            lat="lat", lon="lon",
            color="_source",
            hover_name="_label",
            hover_data={"_hover": True, "lat": False, "lon": False,
                        "_source": False, "_label": False, "_type": False},
            labels={"_source": "Dataset"},
            color_discrete_map={"Knowledge Hub": "#e15759", "EM-DAT": "#4e79a7"},
            zoom=3, height=620,
            map_style="open-street-map",
        )
        fig.update_traces(marker=dict(size=8, opacity=0.8))
        fig.update_layout(margin={"r": 0, "t": 0, "l": 0, "b": 0}, legend_title_text="Dataset")
        st.plotly_chart(fig, width="stretch")

        with st.expander("Event list (filtered)"):
            st.dataframe(
                plot_df[["_label", "_source", "_type", "lat", "lon"]].rename(columns={
                    "_label": "Event", "_source": "Dataset", "_type": "Disaster Type"
                }).reset_index(drop=True),
                width="stretch", height=300,
            )
    else:
        st.info("Select at least one dataset to display.")


def render_knowledge_hub():
    st.title("AEMI/AIDR Knowledge Hub — Disasters")
    source_box(**DATASET_SOURCES["Knowledge Hub"])

    df = load_knowledge_hub()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        categories = sorted(df["Category"].dropna().unique())
        sel_cats = st.multiselect("Category", categories, placeholder="All categories",
                                  key="kh_cats")
    with col2:
        sel_years = year_slider(df["_num_year"].dropna(), key="kh_years")
    with col3:
        all_states = sorted({s for states in df["_states"] for s in states if s != "Unknown"})
        sel_states = st.multiselect("State / Zone", all_states, placeholder="All zones",
                                    key="kh_states")
    with col4:
        sel_source = st.multiselect("Source", ["Both", "AIDR only", "AGD only"],
                                    placeholder="All sources", key="kh_source")

    mask = (df["_num_year"] >= sel_years[0]) & (df["_num_year"] <= sel_years[1])
    if sel_cats:
        mask &= df["Category"].isin(sel_cats)
    if sel_states:
        mask &= df["_states"].apply(lambda s: bool(set(s) & set(sel_states)))
    if sel_source:
        mask &= df["_source_flag"].isin(sel_source)
    filt = df[mask].copy()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Events", f"{len(filt):,}")
    m2.metric("Total Fatalities", f"{filt['_num_fatalities'].sum():,.0f}")
    m3.metric("Total Injured", f"{filt['_num_injured'].sum():,.0f}")
    n_both = (filt["_source_flag"] == "Both").sum()
    m4.metric("In both sources", f"{n_both:,} of {len(filt):,}")

    tab_map, tab_chart, tab_table = st.tabs(["Map", "Charts", "Table"])

    with tab_map:
        map_df = filt.dropna(subset=["lat", "lon"]).copy()
        map_df = map_df[
            map_df["lat"].between(-90, 90) &
            map_df["lon"].between(-180, 180) &
            ~((map_df["lat"] == 0) & (map_df["lon"] == 0))
        ]
        if not map_df.empty:
            map_df["_start_str"] = map_df["Start Date"].dt.strftime("%d/%m/%Y")
            fig = px.scatter_map(
                map_df,
                lat="lat", lon="lon",
                hover_name="Event",
                hover_data={"_start_str": True, "_num_fatalities": True,
                            "Category": True, "lat": False, "lon": False},
                color="Category",
                labels={"_start_str": "Start Date", "_num_fatalities": "Fatalities"},
                zoom=2, height=560,
                map_style="open-street-map",
            )
            fig.update_layout(margin={"r": 0, "t": 0, "l": 0, "b": 0})
            st.plotly_chart(fig, width="stretch")
            st.caption(f"{len(map_df):,} of {len(filt):,} events have coordinates (from AGD).")
        else:
            st.info("No events with coordinates for current filters.")

    with tab_chart:
        c1, c2 = st.columns(2)
        with c1:
            by_cat = filt["Category"].value_counts().reset_index()
            by_cat.columns = ["Category", "Count"]
            fig = px.bar(by_cat, x="Category", y="Count", title="Events by Category",
                         color="Category", color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_layout(showlegend=False, xaxis_tickangle=-35)
            st.plotly_chart(fig, width="stretch")
        with c2:
            fat_yr = filt.groupby("_num_year")["_num_fatalities"].sum().reset_index()
            fat_yr.columns = ["Year", "Fatalities"]
            fig = px.bar(fat_yr, x="Year", y="Fatalities", title="Fatalities per Year",
                         color_discrete_sequence=["#d62728"])
            st.plotly_chart(fig, width="stretch")

        events_yr = filt.groupby("_num_year").size().reset_index(name="Events")
        fig = px.area(events_yr, x="_num_year", y="Events",
                      title="Disaster Events per Year",
                      labels={"_num_year": "Year"},
                      color_discrete_sequence=["#1f77b4"])
        st.plotly_chart(fig, width="stretch")

        st.subheader("Top 15 Deadliest Events")
        top_dead = filt.nlargest(15, "_num_fatalities")[
            ["Event", "Category", "Zone", "Start Date", "Fatalities", "Injured",
             "Insured Cost", "_source_flag"]
        ].copy()
        st.dataframe(fmt_dates(top_dead).reset_index(drop=True), width="stretch")

    with tab_table:
        display_cols = [
            "Event", "Category", "Zone", "Region",
            "Start Date", "End Date",
            "Fatalities", "Injured", "Insured Cost",
            "Description", "URL", "Source(s)",
            "_source_flag", "lat", "lon",
        ]
        display_cols = [c for c in display_cols if c in filt.columns]
        st.dataframe(
            fmt_dates(filt[display_cols]).reset_index(drop=True),
            column_config={
                "URL": st.column_config.LinkColumn("URL", display_text="Open ↗"),
            },
            width="stretch", height=480,
        )
        download_button(filt[display_cols], "Knowledge Hub", "knowledge_hub_filtered.csv")


def render_drfa_activations():
    st.title("DRFA Activation History by Location")
    source_box(**DATASET_SOURCES["DRFA Activations"])

    df = load_drfa_activations()

    col1, col2, col3 = st.columns(3)
    with col1:
        states = sorted(df["STATE"].dropna().unique())
        sel_states = st.multiselect("State", states, placeholder="All states")
    with col2:
        hazards = sorted(df["hazard_type"].dropna().unique())
        sel_hazards = st.multiselect("Hazard type", hazards, placeholder="All hazards")
    with col3:
        sel_years = year_slider(df["_num_year"].dropna(), key="drfa_act_years")

    mask = (df["_num_year"] >= sel_years[0]) & (df["_num_year"] <= sel_years[1])
    if sel_states:
        mask &= df["STATE"].isin(sel_states)
    if sel_hazards:
        mask &= df["hazard_type"].isin(sel_hazards)
    filt = df[mask].copy()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Activation records", f"{len(filt):,}")
    m2.metric("Unique events", f"{filt['event_name'].nunique():,}")
    m3.metric("Unique LGAs", f"{filt['Location_Name'].nunique():,}")
    m4.metric("States covered", f"{filt['STATE'].nunique()}")

    tab_chart, tab_table = st.tabs(["Charts", "Table"])

    with tab_chart:
        c1, c2 = st.columns(2)
        with c1:
            by_hazard = filt["hazard_type"].value_counts().reset_index()
            by_hazard.columns = ["Hazard", "Activations"]
            by_hazard = by_hazard.sort_values("Activations", ascending=True)
            fig = px.bar(
                by_hazard, x="Activations", y="Hazard", orientation="h",
                title="Activations by Hazard Type",
                color="Activations", color_continuous_scale="Blues",
            )
            fig.update_layout(
                coloraxis_showscale=False,
                yaxis_title=None,
                height=max(320, len(by_hazard) * 28),
            )
            st.plotly_chart(fig, width="stretch")
        with c2:
            by_state = filt["STATE"].value_counts().reset_index()
            by_state.columns = ["State", "Activations"]
            fig = px.bar(by_state, x="State", y="Activations", title="Activations by State",
                         color="State", color_discrete_sequence=px.colors.qualitative.Pastel)
            fig.update_layout(showlegend=False)
            st.plotly_chart(fig, width="stretch")

        by_year = filt.groupby("_num_year").size().reset_index(name="Activations")
        fig = px.bar(by_year, x="_num_year", y="Activations", title="Activations per Year",
                     labels={"_num_year": "Year"},
                     color_discrete_sequence=["#1f77b4"])
        st.plotly_chart(fig, width="stretch")

        cat_cols = ["cat_A", "cat_B", "cat_C", "cat_D"]
        present = [c for c in cat_cols if c in filt.columns]
        if present:
            cat_sums = filt[present].sum().reset_index()
            cat_sums.columns = ["Category", "Activations"]
            fig = px.bar(cat_sums, x="Category", y="Activations",
                         title="DRFA Category Breakdown (LGA-event pairs with each category)",
                         color="Category",
                         color_discrete_sequence=px.colors.qualitative.Set1)
            fig.update_layout(showlegend=False)
            st.plotly_chart(fig, width="stretch")

    with tab_table:
        st.dataframe(fmt_dates(filt[raw_cols(filt)]).reset_index(drop=True),
                     width="stretch", height=420)
        download_button(filt, "DRFA Activations", "drfa_activations_filtered.csv")


def render_drfa_payments():
    st.title("DRFA Disaster Payment History")
    source_box(**DATASET_SOURCES["DRFA Payments"])

    df = load_drfa_payments()

    # ── Methodological warnings ───────────────────────────────────────────────
    if "_state_infer_conf" in df.columns:
        n_multi   = (df["_state_infer_conf"] == "MULTI").sum()
        n_unknown = (df["_state_infer_conf"] == "UNKNOWN").sum()
        n_exact   = (df["_state_infer_conf"] == "EXACT").sum()
        if n_multi > 0 or n_unknown > 0:
            with st.expander("⚠️ State inference uncertainty — click to expand", expanded=False):
                st.markdown(
                    f"**{n_exact + n_multi + n_unknown}** rows in this dataset had 'Unknown' state labels "
                    f"in the source CSV and were assigned a state by regex matching on the disaster name.\n\n"
                    f"| Confidence class | Rows |\n|---|---|\n"
                    f"| EXACT (single state match) | {n_exact} |\n"
                    f"| MULTI (matched ≥2 states — first used) | {n_multi} |\n"
                    f"| UNKNOWN (no pattern matched) | {n_unknown} |\n\n"
                    "MULTI and UNKNOWN rows may be misclassified. Run "
                    "`python -m src.validation.state_inference_audit` for a full audit report."
                )

    # Sentinel value note for claim counts
    if "_num_eligible" in df.columns:
        from src.methods.sentinel_strategy import count_sentinels
        n_sentinel_eligible = count_sentinels(df.get("Eligible Claims (No.)", pd.Series(dtype=str)))
        if n_sentinel_eligible > 0:
            st.caption(
                f"ℹ️ **Sentinel values:** {n_sentinel_eligible} cells in 'Eligible Claims' contain "
                f"'<20' (privacy-suppressed counts, substituted with 10 as midpoint). "
                "Claim count totals are approximate. See `src/methods/sentinel_strategy.py` "
                "for lower/upper bound alternatives."
            )

    col1, col2, col3 = st.columns(3)
    with col1:
        states = sorted(df["State Name"].dropna().unique())
        sel_states = st.multiselect("State", states, placeholder="All states")
    with col2:
        pay_types = sorted(df["Payment Type Name"].dropna().unique())
        sel_pay = st.multiselect("Payment type", pay_types, placeholder="All types")
    with col3:
        search = st.text_input("Search disaster name")

    mask = pd.Series(True, index=df.index)
    if sel_states:
        mask &= df["State Name"].isin(sel_states)
    if sel_pay:
        mask &= df["Payment Type Name"].isin(sel_pay)
    if search:
        mask &= df["Disaster Name"].str.contains(search, case=False, na=False)
    filt = df[mask].copy()

    m1, m2, m3 = st.columns(3)
    m1.metric("Records", f"{len(filt):,}")
    m2.metric("Total Paid", f"${filt['_num_paid'].sum():,.0f}")
    m3.metric("Total Granted", f"${filt['_num_granted'].sum():,.0f}")
    st.caption(
        "⚠️ **Selection bias — payments data covers only a subset of DRFA activations.** "
        "Only events where individual or household claims were submitted and processed appear here. "
        "Many DRFA-activated events (particularly infrastructure/public asset payments under DRA/DRFA) "
        "are not captured in this dataset. Absence of a payment record does not mean no Commonwealth "
        "support was provided."
    )

    tab_chart, tab_table = st.tabs(["Charts", "Table"])

    with tab_chart:
        c1, c2 = st.columns(2)
        with c1:
            by_pay_type = filt.groupby("Payment Type Name")["_num_paid"].sum().reset_index()
            by_pay_type.columns = ["Payment Type", "Dollars Paid"]
            by_pay_type = by_pay_type.nlargest(15, "Dollars Paid").sort_values("Dollars Paid", ascending=True)
            fig = px.bar(
                by_pay_type, x="Dollars Paid", y="Payment Type", orientation="h",
                title="$ Paid by Payment Type (top 15)",
                color="Dollars Paid", color_continuous_scale="Greens",
            )
            fig.update_layout(
                coloraxis_showscale=False,
                yaxis_title=None,
                xaxis_tickprefix="$",
                xaxis_tickformat=",.0f",
                height=max(320, len(by_pay_type) * 32),
            )
            st.plotly_chart(fig, width="stretch")
        with c2:
            by_state = filt.groupby("State Name")["_num_paid"].sum().reset_index()
            by_state.columns = ["State", "Dollars Paid"]
            by_state = by_state.sort_values("Dollars Paid", ascending=True)
            fig = px.bar(by_state, x="Dollars Paid", y="State", orientation="h",
                         title="$ Paid by State", color_discrete_sequence=["#2ca02c"])
            st.plotly_chart(fig, width="stretch")

        top_disasters = (
            filt.groupby("Disaster Name")["_num_paid"].sum()
            .nlargest(15).reset_index()
        )
        top_disasters.columns = ["Disaster", "Dollars Paid"]
        fig = px.bar(top_disasters, x="Dollars Paid", y="Disaster", orientation="h",
                     title="Top 15 Disasters by $ Paid",
                     color="Dollars Paid", color_continuous_scale="Blues")
        fig.update_layout(yaxis=dict(autorange="reversed"))
        st.plotly_chart(fig, width="stretch")

    with tab_table:
        st.dataframe(fmt_dates(filt[raw_cols(filt)]).reset_index(drop=True),
                     width="stretch", height=420)
        download_button(filt, "DRFA Payments", "drfa_payments_filtered.csv")


def render_emdat():
    st.title("EM-DAT: Australian Disasters")
    source_box(**DATASET_SOURCES["EMDAT"])

    df = load_emdat()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        dis_types = sorted(df["Disaster Type"].dropna().unique())
        sel_types = st.multiselect("Disaster type", dis_types, placeholder="All types")
    with col2:
        yr_min = int(df["Start Year"].min())
        yr_max = int(df["Start Year"].max())
        sel_years = st.slider("Year range", yr_min, yr_max, (yr_min, yr_max), key="emdat_years")
    with col3:
        subtypes = (
            sorted(df["Disaster Subtype"].dropna().unique())
            if "Disaster Subtype" in df.columns else []
        )
        sel_subtypes = st.multiselect("Disaster subtype", subtypes, placeholder="All subtypes")
    with col4:
        compound_only = st.checkbox("Compound events only", value=False, key="emdat_compound_only")

    mask = (df["Start Year"] >= sel_years[0]) & (df["Start Year"] <= sel_years[1])
    if sel_types:
        mask &= df["Disaster Type"].isin(sel_types)
    if sel_subtypes:
        mask &= df["Disaster Subtype"].isin(sel_subtypes)
    if compound_only:
        mask &= df["_is_compound"]
    filt = df[mask].copy()

    n_compound = filt["_is_compound"].sum()
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Events", f"{len(filt):,}")
    m2.metric("Compound events", f"{n_compound:,}",
              help="Events where CRED recorded one or more co-occurring hazard types in 'Associated Types'")
    m3.metric("Total Deaths", f"{filt['Total Deaths'].sum():,.0f}")
    m4.metric("Total Affected", f"{filt['Total Affected'].sum():,.0f}")
    total_dmg = filt["Total Damage, Adjusted ('000 US$)"].sum()
    m5.metric("Total Damage (adj. '000 USD)", f"${total_dmg:,.0f}")

    tab_chart, tab_compound, tab_table = st.tabs(["📊 Charts", "🔗 Compound Events", "📋 Table"])

    with tab_chart:
        c1, c2 = st.columns(2)
        with c1:
            by_type = filt["Disaster Type"].value_counts().reset_index()
            by_type.columns = ["Type", "Count"]
            fig = px.bar(by_type, x="Type", y="Count", title="Events by Disaster Type",
                         color="Type", color_discrete_sequence=px.colors.qualitative.Set1)
            fig.update_layout(showlegend=False, xaxis_tickangle=-30)
            st.plotly_chart(fig, width="stretch")
        with c2:
            deaths_by_year = filt.groupby("Start Year")["Total Deaths"].sum().reset_index()
            fig = px.bar(deaths_by_year, x="Start Year", y="Total Deaths",
                         title="Deaths per Year", color_discrete_sequence=["#d62728"])
            st.plotly_chart(fig, width="stretch")

        dmg_by_year = filt.groupby("Start Year")["Total Damage, Adjusted ('000 US$)"].sum().reset_index()
        fig = px.area(dmg_by_year, x="Start Year", y="Total Damage, Adjusted ('000 US$)",
                      title="Total Damage per Year (Adjusted '000 USD)",
                      color_discrete_sequence=["#ff7f0e"])
        st.plotly_chart(fig, width="stretch")

        top_deaths = filt.nlargest(15, "Total Deaths")[
            ["DisNo.", "Disaster Type", "Disaster Subtype", "Location", "Start Year", "Total Deaths"]
        ]
        st.subheader("Top 15 Deadliest Events")
        st.dataframe(
            top_deaths.reset_index(drop=True),
            column_config={
                "Total Deaths": st.column_config.ProgressColumn(
                    "Total Deaths",
                    format="%d",
                    min_value=0,
                    max_value=int(top_deaths["Total Deaths"].max()),
                ),
            },
            width="stretch",
        )

    with tab_compound:
        st.info(
            "EM-DAT records an **Associated Types** field when CRED determined that a second hazard "
            "co-occurred with the primary event (e.g., a cyclone that also caused flooding, or a "
            "wildfire preceded by a heat wave). This is CRED's own contemporaneous classification — "
            "not a derived proximity measure — making it a scientifically defensible source for "
            "compound hazard identification. "
            "Source: Cuthbertson et al. (2021), *Prehospital and Disaster Medicine*.",
            icon="📚",
        )

        compound_df = filt[filt["_is_compound"]].copy()
        n_tot = len(filt)
        n_comp = len(compound_df)

        ca, cb, cc = st.columns(3)
        ca.metric("Compound events (filtered)", f"{n_comp:,}")
        cb.metric("As % of filtered events", f"{100 * n_comp / n_tot:.1f}%" if n_tot else "—")
        cc.metric("Unique associated type combinations",
                  f"{compound_df['Associated Types'].nunique():,}")

        if compound_df.empty:
            st.info("No compound events in the current filter selection.")
        else:
            # ── Associated type frequency ──────────────────────────────────
            st.subheader("Associated hazard types")
            all_assoc: list[str] = []
            for val in compound_df["Associated Types"].dropna():
                all_assoc.extend([t.strip() for t in str(val).split("|")])
            assoc_counts = (
                pd.Series(all_assoc)
                .value_counts()
                .reset_index()
            )
            assoc_counts.columns = ["Associated Type", "Count"]
            fig_assoc = px.bar(
                assoc_counts, x="Associated Type", y="Count",
                title="Frequency of Associated Hazard Types across Compound Events",
                color="Associated Type",
                color_discrete_sequence=px.colors.qualitative.Set2,
            )
            fig_assoc.update_layout(showlegend=False, xaxis_tickangle=-30)
            st.plotly_chart(fig_assoc, width="stretch")

            # ── Primary × associated heatmap ───────────────────────────────
            st.subheader("Primary disaster type × associated type")
            rows = []
            for _, row in compound_df.iterrows():
                for assoc in str(row["Associated Types"]).split("|"):
                    rows.append({"Primary": row["Disaster Type"], "Associated": assoc.strip()})
            cross_df = pd.DataFrame(rows)
            pivot = cross_df.groupby(["Primary", "Associated"]).size().reset_index(name="Count")
            fig_heat = px.density_heatmap(
                pivot, x="Primary", y="Associated", z="Count",
                title="Co-occurrence of Primary and Associated Disaster Types",
                color_continuous_scale="Oranges",
                text_auto=True,
            )
            fig_heat.update_layout(xaxis_tickangle=-30)
            st.plotly_chart(fig_heat, width="stretch")

            # ── Event list ────────────────────────────────────────────────
            st.subheader("Compound event records")
            display_cols = [
                c for c in [
                    "DisNo.", "Event Name", "Disaster Type", "Disaster Subgroup",
                    "Associated Types", "Location", "Start Year",
                    "Total Deaths", "Total Affected", "Total Damage, Adjusted ('000 US$)",
                ] if c in compound_df.columns
            ]
            st.dataframe(
                compound_df[display_cols].sort_values("Start Year", ascending=False).reset_index(drop=True),
                width="stretch", height=420,
            )
            download_button(compound_df, "EMDAT Compound Events", "emdat_compound_events.csv")

    with tab_table:
        st.dataframe(
            fmt_dates(filt[raw_cols(filt)]).reset_index(drop=True),
            column_config={
                "Total Deaths":   st.column_config.NumberColumn("Total Deaths",   format="%d"),
                "No. Injured":    st.column_config.NumberColumn("No. Injured",    format="%d"),
                "Total Affected": st.column_config.NumberColumn("Total Affected", format="%d"),
                "Total Damage, Adjusted ('000 US$)": st.column_config.NumberColumn(
                    "Damage Adj. ('000 US$)", format="$%,.0f"
                ),
            },
            width="stretch", height=420,
        )
        download_button(filt, "EMDAT", "emdat_filtered.csv")


def render_ica():
    st.title("ICA Historical Normalised Catastrophes")
    source_box(**DATASET_SOURCES["ICA Catastrophes"])

    df = load_ica()

    # Sentinel warning: ICA claim counts also use <20 privacy suppression
    try:
        from src.methods.sentinel_strategy import count_sentinels
        _raw_ica = __import__("pandas").read_csv(
            DATA_DIR / "ICA-Historical-Normalised-Catastrophe-Master-Updated-2026_02.csv",
            usecols=["TOTAL CLAIMS RECEIVED"],
            low_memory=False,
        )
        n_ica_sentinel = count_sentinels(_raw_ica["TOTAL CLAIMS RECEIVED"])
        if n_ica_sentinel > 0:
            st.caption(
                f"ℹ️ **Sentinel values:** {n_ica_sentinel} ICA events have '<20' claim counts "
                "(privacy-suppressed, substituted with 10). Claim count totals are approximate."
            )
    except Exception:
        pass

    col1, col2, col3 = st.columns(3)
    with col1:
        cat_types = sorted(df["Type"].dropna().unique()) if "Type" in df.columns else []
        sel_types = st.multiselect("Event type", cat_types, placeholder="All types")
    with col2:
        states_raw = df["State"].dropna().unique()
        all_states = sorted(set(
            s.strip()
            for entry in states_raw
            for s in str(entry).split(",")
        ))
        sel_states = st.multiselect("State", all_states, placeholder="All states")
    with col3:
        yr_min = int(df["Year"].min())
        yr_max = int(df["Year"].max())
        sel_years = st.slider("Year range", yr_min, yr_max, (yr_min, yr_max), key="ica_years")

    mask = (df["Year"] >= sel_years[0]) & (df["Year"] <= sel_years[1])
    if sel_types:
        mask &= df["Type"].isin(sel_types)
    if sel_states:
        mask &= df["State"].apply(
            lambda x: any(s in str(x) for s in sel_states) if pd.notna(x) else False
        )
    filt = df[mask].copy()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Events", f"{len(filt):,}")
    m2.metric("Total Claims", f"{filt['_num_total_claims'].sum():,.0f}")
    orig = filt["_num_orig_loss"].sum()
    norm = filt["_num_norm_loss"].sum()
    m3.metric("Original Loss", f"${orig/1e9:.2f}B")
    m4.metric("Normalised Loss (2022)", f"${norm/1e9:.2f}B")
    st.caption(
        "⚠️ **Insured losses only.** ICA captures events that triggered significant insurance claims. "
        "Events with low insurance penetration (remote/regional, low-income communities, agricultural losses) "
        "are systematically under-represented or absent. Loss figures reflect private insurance industry "
        "exposure, not total economic or social impact."
    )

    tab_chart, tab_simul, tab_table = st.tabs(["Charts", "Simultaneity", "Table"])

    with tab_chart:
        c1, c2 = st.columns(2)
        with c1:
            by_type = filt.groupby("Type")["_num_norm_loss"].sum().reset_index()
            by_type.columns = ["Type", "Normalised Loss"]
            fig = px.pie(by_type, names="Type", values="Normalised Loss",
                         title="Normalised Loss by Event Type",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            st.plotly_chart(fig, width="stretch")
        with c2:
            by_year = filt.groupby("Year")["_num_norm_loss"].sum().reset_index()
            by_year.columns = ["Year", "Normalised Loss"]
            fig = px.bar(by_year, x="Year", y="Normalised Loss",
                         title="Normalised Loss per Year (2022 AUD)",
                         color_discrete_sequence=["#9467bd"])
            st.plotly_chart(fig, width="stretch")

        top_events = filt.nlargest(15, "_num_norm_loss")[
            ["CAT Name", "Event Name", "Type", "State", "Year",
             "ORIGINAL LOSS VALUE", "NORMALISED LOSS VALUE (2022)", "TOTAL CLAIMS RECEIVED"]
        ].copy()
        st.subheader("Top 15 Events by Normalised Loss (2022 AUD)")
        st.dataframe(
            top_events.reset_index(drop=True),
            column_config={
                "NORMALISED LOSS VALUE (2022)": st.column_config.TextColumn("Norm. Loss (2022 AUD)"),
                "ORIGINAL LOSS VALUE":           st.column_config.TextColumn("Orig. Loss"),
                "TOTAL CLAIMS RECEIVED":         st.column_config.TextColumn("Claims"),
            },
            width="stretch",
        )

        claims_by_year = filt.groupby("Year")["_num_total_claims"].sum().reset_index()
        claims_by_year.columns = ["Year", "Total Claims"]
        fig = px.line(claims_by_year, x="Year", y="Total Claims",
                      title="Total Claims per Year", markers=True,
                      color_discrete_sequence=["#e377c2"])
        st.plotly_chart(fig, width="stretch")

    with tab_simul:
        st.caption(
            "Two ICA events are 'concurrent' if their active periods overlap. "
            "Active period = Event Start → Event Finish (start + 30 days when no Finish date is recorded). "
            "Computed across all **744 ICA events** (1967–2026), independent of DRFA matching."
        )

        with st.spinner("Computing period overlap across all ICA events…"):
            ica_simul = compute_ica_simultaneity()

        # Apply the same user filters (year, type, state) to the annotated full set
        simul_filt = ica_simul[
            (ica_simul["Year"] >= sel_years[0]) & (ica_simul["Year"] <= sel_years[1])
        ].copy()
        if sel_types:
            simul_filt = simul_filt[simul_filt["Type"].isin(sel_types)]
        if sel_states:
            simul_filt = simul_filt[simul_filt["State"].apply(
                lambda x: any(s in str(x) for s in sel_states) if pd.notna(x) else False
            )]

        n_conc   = (simul_filt["concurrent_count"] > 0).sum()
        n_total  = len(simul_filt)
        worst_ev = simul_filt.nlargest(1, "concurrent_count")

        sm1, sm2, sm3 = st.columns(3)
        sm1.metric("Events with ≥1 concurrent disaster", f"{n_conc} of {n_total}")
        sm2.metric("Max concurrent at any one time", f"{simul_filt['concurrent_count'].max():.0f}")
        sm3.metric("Worst instance",
                   worst_ev["Event Name"].iloc[0] if len(worst_ev) else "—",
                   delta=f"{int(worst_ev['concurrent_count'].iloc[0])} concurrent" if len(worst_ev) else None,
                   delta_color="off")

        # Peak concurrent by year
        peak_yr = (
            simul_filt.groupby("Year")["concurrent_count"].max().reset_index()
        )
        peak_yr.columns = ["Year", "Max Concurrent Events"]
        fig = px.bar(
            peak_yr, x="Year", y="Max Concurrent Events",
            title="Peak Concurrent ICA Events Active per Year",
            color="Max Concurrent Events", color_continuous_scale="YlOrRd",
        )
        fig.update_layout(coloraxis_showscale=False)
        st.plotly_chart(fig, width="stretch")

        # Gantt of events with concurrent companions
        gantt_df = simul_filt[simul_filt["concurrent_count"] > 0].copy()
        gantt_df = gantt_df.sort_values("_p_start")
        if not gantt_df.empty:
            gantt_df["_label"] = (
                gantt_df["Event Name"].str[:40]
                + " (" + gantt_df["State"].fillna("?") + ")"
            )
            fig_g = px.timeline(
                gantt_df,
                x_start="_p_start", x_end="_p_end",
                y="_label",
                color="Type",
                hover_data={
                    "concurrent_count":  True,
                    "concurrent_states": True,
                    "concurrent_events": True,
                    "_p_start": False, "_p_end": False,
                },
                labels={
                    "_label":            "ICA Event",
                    "Type":              "Hazard Type",
                    "concurrent_count":  "# Concurrent",
                    "concurrent_states": "Concurrent states",
                    "concurrent_events": "Concurrent event names",
                },
                title="ICA Events with Concurrent Active Disasters",
                height=max(350, len(gantt_df) * 20),
            )
            fig_g.update_yaxes(autorange="reversed")
            fig_g.update_layout(margin={"l": 10, "r": 10},
                                legend=dict(orientation="h", y=-0.15))
            st.plotly_chart(fig_g, width="stretch")
        else:
            st.info("No concurrent events found for the current filter selection.")

        # Worst-instances table
        st.subheader("Most simultaneous instances")
        worst_tbl = (
            simul_filt[simul_filt["concurrent_count"] > 0]
            .sort_values("concurrent_count", ascending=False)
            [["Event Name", "Type", "State", "Event Start", "Event Finish",
              "concurrent_count", "concurrent_states", "concurrent_events"]]
            .rename(columns={
                "Event Name":        "ICA Event",
                "Event Start":       "Start",
                "Event Finish":      "Finish",
                "concurrent_count":  "# Concurrent",
                "concurrent_states": "Concurrent States",
                "concurrent_events": "Concurrent Events",
            })
        )
        st.dataframe(fmt_dates(worst_tbl).reset_index(drop=True),
                     width="stretch", height=400)
        csv_sim = worst_tbl.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download simultaneity table", data=csv_sim,
                           file_name="ica_simultaneous_disasters.csv", mime="text/csv")

    with tab_table:
        st.dataframe(fmt_dates(filt[raw_cols(filt)]).reset_index(drop=True),
                     width="stretch", height=420)
        download_button(filt, "ICA Catastrophes", "ica_catastrophes_filtered.csv")


def render_drfa_merged():
    st.title("DRFA Merged: Activations + Payments")
    source_box(**DATASET_SOURCES["DRFA Merged"])

    df = load_drfa_merged()

    col1, col2, col3 = st.columns(3)
    with col1:
        states = sorted(df["STATE"].dropna().unique())
        sel_states = st.multiselect("State", states, placeholder="All states")
    with col2:
        hazards = sorted(df["hazard_type"].dropna().unique())
        sel_hazards = st.multiselect("Hazard type", hazards, placeholder="All hazards")
    with col3:
        sel_years = year_slider(df["_num_year"].dropna(), key="drfa_merged_years")

    mask = (df["_num_year"] >= sel_years[0]) & (df["_num_year"] <= sel_years[1])
    if sel_states:
        mask &= df["STATE"].isin(sel_states)
    if sel_hazards:
        mask &= df["hazard_type"].isin(sel_hazards)
    filt = df[mask].copy()

    has_payment = filt["Has_Payment_Data"] == "Yes"
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total rows", f"{len(filt):,}")
    m2.metric("Rows with payment data", f"{has_payment.sum():,}")
    m3.metric("Total Paid (matched rows)", f"${filt.loc[has_payment, '_num_paid'].sum():,.0f}")
    m4.metric("Total Claims (matched rows)", f"{filt.loc[has_payment, '_num_total_received'].sum():,.0f}")

    tab_chart, tab_table = st.tabs(["Charts", "Table"])

    with tab_chart:
        c1, c2 = st.columns(2)
        with c1:
            hazard_grp = (
                filt["hazard_type"].str.split(",").str[0].str.strip()
                .map(_DRFA_HAZARD_SIMPLE).fillna("Other")
            )
            by_hazard = hazard_grp.value_counts().reset_index()
            by_hazard.columns = ["Hazard Group", "Activations"]
            by_hazard = by_hazard.sort_values("Activations", ascending=True)
            fig = px.bar(
                by_hazard, x="Activations", y="Hazard Group", orientation="h",
                title="Activations by Hazard Group",
                color="Activations", color_continuous_scale="Blues",
            )
            fig.update_layout(
                coloraxis_showscale=False,
                yaxis_title=None,
                height=max(300, len(by_hazard) * 36),
            )
            st.plotly_chart(fig, width="stretch")
            st.caption("Hazard group is derived from the primary (first-listed) hazard in the DRFA record. Flood/Cyclone = activations where the primary hazard spans both peril types.")
        with c2:
            paid_by_hazard = (
                filt[has_payment]
                .groupby("hazard_type")["_num_paid"].sum()
                .reset_index()
            )
            paid_by_hazard.columns = ["Hazard", "Total Paid"]
            paid_by_hazard = paid_by_hazard.sort_values("Total Paid", ascending=True)
            fig = px.bar(paid_by_hazard, x="Total Paid", y="Hazard", orientation="h",
                         title="Total $ Paid by Hazard Type",
                         color_discrete_sequence=["#2ca02c"])
            st.plotly_chart(fig, width="stretch")

        by_year = filt.groupby("_num_year").size().reset_index(name="Rows")
        fig = px.bar(by_year, x="_num_year", y="Rows",
                     title="Merged Rows per Year",
                     labels={"_num_year": "Year"},
                     color_discrete_sequence=["#1f77b4"])
        st.plotly_chart(fig, width="stretch")

        # ── Payment total per year time series ───────────────────────────────
        pay_rows = filt[has_payment].copy()
        if pay_rows.empty:
            st.info("No payment data in the current filter selection.")
        else:
            pay_by_year = (
                pay_rows.groupby("_num_year")
                .agg(
                    total_paid=("_num_paid", "sum"),
                    total_granted=("_num_granted", "sum"),
                    n_events=("agrn", "nunique"),
                )
                .reset_index()
            )
            pay_by_year["Total Paid (A$M)"]    = pay_by_year["total_paid"]    / 1e6
            pay_by_year["Total Granted (A$M)"] = pay_by_year["total_granted"] / 1e6

            fig_pay = go.Figure()
            fig_pay.add_trace(go.Bar(
                x=pay_by_year["_num_year"],
                y=pay_by_year["Total Granted (A$M)"],
                name="Granted",
                marker_color="lightsteelblue",
                opacity=0.85,
            ))
            fig_pay.add_trace(go.Bar(
                x=pay_by_year["_num_year"],
                y=pay_by_year["Total Paid (A$M)"],
                name="Paid",
                marker_color="#2ca02c",
                opacity=0.85,
            ))
            fig_pay.add_trace(go.Scatter(
                x=pay_by_year["_num_year"],
                y=pay_by_year["n_events"],
                name="Events with payment data",
                mode="lines+markers",
                yaxis="y2",
                line=dict(color="crimson", width=2),
                marker=dict(size=7),
            ))
            fig_pay.update_layout(
                title="DRFA Payment Totals per Year (disaster onset year)",
                xaxis_title="Year",
                yaxis=dict(title="A$ Million", tickprefix="$", ticksuffix="M"),
                yaxis2=dict(
                    title="Events with payment records",
                    overlaying="y",
                    side="right",
                    showgrid=False,
                ),
                barmode="overlay",
                legend=dict(orientation="h", y=-0.2),
                hovermode="x unified",
            )
            st.plotly_chart(fig_pay, width="stretch")
            st.caption(
                f"Only {pay_rows['agrn'].nunique()} of {filt['agrn'].nunique()} events in the current "
                "selection have published payment records (Services Australia). "
                "**Known data gaps:** The payments dataset covers AGRN 840+ (2019 onwards) — pre-2019 "
                "payments are not published here. **2020 has no records** despite 2020 AGRNs being in range "
                "— these are simply not published in this dataset. **2024–2026 records are sparse** as "
                "many activations are still being processed. Years with no bar = no published records, "
                "not zero payments."
            )

        top_paid = (
            filt[has_payment]
            .nlargest(15, "_num_paid")
            [["Location_Name", "STATE", "agrn", "event_name", "hazard_type",
              "Payment_Types", "_num_paid", "_num_total_received"]]
            .rename(columns={
                "Payment_Types":       "Payment Types",
                "_num_paid":           "Dollars Paid ($)",
                "_num_total_received": "Claims Received",
            })
            .copy()
        )
        if not top_paid.empty:
            st.subheader("Top 15 Rows by $ Paid")
            st.dataframe(top_paid.reset_index(drop=True), width="stretch")

    with tab_table:
        st.dataframe(fmt_dates(filt[raw_cols(filt)]).reset_index(drop=True),
                     width="stretch", height=420)
        download_button(filt, "DRFA Merged", "drfa_merged_filtered.csv")


@st.fragment
def _fragment_ica_compound():
    """Fragment: renders the FY-filtered results section of the ICA compound page."""
    from itertools import combinations as _combos
    ev_all       = st.session_state["_ica_ev_all"]
    cl_all       = st.session_state["_ica_cl_all"]
    fy_min       = st.session_state["_ica_fy_min"]
    fy_max       = st.session_state["_ica_fy_max"]
    thresh_lbl   = st.session_state["_ica_thresh_lbl"]

    sel_fy = st.slider("Financial year range", fy_min, fy_max, (fy_min, fy_max), key="cd_fy")

    ev = ev_all[(ev_all["_fy"] >= sel_fy[0]) & (ev_all["_fy"] <= sel_fy[1])].copy()
    cl = cl_all[(cl_all["fy"] >= sel_fy[0]) & (cl_all["fy"] <= sel_fy[1])].copy()

    n_ev       = len(ev)
    n_compound = ev["_is_compound"].sum()
    n_fy_total = cl["fy"].nunique()
    n_fy_comp  = cl[cl["_is_compound"]]["fy"].nunique()
    total_nl   = ev["_num_norm_loss"].sum()
    comp_nl    = ev[ev["_is_compound"]]["_num_norm_loss"].sum()
    worst_cl   = cl[cl["_is_compound"]].nlargest(1, "n_events")

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("ICA events (≥ threshold)", f"{n_ev:,}")
    m2.metric("In compound clusters",
              f"{n_compound:,}",
              delta=f"{100*n_compound/max(n_ev,1):.0f}% of events",
              delta_color="off")
    m3.metric("Compound financial years", f"{n_fy_comp} of {n_fy_total}")
    m4.metric("Largest cluster",
              f"{int(worst_cl['n_events'].iloc[0])} events" if len(worst_cl) else "—",
              delta=f"FY{int(worst_cl['fy'].iloc[0])}" if len(worst_cl) else None,
              delta_color="off")
    m5.metric("Compound share of NL",
              f"{100*comp_nl/max(total_nl,1):.0f}%",
              delta=f"A${comp_nl/1e9:.1f}B of A${total_nl/1e9:.1f}B",
              delta_color="off")

    t_annual, t_clusters, t_pairs, t_cdms, t_climate = st.tabs([
        "Annual Overview", "Compound Clusters", "Peril Pairs", "Magnitude Scale", "Climate Drivers",
    ])

    with t_annual:
        import numpy as np
        import math
        from plotly.subplots import make_subplots as _make_subplots

        # ── Colour palette (consistent with Gissing et al. 2022 Figure 1) ──────
        _PC = {
            "Tropical Cyclone": "#2ca02c",
            "Flood":            "#1f77b4",
            "Storm":            "#aec7e8",
            "Bushfire":         "#d62728",
            "Heatwave":         "#ff7f0e",
            "Landslide":        "#7f7f7f",
            "Earthquake":       "#8c564b",
            "Other":            "#c7c7c7",
        }
        _PERIL_ORDER = [
            "Tropical Cyclone", "Flood", "Storm",
            "Bushfire", "Heatwave", "Landslide", "Earthquake", "Other",
        ]

        # ── Data preparation ─────────────────────────────────────────────────
        fy_all = sorted(ev["_fy"].unique())
        fy_lo, fy_hi = fy_all[0], fy_all[-1]

        fy_peril_counts = ev.groupby(["_fy", "_peril"]).size().reset_index(name="n")

        fy_comp = (
            ev[ev["_is_compound"]].groupby("_fy").size()
            .reindex(fy_all, fill_value=0)
            .reset_index()
        )
        fy_comp.columns = ["_fy", "compound_n"]

        # ── OLS linear trend on compound count (numpy only) ──────────────────
        x_f = np.array(fy_comp["_fy"].values, dtype=float)
        y_f = np.array(fy_comp["compound_n"].values, dtype=float)
        n_f  = len(x_f)
        xm, ym = x_f.mean(), y_f.mean()
        SS_xx = float(np.sum((x_f - xm) ** 2))
        SS_xy = float(np.sum((x_f - xm) * (y_f - ym)))
        _slope     = SS_xy / SS_xx if SS_xx > 0 else 0.0
        _intercept = ym - _slope * xm
        y_trend    = _slope * x_f + _intercept
        SS_res = float(np.sum((y_f - y_trend) ** 2))
        if SS_xx > 0 and n_f > 2:
            se_s  = math.sqrt((SS_res / (n_f - 2)) / SS_xx) if SS_res > 0 else 0.0
            t_s   = _slope / se_s if se_s > 0 else 0.0
            df_t  = n_f - 2
            z_a   = abs(t_s) * (1 - 1 / (4 * df_t))
            p_val = round(max(0.01, min(0.99, 2 * (1 - 0.5 * (1 + math.erf(z_a / math.sqrt(2)))))), 2)
        else:
            p_val = 1.0

        # ── Context stressors (Australian historical record, FY-based) ───────
        # FY year = calendar year in which 1 July falls (e.g. FY2010 = Jul 2010–Jun 2011)
        _BANDS = [
            (1962, 1971, "Vietnam",            "War"),
            (1990, 1990, "Gulf War I",         "War"),
            (2001, 2012, "Afghanistan/Iraq",   "War"),
            (1968, 1969, "Hong Kong Flu", "Pandemic"),
            (2009, 2009, "Swine Flu",     "Pandemic"),
            (2020, 2022, "COVID-19",      "Pandemic"),
            (1974, 1974, "", "Recession"),
            (1982, 1982, "", "Recession"),
            (1990, 1990, "", "Recession"),
            (2020, 2020, "", "Recession"),
        ]
        _STRESSOR_Y     = {"War": 3, "Pandemic": 2, "Recession": 1}
        _STRESSOR_COLOR = {"War": "#8B0000", "Pandemic": "#CC6600", "Recession": "#555555"}
        _NEWCASTLE_FY   = 1989

        # ── Build figure (main chart + context strip) ─────────────────────────
        fig_a = _make_subplots(
            rows=2, cols=1,
            row_heights=[0.78, 0.22],
            shared_xaxes=True,
            vertical_spacing=0.03,
        )

        # Stacked bars: event count by peril
        for peril in _PERIL_ORDER:
            pdata = fy_peril_counts[fy_peril_counts["_peril"] == peril]
            if pdata.empty:
                continue
            fy_vals = pdata.set_index("_fy")["n"].reindex(fy_all, fill_value=0)
            fig_a.add_trace(go.Bar(
                x=fy_all,
                y=fy_vals.values,
                name=peril,
                marker_color=_PC.get(peril, "#aaaaaa"),
                legendgroup=peril,
                hovertemplate=f"<b>{peril}</b><br>FY%{{x}}: %{{y}} event(s)<extra></extra>",
            ), row=1, col=1)

        # Compound event count line
        fig_a.add_trace(go.Scatter(
            x=fy_comp["_fy"].tolist(),
            y=fy_comp["compound_n"].tolist(),
            name="Number of compound events",
            mode="lines+markers",
            line=dict(color="black", width=2),
            marker=dict(size=3, color="black"),
            hovertemplate="FY%{x}: %{y} compound event(s)<extra></extra>",
        ), row=1, col=1)

        # OLS trend line on compound count
        fig_a.add_trace(go.Scatter(
            x=fy_comp["_fy"].tolist(),
            y=[round(v, 3) for v in y_trend.tolist()],
            name=f"Trend in compound events (p = {p_val:.2f})",
            mode="lines",
            line=dict(color="black", width=1.5, dash="dot"),
            hoverinfo="skip",
        ), row=1, col=1)

        # Context stressor shapes (row 2)
        for fy_s, fy_e, label, cat in _BANDS:
            fy_s_c = max(fy_s, fy_lo)
            fy_e_c = min(fy_e, fy_hi)
            if fy_s_c > fy_hi or fy_e_c < fy_lo:
                continue
            y_pos = _STRESSOR_Y[cat]
            fig_a.add_shape(
                type="rect", xref="x", yref="y2",
                x0=fy_s_c - 0.45, x1=fy_e_c + 0.45,
                y0=y_pos - 0.38,  y1=y_pos + 0.38,
                fillcolor=_STRESSOR_COLOR[cat],
                line=dict(width=0), opacity=0.88,
            )
            if label and (fy_e_c - fy_s_c) >= 2:
                fig_a.add_annotation(
                    xref="x", yref="y2",
                    x=(fy_s_c + fy_e_c) / 2, y=y_pos,
                    text=label, showarrow=False,
                    font=dict(color="white", size=8, family="Arial"),
                )

        # Newcastle Earthquake (row 2)
        if fy_lo <= _NEWCASTLE_FY <= fy_hi:
            fig_a.add_shape(
                type="rect", xref="x", yref="y2",
                x0=_NEWCASTLE_FY - 0.45, x1=_NEWCASTLE_FY + 0.45,
                y0=0.12, y1=0.88,
                fillcolor="#222222", line=dict(width=0),
            )

        # Dummy legend entries for stressors (so they appear in the legend)
        for cat, color in _STRESSOR_COLOR.items():
            fig_a.add_trace(go.Scatter(
                x=[None], y=[None], mode="markers",
                marker=dict(color=color, size=10, symbol="square"),
                name=cat, legendgroup=f"ctx_{cat}",
            ), row=2, col=1)
        fig_a.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(color="#222222", size=10, symbol="square"),
            name="Earthquake", legendgroup="ctx_eq",
        ), row=2, col=1)

        # Layout
        n_comp_yrs = int((fy_comp["compound_n"] > 0).sum())
        fig_a.update_layout(
            barmode="stack",
            title=dict(
                text=(
                    f"ICA Compound Disaster Frequency — Events ≥ {thresh_lbl} NL "
                    f"· FY{fy_lo}–FY{fy_hi} · Gissing et al. (2022) methodology"
                ),
                font=dict(size=13),
            ),
            legend=dict(orientation="h", y=-0.12, x=0, xanchor="left", font=dict(size=11)),
            hovermode="x unified",
            height=590,
            plot_bgcolor="white",
            paper_bgcolor="white",
            margin=dict(t=60, b=110),
        )
        fig_a.update_yaxes(
            title_text="Number of ICA Events",
            row=1, col=1,
            gridcolor="#eeeeee", showline=True, linecolor="#cccccc",
            rangemode="nonnegative",
        )
        fig_a.update_yaxes(
            row=2, col=1,
            tickvals=[1, 2, 3],
            ticktext=["Recession", "Pandemic", "War"],
            range=[0, 3.8], showgrid=False, zeroline=False,
            title_text="Stressor", title_font=dict(size=10),
            tickfont=dict(size=9),
        )
        fig_a.update_xaxes(showgrid=False, row=1, col=1)
        fig_a.update_xaxes(showgrid=False, row=2, col=1, title_text="Financial Year (July start)")

        st.plotly_chart(fig_a, width="stretch")
        trend_sig = "no significant trend" if p_val >= 0.05 else "a statistically significant trend"
        thresh_lbl_esc = thresh_lbl.replace("$", r"\$")
        st.caption(
            f"**{n_comp_yrs} compound disaster year{'s' if n_comp_yrs != 1 else ''}** "
            f"at the {thresh_lbl_esc} threshold (FY{fy_lo}–FY{fy_hi}). "
            f"Trend: {trend_sig} (p = {p_val:.2f}); "
            "Gissing et al. (2022) reported p = 0.79 for FY1967–FY2020 at A\\$100M NL. "
            "Context strip: Australian war, pandemic, and recession periods; "
            "black marker = Newcastle earthquake (1989)."
        )

        st.divider()

        # ── Normalised Loss per FY ────────────────────────────────────────────
        fy_nl = (
            ev.groupby(["_fy", "_peril"])["_num_norm_loss"]
            .sum()
            .reset_index()
        )
        fy_nl["NL (A$B)"] = fy_nl["_num_norm_loss"] / 1e9

        compound_fy_set = set(fy_comp[fy_comp["compound_n"] > 0]["_fy"])
        fy_total_nl = fy_nl.groupby("_fy")["NL (A$B)"].sum()
        max_nl_val  = float(fy_total_nl.max()) if not fy_total_nl.empty else 1.0

        fig2 = px.bar(
            fy_nl, x="_fy", y="NL (A$B)", color="_peril",
            barmode="stack",
            color_discrete_map=_PC,
            category_orders={"_peril": _PERIL_ORDER},
            title=f"Normalised Insurance Loss per Financial Year — by Peril (2022 AUD, ≥ {thresh_lbl})",
            labels={"_fy": "Financial Year (start)", "_peril": "Peril", "NL (A$B)": "Normalised Loss (A$B)"},
        )
        for fy_c in sorted(compound_fy_set):
            if fy_lo <= fy_c <= fy_hi and fy_c in fy_total_nl.index:
                fig2.add_annotation(
                    x=fy_c,
                    y=float(fy_total_nl.loc[fy_c]) + max_nl_val * 0.025,
                    text="★", showarrow=False,
                    font=dict(size=9, color="#222222"),
                    yanchor="bottom",
                )
        fig2.update_layout(
            legend=dict(orientation="h", y=-0.22),
            plot_bgcolor="white",
            yaxis=dict(gridcolor="#eeeeee"),
            hovermode="x unified",
        )
        st.plotly_chart(fig2, width="stretch")
        st.caption("★ = compound disaster year (≥ 2 qualifying events within a 91-day chain-linked window).")

    with t_clusters:
        st.subheader("Compound disaster clusters (≥ 2 events)")
        compound_cl = cl[cl["_is_compound"]].copy()
        if compound_cl.empty:
            st.info("No compound clusters found at the selected threshold and year range.")
        else:
            compound_cl["NL (2022 AUD)"] = compound_cl["total_nl"].apply(
                lambda x: (f"${x/1e9:.2f}B" if x >= 1e9 else f"${x/1e6:.0f}M") if pd.notna(x) else ""
            )
            compound_cl["First Event"] = pd.to_datetime(compound_cl["cluster_start"]).dt.strftime("%d/%m/%Y")
            display = compound_cl.rename(columns={
                "fy":         "FY",
                "n_events":   "# Events",
                "_magnitude": "CDMS",
                "perils":     "Perils",
                "event_names":"Events (first 5)",
            })[["FY", "First Event", "# Events", "NL (2022 AUD)", "CDMS", "Perils", "Events (first 5)"]]
            st.dataframe(display.reset_index(drop=True), width="stretch", height=500)
            csv_cl = compound_cl.drop(columns=["_is_compound"], errors="ignore").to_csv(index=False).encode("utf-8")
            st.download_button("⬇ Download compound clusters CSV", data=csv_cl,
                               file_name="compound_clusters.csv", mime="text/csv")

    with t_pairs:
        st.subheader("Peril co-occurrence in compound clusters")
        compound_events = ev[ev["_is_compound"]].copy()
        if compound_events.empty:
            st.info("No compound events at the selected threshold and year range.")
        else:
            peril_pairs = []
            for _cid, grp in compound_events.groupby("_cluster_id"):
                perils = sorted(grp["_peril"].dropna().tolist())
                for a, b in _combos(perils, 2):
                    peril_pairs.append((min(a, b), max(a, b)))

            if not peril_pairs:
                st.info("No within-cluster peril pairs found.")
            else:
                pairs_df = pd.DataFrame(peril_pairs, columns=["Peril A", "Peril B"])
                pair_counts = (
                    pairs_df.groupby(["Peril A", "Peril B"])
                    .size()
                    .reset_index(name="Count")
                )
                all_perils = sorted(
                    set(pair_counts["Peril A"].unique()) | set(pair_counts["Peril B"].unique())
                )
                matrix = pd.DataFrame(0, index=all_perils, columns=all_perils, dtype=int)
                for _, row in pair_counts.iterrows():
                    matrix.at[row["Peril A"], row["Peril B"]] += row["Count"]
                    matrix.at[row["Peril B"], row["Peril A"]] += row["Count"]

                fig_hm = px.imshow(
                    matrix, text_auto=True, color_continuous_scale="YlOrRd",
                    title="Peril Co-occurrence in Compound Clusters (symmetric)",
                    labels={"color": "Co-occurrences"},
                )
                st.plotly_chart(fig_hm, width="stretch")

                pair_counts["Pair"] = pair_counts["Peril A"] + " + " + pair_counts["Peril B"]
                pair_counts = pair_counts.sort_values("Count", ascending=False)
                fig_bar = px.bar(
                    pair_counts, x="Pair", y="Count",
                    title="Most common peril pairs in compound disasters",
                    color="Count", color_continuous_scale="YlOrRd",
                )
                fig_bar.update_layout(xaxis_tickangle=-30, coloraxis_showscale=False)
                st.plotly_chart(fig_bar, width="stretch")
                st.dataframe(pair_counts[["Pair", "Count"]].reset_index(drop=True),
                             width="stretch", height=300)

    with t_cdms:
        st.subheader("Compound Disaster Magnitude Scale (Gissing et al. 2022, Table 3)")
        st.caption(
            "Each tier requires **every component event** to individually exceed the loss threshold. "
            "A cluster of two A\\$1.5B events rates CDMS III — not CDMS I — because each event exceeds A\\$1B."
        )
        st.markdown("""
| CDMS | # Component Events | Loss threshold per component event (NL) |
|:---:|:---:|:---|
| **I** | 2–3 | ≥ A\\$100M each |
| **II** | > 3 | ≥ A\\$100M each |
| **III** | 2 | ≥ A\\$1B each |
| **IV** | > 2 | ≥ A\\$1B each |
| **V** | 2 | ≥ A\\$5B each |
| **VI** | > 2 | ≥ A\\$5B each |
| **VII** | 2 | ≥ A\\$20B each |
| **VIII** | > 2 | ≥ A\\$20B each |
        """)

        compound_cl = cl[cl["_is_compound"] & cl["_magnitude"].notna()].copy()
        if compound_cl.empty:
            st.info("No compound clusters with CDMS rating in current selection.")
        else:
            _MAG_ORDER = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII"]
            mag_counts = (
                compound_cl["_magnitude"]
                .value_counts()
                .reindex(_MAG_ORDER, fill_value=0)
                .reset_index()
            )
            mag_counts.columns = ["CDMS", "Count"]
            fig_mag = px.bar(
                mag_counts, x="CDMS", y="Count",
                title="Compound Disasters by Magnitude Scale",
                color="Count", color_continuous_scale="YlOrRd",
                category_orders={"CDMS": _MAG_ORDER},
            )
            fig_mag.update_layout(coloraxis_showscale=False)
            st.plotly_chart(fig_mag, width="stretch")

            compound_cl["Total NL (A$B)"] = compound_cl["total_nl"] / 1e9
            fig_sc = px.scatter(
                compound_cl, x="n_events", y="Total NL (A$B)",
                color="_magnitude",
                category_orders={"_magnitude": _MAG_ORDER},
                hover_data={"fy": True, "perils": True, "event_names": True},
                labels={
                    "n_events":   "# Component Events",
                    "_magnitude": "CDMS",
                    "fy":         "Financial Year",
                    "perils":     "Perils",
                    "event_names":"Events",
                },
                title="Compound Clusters — Component Count vs Total NL",
                size="Total NL (A$B)", size_max=25,
            )
            st.plotly_chart(fig_sc, width="stretch")

    with t_climate:
        st.subheader("ENSO (ONI) and compound disaster seasons")
        try:
            oni = fetch_oni_data()
            oni["_fy"] = oni["date"].apply(lambda d: d.year if d.month >= 7 else d.year - 1)
            oni_fy = oni.groupby("_fy")["oni"].mean().reset_index()
            oni_fy.columns = ["fy", "mean_oni"]

            fy_compound = cl.groupby("fy")["_is_compound"].any().reset_index()
            fy_compound.columns = ["fy", "has_compound"]

            oni_merged = oni_fy.merge(fy_compound, on="fy", how="left")
            oni_merged["has_compound"] = oni_merged["has_compound"].astype("boolean").fillna(False).astype(bool)
            oni_merged["Season"] = oni_merged["has_compound"].map({
                True:  "Compound disaster season",
                False: "No compound disasters",
            })
            oni_sel = oni_merged[oni_merged["fy"].between(sel_fy[0], sel_fy[1])]

            _CMAP = {
                "Compound disaster season": "#d62728",
                "No compound disasters":    "#1f77b4",
            }
            fig_oni = px.bar(
                oni_sel, x="fy", y="mean_oni", color="Season",
                color_discrete_map=_CMAP,
                title="Mean ONI by Financial Year — compound vs non-compound seasons",
                labels={"fy": "Financial Year (start)", "mean_oni": "Mean ONI (Niño 3.4)"},
            )
            fig_oni.add_hline(y=0.5,  line_dash="dot", line_color="orange",    annotation_text="El Niño threshold (+0.5)")
            fig_oni.add_hline(y=-0.5, line_dash="dot", line_color="steelblue", annotation_text="La Niña threshold (−0.5)")
            fig_oni.update_layout(legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(fig_oni, width="stretch")

            fig_box = px.box(
                oni_sel, x="Season", y="mean_oni", color="Season",
                color_discrete_map=_CMAP, points="all",
                title="ONI distribution: compound vs non-compound financial years",
                labels={"mean_oni": "Mean ONI (Jul–Jun)"},
            )
            fig_box.add_hline(y=0.5,  line_dash="dot", line_color="orange")
            fig_box.add_hline(y=-0.5, line_dash="dot", line_color="steelblue")
            st.plotly_chart(fig_box, width="stretch")

            summary = (
                oni_sel.groupby("Season")["mean_oni"]
                .agg(n="count", mean="mean", median="median", std="std")
                .reset_index()
                .rename(columns={"n": "N years", "mean": "Mean ONI", "median": "Median ONI", "std": "Std ONI"})
            )
            for col in ("Mean ONI", "Median ONI", "Std ONI"):
                summary[col] = summary[col].round(3)
            st.dataframe(summary.reset_index(drop=True), width="stretch")
            st.caption(
                "Mean ONI computed over the 12 months of each Australian financial year (Jul–Jun). "
                "Positive ONI → El Niño (drought / fire risk); Negative → La Niña (flood / cyclone risk)."
            )
        except Exception as _e:
            st.warning(f"Could not load ONI data: {_e}")

        st.divider()
        st.subheader("SAM (Southern Annular Mode) and compound disaster seasons")
        st.caption(
            "Positive SAM = enhanced westerlies at 60°S → reduced southern Australian rainfall → elevated fire and drought risk.  \n"
            "Negative SAM = weakened westerlies at 60°S → enhanced mid-latitude rainfall → elevated flood and cyclone risk in southern Australia."
        )
        try:
            sam = fetch_sam_data()
            sam["_fy"] = sam["date"].apply(lambda d: d.year if d.month >= 7 else d.year - 1)
            sam_fy = sam.groupby("_fy")["sam"].mean().reset_index()
            sam_fy.columns = ["fy", "mean_sam"]

            sam_merged = sam_fy.merge(fy_compound, on="fy", how="left")
            sam_merged["has_compound"] = sam_merged["has_compound"].astype("boolean").fillna(False).astype(bool)
            sam_merged["Season"] = sam_merged["has_compound"].map({
                True:  "Compound disaster season",
                False: "No compound disasters",
            })
            sam_sel = sam_merged[sam_merged["fy"].between(sel_fy[0], sel_fy[1])]

            fig_sam = px.bar(
                sam_sel, x="fy", y="mean_sam", color="Season",
                color_discrete_map=_CMAP,
                title="Mean SAM by Financial Year — compound vs non-compound seasons",
                labels={"fy": "Financial Year (start)", "mean_sam": "Mean SAM (Marshall index)"},
            )
            fig_sam.add_hline(y= 1.0, line_dash="dot", line_color="firebrick",   annotation_text="Positive SAM threshold (+1.0)")
            fig_sam.add_hline(y=-1.0, line_dash="dot", line_color="steelblue",   annotation_text="Negative SAM threshold (−1.0)")
            fig_sam.add_hline(y= 0,   line_dash="dash", line_color="grey", line_width=1)
            fig_sam.update_layout(legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(fig_sam, width="stretch")

            fig_sam_box = px.box(
                sam_sel, x="Season", y="mean_sam", color="Season",
                color_discrete_map=_CMAP, points="all",
                title="SAM distribution: compound vs non-compound financial years",
                labels={"mean_sam": "Mean SAM (Jul–Jun)"},
            )
            fig_sam_box.add_hline(y= 1.0, line_dash="dot", line_color="firebrick")
            fig_sam_box.add_hline(y=-1.0, line_dash="dot", line_color="steelblue")
            st.plotly_chart(fig_sam_box, width="stretch")

            sam_summary = (
                sam_sel.groupby("Season")["mean_sam"]
                .agg(n="count", mean="mean", median="median", std="std")
                .reset_index()
                .rename(columns={"n": "N years", "mean": "Mean SAM", "median": "Median SAM", "std": "Std SAM"})
            )
            for col in ("Mean SAM", "Median SAM", "Std SAM"):
                sam_summary[col] = sam_summary[col].round(3)
            st.dataframe(sam_summary.reset_index(drop=True), width="stretch")
            st.caption(
                "Mean SAM computed over the 12 months of each Australian financial year (Jul–Jun). "
                "Source: Marshall (2003) SAM index, British Antarctic Survey. Baseline: 1971–2000."
            )
        except Exception as _e:
            st.warning(f"Could not load SAM data: {_e}")

        st.divider()
        st.subheader("IOD (Dipole Mode Index) and compound disaster seasons")
        st.caption(
            "Positive IOD (DMI ≥ +0.4 °C) = warm western / cool eastern Indian Ocean → drought and fire risk in southern/eastern Australia.  \n"
            "Negative IOD (DMI ≤ −0.4 °C) = cool western / warm eastern Indian Ocean → elevated flood and cyclone risk.  \n"
            "IOD is most active during austral winter–spring (June–November)."
        )
        try:
            iod = fetch_iod_data()
            iod["_fy"] = iod["date"].apply(lambda d: d.year if d.month >= 7 else d.year - 1)
            iod_fy = iod.groupby("_fy")["dmi"].mean().reset_index()
            iod_fy.columns = ["fy", "mean_dmi"]

            iod_merged = iod_fy.merge(fy_compound, on="fy", how="left")
            iod_merged["has_compound"] = iod_merged["has_compound"].astype("boolean").fillna(False).astype(bool)
            iod_merged["Season"] = iod_merged["has_compound"].map({
                True:  "Compound disaster season",
                False: "No compound disasters",
            })
            iod_sel = iod_merged[iod_merged["fy"].between(sel_fy[0], sel_fy[1])]

            fig_iod = px.bar(
                iod_sel, x="fy", y="mean_dmi", color="Season",
                color_discrete_map=_CMAP,
                title="Mean DMI by Financial Year — compound vs non-compound seasons",
                labels={"fy": "Financial Year (start)", "mean_dmi": "Mean DMI (°C)"},
            )
            fig_iod.add_hline(y= 0.4, line_dash="dot", line_color="firebrick",  annotation_text="pIOD threshold (+0.4 °C)")
            fig_iod.add_hline(y=-0.4, line_dash="dot", line_color="steelblue",  annotation_text="nIOD threshold (−0.4 °C)")
            fig_iod.add_hline(y= 0,   line_dash="dash", line_color="grey", line_width=1)
            fig_iod.update_layout(legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(fig_iod, width="stretch")

            fig_iod_box = px.box(
                iod_sel, x="Season", y="mean_dmi", color="Season",
                color_discrete_map=_CMAP, points="all",
                title="DMI distribution: compound vs non-compound financial years",
                labels={"mean_dmi": "Mean DMI (°C, Jul–Jun)"},
            )
            fig_iod_box.add_hline(y= 0.4, line_dash="dot", line_color="firebrick")
            fig_iod_box.add_hline(y=-0.4, line_dash="dot", line_color="steelblue")
            st.plotly_chart(fig_iod_box, width="stretch")

            iod_summary = (
                iod_sel.groupby("Season")["mean_dmi"]
                .agg(n="count", mean="mean", median="median", std="std")
                .reset_index()
                .rename(columns={"n": "N years", "mean": "Mean DMI", "median": "Median DMI", "std": "Std DMI"})
            )
            for col in ("Mean DMI", "Median DMI", "Std DMI"):
                iod_summary[col] = iod_summary[col].round(3)
            st.dataframe(iod_summary.reset_index(drop=True), width="stretch")
            st.caption(
                "Mean DMI computed over the 12 months of each Australian financial year (Jul–Jun). "
                "Source: NOAA PSL / HadISST1.1 (Saji & Yamagata 2003). Threshold: ±0.4 °C (BoM standard)."
            )
        except Exception as _e:
            st.warning(f"Could not load IOD data: {_e}")


@st.fragment
def _fragment_drfa_compound():
    """Fragment: renders the FY-filtered results section of the DRFA compound page."""
    from itertools import combinations as _combos
    ev_all = st.session_state["_drfa_ev_all"]
    cl_all = st.session_state["_drfa_cl_all"]
    fy_min = st.session_state["_drfa_fy_min"]
    fy_max = st.session_state["_drfa_fy_max"]

    sel_fy = st.slider("Financial year range", fy_min, fy_max, (fy_min, fy_max), key="drfa_cd_fy")

    ev = ev_all[(ev_all["_fy"] >= sel_fy[0]) & (ev_all["_fy"] <= sel_fy[1])].copy()
    cl = cl_all[(cl_all["fy"] >= sel_fy[0]) & (cl_all["fy"] <= sel_fy[1])].copy()

    n_ev        = len(ev)
    n_compound  = ev["_is_compound"].sum()
    n_fy_total  = cl["fy"].nunique()
    n_fy_comp   = cl[cl["_is_compound"]]["fy"].nunique()
    total_lgas  = ev["_lga_count"].sum()
    comp_lgas   = ev[ev["_is_compound"]]["_lga_count"].sum()
    worst_cl    = cl[cl["_is_compound"]].nlargest(1, "n_events")

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("DRFA events", f"{n_ev:,}")
    m2.metric("In compound clusters",
              f"{n_compound:,}",
              delta=f"{100*n_compound/max(n_ev,1):.0f}% of events",
              delta_color="off")
    m3.metric("Compound financial years", f"{n_fy_comp} of {n_fy_total}")
    m4.metric("Largest cluster",
              f"{int(worst_cl['n_events'].iloc[0])} events" if len(worst_cl) else "—",
              delta=f"FY{int(worst_cl['fy'].iloc[0])}" if len(worst_cl) else None,
              delta_color="off")
    m5.metric("Compound share of LGA activations",
              f"{100*comp_lgas/max(total_lgas,1):.0f}%",
              delta=f"{int(comp_lgas):,} of {int(total_lgas):,} LGA-events",
              delta_color="off")

    t_annual, t_clusters, t_multistate, t_dgms, t_climate, t_compare = st.tabs([
        "Annual Overview", "Compound Clusters", "Multi-State Demand",
        "Magnitude Scale", "Climate Drivers", "ICA Comparison",
    ])

    with t_annual:
        import numpy as np
        import math
        from plotly.subplots import make_subplots as _make_subplots

        # ── Colour palette (shared with ICA page for cross-page consistency) ──
        _DPC = {
            "Tropical Cyclone": "#2ca02c",
            "Flood":            "#1f77b4",
            "Storm":            "#aec7e8",
            "Bushfire":         "#d62728",
            "Heatwave":         "#ff7f0e",
            "Landslide":        "#7f7f7f",
            "Earthquake":       "#8c564b",
            "Other":            "#c7c7c7",
        }
        _DRFA_PERIL_ORDER = [
            "Tropical Cyclone", "Flood", "Storm",
            "Bushfire", "Heatwave", "Landslide", "Earthquake", "Other",
        ]

        # ── Data preparation ─────────────────────────────────────────────────
        fy_all = sorted(ev["_fy"].unique())
        fy_lo, fy_hi = fy_all[0], fy_all[-1]

        fy_peril_counts = ev.groupby(["_fy", "_peril"]).size().reset_index(name="n")

        fy_comp = (
            ev[ev["_is_compound"]].groupby("_fy").size()
            .reindex(fy_all, fill_value=0)
            .reset_index()
        )
        fy_comp.columns = ["_fy", "compound_n"]

        # ── OLS linear trend on compound count (numpy only) ──────────────────
        x_f = np.array(fy_comp["_fy"].values, dtype=float)
        y_f = np.array(fy_comp["compound_n"].values, dtype=float)
        n_f  = len(x_f)
        xm, ym = x_f.mean(), y_f.mean()
        SS_xx = float(np.sum((x_f - xm) ** 2))
        SS_xy = float(np.sum((x_f - xm) * (y_f - ym)))
        _slope     = SS_xy / SS_xx if SS_xx > 0 else 0.0
        _intercept = ym - _slope * xm
        y_trend    = _slope * x_f + _intercept
        SS_res = float(np.sum((y_f - y_trend) ** 2))
        if SS_xx > 0 and n_f > 2:
            se_s  = math.sqrt((SS_res / (n_f - 2)) / SS_xx) if SS_res > 0 else 0.0
            t_s   = _slope / se_s if se_s > 0 else 0.0
            df_t  = n_f - 2
            z_a   = abs(t_s) * (1 - 1 / (4 * df_t))
            p_val = round(max(0.01, min(0.99, 2 * (1 - 0.5 * (1 + math.erf(z_a / math.sqrt(2)))))), 2)
        else:
            p_val = 1.0

        # ── Context stressors (clipped to DRFA coverage from 2006) ──────────
        _BANDS_D = [
            (2001, 2012, "Afghanistan/Iraq", "War"),
            (2009, 2009, "Swine Flu",        "Pandemic"),
            (2020, 2022, "COVID-19",    "Pandemic"),
            (2020, 2020, "", "Recession"),
        ]
        _STRESSOR_Y_D     = {"War": 3, "Pandemic": 2, "Recession": 1}
        _STRESSOR_COLOR_D = {"War": "#8B0000", "Pandemic": "#CC6600", "Recession": "#555555"}

        # ── Build figure (main chart + context strip) ─────────────────────────
        fig_d = _make_subplots(
            rows=2, cols=1,
            row_heights=[0.78, 0.22],
            shared_xaxes=True,
            vertical_spacing=0.03,
        )

        # Stacked bars: DRFA event count by peril
        for peril in _DRFA_PERIL_ORDER:
            pdata = fy_peril_counts[fy_peril_counts["_peril"] == peril]
            if pdata.empty:
                continue
            fy_vals = pdata.set_index("_fy")["n"].reindex(fy_all, fill_value=0)
            fig_d.add_trace(go.Bar(
                x=fy_all,
                y=fy_vals.values,
                name=peril,
                marker_color=_DPC.get(peril, "#aaaaaa"),
                legendgroup=peril,
                hovertemplate=f"<b>{peril}</b><br>FY%{{x}}: %{{y}} DRFA event(s)<extra></extra>",
            ), row=1, col=1)

        # Compound event count line
        fig_d.add_trace(go.Scatter(
            x=fy_comp["_fy"].tolist(),
            y=fy_comp["compound_n"].tolist(),
            name="Number of compound events",
            mode="lines+markers",
            line=dict(color="black", width=2),
            marker=dict(size=3, color="black"),
            hovertemplate="FY%{x}: %{y} compound event(s)<extra></extra>",
        ), row=1, col=1)

        # OLS trend line
        fig_d.add_trace(go.Scatter(
            x=fy_comp["_fy"].tolist(),
            y=[round(v, 3) for v in y_trend.tolist()],
            name=f"Trend in compound events (p = {p_val:.2f})",
            mode="lines",
            line=dict(color="black", width=1.5, dash="dot"),
            hoverinfo="skip",
        ), row=1, col=1)

        # Context stressor shapes (row 2)
        for fy_s, fy_e, label, cat in _BANDS_D:
            fy_s_c = max(fy_s, fy_lo)
            fy_e_c = min(fy_e, fy_hi)
            if fy_s_c > fy_hi or fy_e_c < fy_lo:
                continue
            y_pos = _STRESSOR_Y_D[cat]
            fig_d.add_shape(
                type="rect", xref="x", yref="y2",
                x0=fy_s_c - 0.45, x1=fy_e_c + 0.45,
                y0=y_pos - 0.38,  y1=y_pos + 0.38,
                fillcolor=_STRESSOR_COLOR_D[cat],
                line=dict(width=0), opacity=0.88,
            )
            if label and (fy_e_c - fy_s_c) >= 1:
                fig_d.add_annotation(
                    xref="x", yref="y2",
                    x=(fy_s_c + fy_e_c) / 2, y=y_pos,
                    text=label, showarrow=False,
                    font=dict(color="white", size=8, family="Arial"),
                )

        # Dummy legend entries for stressors
        for cat, color in _STRESSOR_COLOR_D.items():
            fig_d.add_trace(go.Scatter(
                x=[None], y=[None], mode="markers",
                marker=dict(color=color, size=10, symbol="square"),
                name=cat, legendgroup=f"dctx_{cat}",
            ), row=2, col=1)

        # Layout
        n_comp_yrs_d = int((fy_comp["compound_n"] > 0).sum())
        fig_d.update_layout(
            barmode="stack",
            title=dict(
                text=(
                    f"DRFA Compound Activation Frequency — Gissing et al. (2022) methodology "
                    f"· FY{fy_lo}–FY{fy_hi}"
                ),
                font=dict(size=13),
            ),
            legend=dict(orientation="h", y=-0.12, x=0, xanchor="left", font=dict(size=11)),
            hovermode="x unified",
            height=590,
            plot_bgcolor="white",
            paper_bgcolor="white",
            margin=dict(t=60, b=110),
        )
        fig_d.update_yaxes(
            title_text="Number of DRFA Events",
            row=1, col=1,
            gridcolor="#eeeeee", showline=True, linecolor="#cccccc",
            rangemode="nonnegative",
        )
        fig_d.update_yaxes(
            row=2, col=1,
            tickvals=[1, 2, 3],
            ticktext=["Recession", "Pandemic", "War"],
            range=[0, 3.8], showgrid=False, zeroline=False,
            title_text="Stressor", title_font=dict(size=10),
            tickfont=dict(size=9),
        )
        fig_d.update_xaxes(showgrid=False, row=1, col=1)
        fig_d.update_xaxes(showgrid=False, row=2, col=1, title_text="Financial Year (July start)")

        st.plotly_chart(fig_d, width="stretch")
        trend_sig_d = "no significant trend" if p_val >= 0.05 else "a statistically significant trend"
        st.caption(
            f"**{n_comp_yrs_d} compound activation year{'s' if n_comp_yrs_d != 1 else ''}** identified "
            f"(FY{fy_lo}–FY{fy_hi}). Linear regression shows {trend_sig_d} in compound event frequency "
            f"(p = {p_val:.2f}). "
            "Note: DRFA activations reflect government funding decisions, not physical disaster severity. "
            "Compound years here indicate concurrent Commonwealth funding obligations across multiple declared events."
        )

        st.divider()

        # ── LGA Activations per FY ────────────────────────────────────────────
        fy_lga = (
            ev.groupby(["_fy", "_peril"])["_lga_count"]
            .sum()
            .reset_index()
        )
        fy_lga.columns = ["_fy", "_peril", "LGA Activations"]

        compound_fy_set_d = set(fy_comp[fy_comp["compound_n"] > 0]["_fy"])
        fy_total_lga = fy_lga.groupby("_fy")["LGA Activations"].sum()
        max_lga_val  = float(fy_total_lga.max()) if not fy_total_lga.empty else 1.0

        fig2 = px.bar(
            fy_lga, x="_fy", y="LGA Activations", color="_peril",
            barmode="stack",
            color_discrete_map=_DPC,
            category_orders={"_peril": _DRFA_PERIL_ORDER},
            title="Total LGA Activations per Financial Year — by Peril",
            labels={"_fy": "Financial Year (start)", "_peril": "Peril"},
        )
        for fy_c in sorted(compound_fy_set_d):
            if fy_lo <= fy_c <= fy_hi and fy_c in fy_total_lga.index:
                fig2.add_annotation(
                    x=fy_c,
                    y=float(fy_total_lga.loc[fy_c]) + max_lga_val * 0.025,
                    text="★", showarrow=False,
                    font=dict(size=9, color="#222222"),
                    yanchor="bottom",
                )
        fig2.update_layout(
            legend=dict(orientation="h", y=-0.22),
            plot_bgcolor="white",
            yaxis=dict(gridcolor="#eeeeee"),
            hovermode="x unified",
        )
        st.plotly_chart(fig2, width="stretch")
        st.caption("★ = compound activation year (≥ 2 DRFA events within a 91-day chain-linked window). LGA count = number of unique LGA-level activations.")

    with t_clusters:
        st.subheader("DRFA compound disaster clusters (≥ 2 events)")
        compound_cl = cl[cl["_is_compound"]].copy()
        if compound_cl.empty:
            st.info("No compound clusters found at the selected year range.")
        else:
            compound_cl["First Event"] = pd.to_datetime(compound_cl["cluster_start"]).dt.strftime("%d/%m/%Y")
            display = compound_cl.rename(columns={
                "fy":          "FY",
                "n_events":    "# Events",
                "total_lgas":  "Total LGAs",
                "_magnitude":  "DGMS",
                "perils":      "Perils",
                "states":      "States",
                "event_names": "Events (first 5)",
            })[["FY", "First Event", "# Events", "Total LGAs", "DGMS", "Perils", "States", "Events (first 5)"]]
            st.dataframe(display.reset_index(drop=True), width="stretch", height=500)
            csv_cl = compound_cl.drop(columns=["_is_compound"], errors="ignore").to_csv(index=False).encode("utf-8")
            st.download_button("⬇ Download compound clusters CSV", data=csv_cl,
                               file_name="drfa_compound_clusters.csv", mime="text/csv")

        st.divider()
        st.subheader("Peril co-occurrence in DRFA compound clusters")
        compound_events = ev[ev["_is_compound"]].copy()
        if not compound_events.empty:
            peril_pairs = []
            for _cid, grp in compound_events.groupby("_cluster_id"):
                perils = sorted(grp["_peril"].dropna().tolist())
                for a, b in _combos(perils, 2):
                    peril_pairs.append((min(a, b), max(a, b)))

            if peril_pairs:
                pairs_df = pd.DataFrame(peril_pairs, columns=["Peril A", "Peril B"])
                pair_counts = (
                    pairs_df.groupby(["Peril A", "Peril B"])
                    .size()
                    .reset_index(name="Count")
                )
                all_perils = sorted(
                    set(pair_counts["Peril A"].unique()) | set(pair_counts["Peril B"].unique())
                )
                matrix = pd.DataFrame(0, index=all_perils, columns=all_perils, dtype=int)
                for _, row in pair_counts.iterrows():
                    matrix.at[row["Peril A"], row["Peril B"]] += row["Count"]
                    matrix.at[row["Peril B"], row["Peril A"]] += row["Count"]
                fig_hm = px.imshow(
                    matrix, text_auto=True, color_continuous_scale="YlOrRd",
                    title="Peril Co-occurrence in DRFA Compound Clusters (symmetric)",
                    labels={"color": "Co-occurrences"},
                )
                st.plotly_chart(fig_hm, width="stretch")

    with t_multistate:
        st.subheader("Cross-jurisdictional simultaneous demand")
        st.caption(
            "For each compound cluster, which states were simultaneously responding to active DRFA events? "
            "This directly measures the simultaneity problem: multiple jurisdictions drawing on Commonwealth "
            "support capacity at the same time."
        )
        compound_cl = cl[cl["_is_compound"]].copy()
        if compound_cl.empty:
            st.info("No compound clusters in the selected year range.")
        else:
            compound_cl["_state_list"] = compound_cl["states"].apply(
                lambda s: sorted(s.split("; ")) if pd.notna(s) else []
            )
            compound_cl["_n_states"] = compound_cl["_state_list"].apply(len)
            compound_cl["_multi_state"] = compound_cl["_n_states"] > 1

            n_multi = compound_cl["_multi_state"].sum()
            n_total = len(compound_cl)
            st.metric(
                "Multi-jurisdictional compound clusters",
                f"{n_multi} of {n_total}",
                delta=f"{100*n_multi/max(n_total,1):.0f}% of compound clusters",
                delta_color="off",
            )

            state_dist = compound_cl["_n_states"].value_counts().sort_index().reset_index()
            state_dist.columns = ["States Simultaneously Active", "Compound Clusters"]
            fig_sd = px.bar(
                state_dist, x="States Simultaneously Active", y="Compound Clusters",
                title="Number of Jurisdictions Simultaneously Active in Compound Clusters",
                color="Compound Clusters", color_continuous_scale="YlOrRd",
            )
            fig_sd.update_layout(coloraxis_showscale=False)
            st.plotly_chart(fig_sd, width="stretch")

            state_pairs = []
            for _, row in compound_cl.iterrows():
                sl = row["_state_list"]
                for a, b in _combos(sl, 2):
                    state_pairs.append((min(a, b), max(a, b)))

            if state_pairs:
                sp_df = pd.DataFrame(state_pairs, columns=["State A", "State B"])
                sp_counts = (
                    sp_df.groupby(["State A", "State B"])
                    .size()
                    .reset_index(name="Count")
                )
                all_states = sorted(
                    set(sp_counts["State A"].unique()) | set(sp_counts["State B"].unique())
                )
                sp_matrix = pd.DataFrame(0, index=all_states, columns=all_states, dtype=int)
                for _, row in sp_counts.iterrows():
                    sp_matrix.at[row["State A"], row["State B"]] += row["Count"]
                    sp_matrix.at[row["State B"], row["State A"]] += row["Count"]
                fig_sp = px.imshow(
                    sp_matrix, text_auto=True, color_continuous_scale="Blues",
                    title="State/Territory Co-occurrence in Compound DRFA Clusters",
                    labels={"color": "Co-occurrences"},
                )
                st.plotly_chart(fig_sp, width="stretch")
            else:
                st.info("No multi-state compound clusters in the selected year range.")

            multi_cl = compound_cl[compound_cl["_multi_state"]].copy()
            if not multi_cl.empty:
                multi_cl["cluster_start"] = pd.to_datetime(multi_cl["cluster_start"])
                multi_cl["First Event"] = multi_cl["cluster_start"].dt.strftime("%d/%m/%Y")
                display_ms = multi_cl.rename(columns={
                    "fy":       "FY",
                    "n_events": "# Events",
                    "_n_states":"# States",
                    "states":   "States",
                    "perils":   "Perils",
                    "event_names": "Events (first 5)",
                })[["FY", "First Event", "# Events", "# States", "States", "Perils", "Events (first 5)"]]
                st.subheader("Multi-jurisdictional compound clusters")
                st.dataframe(display_ms.reset_index(drop=True), width="stretch", height=400)

    with t_dgms:
        st.subheader("DRFA Government Magnitude Scale (DGMS)")
        st.markdown("""
Adapted from Gissing et al. (2022) Table 3. The normalised loss dimension is replaced by
**total LGA activations within the cluster** — a direct measure of geographic spread and
government response burden.

| DGMS | # Component Events | Total Cluster LGA Activations |
|:---:|---|---|
| **I** | 2 | < 5 LGAs |
| **II** | 2 | 5–20 LGAs · *or* · 3 events < 5 LGAs |
| **III** | 2 | > 20 LGAs · *or* · 3 events 5–20 LGAs · *or* · ≥4 events < 5 LGAs |
| **IV** | 3 | > 20 LGAs · *or* · ≥4 events 5–20 LGAs |
| **V** | ≥ 4 | > 20 LGAs |

*Note: DGMS I–V vs Gissing's CDMS I–VIII reflects the narrower LGA-count range relative to the 4-orders-of-magnitude insurance loss scale.*
        """)

        compound_cl = cl[cl["_is_compound"] & cl["_magnitude"].notna()].copy()
        if compound_cl.empty:
            st.info("No compound clusters with DGMS rating in current selection.")
        else:
            _MAG_ORDER = ["I", "II", "III", "IV", "V"]
            mag_counts = (
                compound_cl["_magnitude"]
                .value_counts()
                .reindex(_MAG_ORDER, fill_value=0)
                .reset_index()
            )
            mag_counts.columns = ["DGMS", "Count"]
            fig_mag = px.bar(
                mag_counts, x="DGMS", y="Count",
                title="DRFA Compound Disasters by Magnitude Scale",
                color="Count", color_continuous_scale="YlOrRd",
                category_orders={"DGMS": _MAG_ORDER},
            )
            fig_mag.update_layout(coloraxis_showscale=False)
            st.plotly_chart(fig_mag, width="stretch")

            fig_sc = px.scatter(
                compound_cl, x="n_events", y="total_lgas",
                color="_magnitude",
                category_orders={"_magnitude": _MAG_ORDER},
                hover_data={"fy": True, "perils": True, "states": True, "event_names": True},
                labels={
                    "n_events":    "# Component Events",
                    "total_lgas":  "Total LGA Activations",
                    "_magnitude":  "DGMS",
                    "fy":          "Financial Year",
                    "perils":      "Perils",
                    "states":      "States",
                    "event_names": "Events",
                },
                title="DRFA Compound Clusters — Component Count vs Total LGA Activations",
                size="total_lgas", size_max=30,
            )
            st.plotly_chart(fig_sc, width="stretch")

    with t_climate:
        _CMAP = {
            "Compound disaster season": "#d62728",
            "No compound disasters":    "#1f77b4",
        }
        fy_compound = cl.groupby("fy")["_is_compound"].any().reset_index()
        fy_compound.columns = ["fy", "has_compound"]

        st.subheader("ENSO (ONI) and DRFA compound disaster seasons")
        try:
            oni = fetch_oni_data()
            oni["_fy"] = oni["date"].apply(lambda d: d.year if d.month >= 7 else d.year - 1)
            oni_fy = oni.groupby("_fy")["oni"].mean().reset_index()
            oni_fy.columns = ["fy", "mean_oni"]
            oni_merged = oni_fy.merge(fy_compound, on="fy", how="left")
            oni_merged["has_compound"] = oni_merged["has_compound"].astype("boolean").fillna(False).astype(bool)
            oni_merged["Season"] = oni_merged["has_compound"].map({
                True: "Compound disaster season", False: "No compound disasters",
            })
            oni_sel = oni_merged[oni_merged["fy"].between(sel_fy[0], sel_fy[1])]
            fig_oni = px.bar(
                oni_sel, x="fy", y="mean_oni", color="Season",
                color_discrete_map=_CMAP,
                title="Mean ONI by Financial Year — DRFA compound vs non-compound seasons",
                labels={"fy": "Financial Year (start)", "mean_oni": "Mean ONI (Niño 3.4)"},
            )
            fig_oni.add_hline(y=0.5,  line_dash="dot", line_color="orange",    annotation_text="El Niño (+0.5)")
            fig_oni.add_hline(y=-0.5, line_dash="dot", line_color="steelblue", annotation_text="La Niña (−0.5)")
            fig_oni.update_layout(legend=dict(orientation="h", y=-0.2))
            st.plotly_chart(fig_oni, width="stretch")
            fig_box = px.box(
                oni_sel, x="Season", y="mean_oni", color="Season",
                color_discrete_map=_CMAP, points="all",
                title="ONI distribution: DRFA compound vs non-compound financial years",
                labels={"mean_oni": "Mean ONI (Jul–Jun)"},
            )
            fig_box.add_hline(y=0.5,  line_dash="dot", line_color="orange")
            fig_box.add_hline(y=-0.5, line_dash="dot", line_color="steelblue")
            st.plotly_chart(fig_box, width="stretch")
            summary = (
                oni_sel.groupby("Season")["mean_oni"]
                .agg(n="count", mean="mean", median="median", std="std")
                .reset_index()
                .rename(columns={"n": "N years", "mean": "Mean ONI", "median": "Median ONI", "std": "Std ONI"})
            )
            for col in ("Mean ONI", "Median ONI", "Std ONI"):
                summary[col] = summary[col].round(3)
            st.dataframe(summary.reset_index(drop=True), width="stretch")
        except Exception as _e:
            st.warning(f"Could not load ONI data: {_e}")

    with t_compare:
        st.subheader("DRFA vs ICA Compound Season Comparison")
        st.caption(
            "Compound financial years identified by DRFA (government response burden) vs ICA "
            "(insured loss burden, A\\$100M threshold). Overlap = seasons where both datasets "
            "independently identify compound activity. Divergence reveals events captured by one "
            "dataset but not the other — e.g. low-insured but high-government-cost events."
        )
        try:
            _, cl_ica = load_compound_disasters(100.0, 91)
            drfa_compound_fy = set(cl[cl["_is_compound"]]["fy"].unique())
            ica_compound_fy  = set(cl_ica[cl_ica["_is_compound"]]["fy"].unique())
            all_fy_range = range(max(fy_min, 2006), sel_fy[1] + 1)
            compare_rows = []
            for fy in all_fy_range:
                in_drfa = fy in drfa_compound_fy
                in_ica  = fy in ica_compound_fy
                if in_drfa and in_ica:
                    label = "Both"
                elif in_drfa:
                    label = "DRFA only"
                elif in_ica:
                    label = "ICA only"
                else:
                    label = "Neither"
                compare_rows.append({"FY": fy, "Compound Season": label})
            comp_df = pd.DataFrame(compare_rows)
            _COMP_CMAP = {
                "Both":      "#2ca02c",
                "DRFA only": "#1f77b4",
                "ICA only":  "#ff7f0e",
                "Neither":   "#d3d3d3",
            }
            _ORDER = ["Both", "DRFA only", "ICA only", "Neither"]
            fig_comp = px.bar(
                comp_df, x="FY", y=[1] * len(comp_df), color="Compound Season",
                color_discrete_map=_COMP_CMAP,
                category_orders={"Compound Season": _ORDER},
                title="Compound Financial Years — DRFA vs ICA (2006 onwards)",
                labels={"FY": "Financial Year (start)", "y": ""},
            )
            fig_comp.update_yaxes(visible=False)
            fig_comp.update_layout(legend=dict(orientation="h", y=-0.2), bargap=0.1)
            st.plotly_chart(fig_comp, width="stretch")
            summary_counts = comp_df["Compound Season"].value_counts().reindex(_ORDER, fill_value=0).reset_index()
            summary_counts.columns = ["Compound Season", "Financial Years"]
            st.dataframe(summary_counts, width="stretch", hide_index=True)
            drfa_only = sorted(drfa_compound_fy - ica_compound_fy)
            ica_only  = sorted(ica_compound_fy  - drfa_compound_fy)
            if drfa_only:
                st.markdown(f"**DRFA-only compound years** (government burden without major insured loss): FY{', FY'.join(str(y) for y in drfa_only)}")
            if ica_only:
                st.markdown(f"**ICA-only compound years** (insured loss without DRFA compound activation): FY{', FY'.join(str(y) for y in ica_only)}")
        except Exception as _e:
            st.warning(f"Could not load ICA comparison data: {_e}")


def render_compound_disasters():  # noqa: C901
    """Gissing et al. (2022) compound disaster analysis applied to the ICA dataset."""
    from itertools import combinations as _combos

    st.title("Compound Disaster Analysis")
    st.caption(
        "Replicates the Gissing et al. (2022) methodology: two or more ICA catastrophe events "
        "within a 3-month (91-day) chain-linked window in the same Australian financial year "
        "(July–June) constitute a compound disaster. Dataset: ICA Historical Normalised Loss."
    )

    with st.expander("Methodology — Gissing et al. (2022)", expanded=False):
        st.markdown("""
**Reference:** Gissing, A., Crompton, R., McAneney, J., & Vidana-Rodriguez, R. (2022).
Compound natural disasters in Australia: a historical analysis.
*International Journal of Disaster Risk Reduction*, 72, 102812.

**Steps applied here:**
1. Filter ICA events to those above the selected normalised loss (NL) threshold (2022 AUD).
2. Assign each event to an **Australian financial year** (1 July – 30 June).
3. Within each FY, sort events by start date and apply **chain-link clustering**:
   if consecutive events start within `window_days` of each other they share a cluster.
   Any cluster of ≥ 2 events is a **compound disaster**.
4. Map ICA *Type* to Gissing's six peril categories:
   Tropical Cyclone · Flood · Storm · Bushfire · Heatwave · Landslide (+ Earthquake / Other).
5. Classify compound clusters on the **Compound Disaster Magnitude Scale (CDMS)** I–VIII
   using component count and total NL.
6. Link compound seasons to ENSO via the Oceanic Niño Index (ONI, NOAA CPC).

**Known limitations:**
- **Financial year boundary:** The algorithm clusters strictly within each FY (Jul–Jun). A sequence
  spanning June–July is split into two separate FYs and cannot form a single compound cluster. Verify
  manually any candidate compound sequence that crosses the July boundary (e.g. 2010–11 QLD floods).
- **Chain-link transitivity:** Events A → B → C chain together even if A and C are separated by
  2× `window_days`. The pair-wise gap only needs to be ≤ window between *consecutive* sorted events.
  Inspect long clusters for whether all members are genuinely co-occurring.
- **ICA covers insured losses only.** High-income states (NSW, VIC) have systematically higher ICA
  footprints due to greater insurance penetration and property values, independent of actual event severity.

*Note: ICA data begins 1967; DRFA matching covers 2006 onwards.*
        """)

    # ── controls ─────────────────────────────────────────────────────────────
    fc1, fc2 = st.columns(2)
    with fc1:
        _THRESH = {"$100M (base)": 100.0, "$500M": 500.0, "$1B": 1000.0, "$5B": 5000.0}
        sel_thresh_lbl = st.selectbox("Normalised loss threshold", list(_THRESH.keys()), key="cd_thresh")
        nl_thresh = _THRESH[sel_thresh_lbl]
    with fc2:
        window_days = st.slider(
            "Compound window (days)", 30, 180, 91, step=7,
            help="Gissing et al. use 91 days (~3 months).", key="cd_window",
        )

    with st.spinner("Computing compound clusters…"):
        ev_all, cl_all = load_compound_disasters(nl_thresh, window_days)

    st.session_state["_ica_ev_all"]     = ev_all
    st.session_state["_ica_cl_all"]     = cl_all
    st.session_state["_ica_fy_min"]     = int(ev_all["_fy"].min())
    st.session_state["_ica_fy_max"]     = int(ev_all["_fy"].max())
    st.session_state["_ica_thresh_lbl"] = sel_thresh_lbl
    _fragment_ica_compound()

def render_compound_disasters_drfa():  # noqa: C901
    """Gissing et al. (2022) compound disaster methodology adapted for the DRFA dataset."""
    from itertools import combinations as _combos

    st.title("DRFA Compound Disaster Analysis")
    st.caption(
        "Applies the Gissing et al. (2022) compound disaster methodology to the DRFA dataset. "
        "Each unique AGRN activation is treated as a disaster event. Events whose onset dates fall "
        "within a 91-day chain-linked window in the same Australian financial year (July–June) "
        "constitute a compound disaster. Severity is measured by total LGA activations (proxy for "
        "government response burden) rather than normalised insured loss."
    )

    with st.expander("Methodology — DRFA Adaptation of Gissing et al. (2022)", expanded=False):
        st.markdown("""
**Reference:** Gissing, A., Crompton, R., McAneney, J., & Vidana-Rodriguez, R. (2022).
Compound natural disasters in Australia: a historical analysis.
*International Journal of Disaster Risk Reduction*, 72, 102812.

**Adaptation for DRFA (this page):**
1. Each unique AGRN activation = one disaster event (809 events, 2006–2026).
2. Assign each event to an **Australian financial year** (1 July – 30 June).
3. Within each FY, sort events by onset date and apply **chain-link clustering**:
   if consecutive events start within `window_days` of each other they share a cluster.
   Any cluster of ≥ 2 events is a **compound disaster**.
4. Map DRFA *hazard_type* to Gissing's peril categories using the primary (first-listed) hazard.
5. Classify compound clusters on the **DRFA Government Magnitude Scale (DGMS)**:
   component count × total cluster LGA activations (replaces normalised loss).
6. Link compound seasons to ENSO via the Oceanic Niño Index (ONI, NOAA CPC).

**Why LGA count as severity proxy?**
Normalised insured loss (used in the ICA version) is unavailable for most DRFA events.
LGA activations are a direct measure of geographic spread and government response burden —
more LGAs activated = more local governments requiring Commonwealth support.

**Coverage:** DRFA 2006–2026 (20 financial years). ICA comparison available from FY2006.
        """)

    # ── controls ─────────────────────────────────────────────────────────────
    st.info(
        "**Note on window size:** Gissing et al. use 91 days for the ICA dataset, which has ~5–15 events "
        "per financial year above the A\\$100M threshold. The DRFA dataset has ~40 events per year, so a "
        "91-day window chains almost everything into one cluster per year (itself a finding — Australia is in "
        "near-continuous compound disaster response). A shorter window (7–30 days) identifies discrete "
        "**disaster bursts** — periods of concentrated concurrent activation — which is more analytically "
        "useful for this dataset. Default: 14 days.",
        icon="ℹ️",
    )
    window_days = st.slider(
        "Compound window (days)", 7, 91, 14, step=7,
        help="Gissing et al. use 91 days for ICA. For DRFA, 14 days identifies discrete disaster bursts. "
             "Increase to see how events chain into larger seasonal clusters.",
        key="drfa_cd_window",
    )

    with st.spinner("Computing DRFA compound clusters…"):
        ev_all, cl_all = load_compound_disasters_drfa(window_days)

    st.session_state["_drfa_ev_all"] = ev_all
    st.session_state["_drfa_cl_all"] = cl_all
    st.session_state["_drfa_fy_min"] = int(ev_all["_fy"].min())
    st.session_state["_drfa_fy_max"] = int(ev_all["_fy"].max())
    _fragment_drfa_compound()


def render_research_analysis():  # noqa: C901
    st.title("Research Analysis: External Support Demand")
    st.caption(
        "Analyses 809 DRFA (Disaster Recovery Funding Arrangements) activations 2006–2026 "
        "as the primary proxy for Commonwealth external support demand. "
        "Each unique AGRN = one declared disaster event. "
        "Use the filters below — all four tabs respond to the same selection."
    )

    with st.expander("Data & methodology notes", expanded=False):
        st.markdown("""
**What DRFA represents:** The Commonwealth activates DRFA when a state/territory cannot meet disaster costs alone.
Categories A (emergency relief), B (recovery), C (reconstruction) and D (extraordinary measures)
represent escalating levels of external assistance.

**Crosscheck (verified against source CSVs):**
| Dataset | Key figures |
|---|---|
| DRFA Activations | 5,967 LGA-event rows · **809 unique events** · 2006–2026 |
| DRFA Payments | **A\\$6.95 B paid** · A\\$6.96 B granted · only 44 of 809 events have published payment records |
| ICA Catastrophes | **744 events** · A\\$168.9 B normalised loss (2022 AUD) · 27 pre-2000 2-digit year artefacts corrected |
| AGD Events | **673 events** · 317,945 deaths (includes 29 non-Australian events: shipwrecks, pandemics, wars) |
| EM-DAT | **224 events** · 2,263 deaths · different inclusion scope to AGD |

**Concurrency method:** For each calendar day, count DRFA events whose active window overlaps that day.
Active window = disaster start date + user-selected duration (default 180 days).
Uses an O(n) sweep-line — not point-in-time simulation.

**⚠️ DRFA as a proxy — key limitations:**
- DRFA is a *funding trigger*, not an operational deployment record. Activation does not mean resources were physically deployed.
- The 2018 transition from NDRRA to DRFA changed eligibility criteria, which inflates apparent event counts post-2018 relative to pre-2018 (see vline annotation on charts).
- Only 44 of 809 events (5.4%) have published individual payment records — the payment dataset is not a complete picture of Commonwealth disbursements.
- DRFA does not record mutual aid deployments, volunteer contributions, or state-funded responses — the most operationally significant capacity information is absent.
        """)

    events = load_drfa_events()

    # ── filters ──────────────────────────────────────────────────────────────
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        all_hazards = sorted(events["hazard_group"].dropna().unique())
        sel_hazard = st.multiselect("Hazard group", all_hazards, placeholder="All types", key="ra_hazard")
    with fc2:
        all_states = sorted(events["state_primary"].dropna().unique())
        sel_states = st.multiselect("Primary state", all_states, placeholder="All states", key="ra_states")
    with fc3:
        yrs = events["year"].dropna().astype(int)
        sel_years = st.slider("Year range", int(yrs.min()), int(yrs.max()),
                              (int(yrs.min()), int(yrs.max())), key="ra_years")

    filt = events.copy()
    if sel_hazard:
        filt = filt[filt["hazard_group"].isin(sel_hazard)]
    if sel_states:
        filt = filt[filt["state_primary"].isin(sel_states)]
    filt = filt[(filt["year"] >= sel_years[0]) & (filt["year"] <= sel_years[1])]

    # ── summary metrics ───────────────────────────────────────────────────────
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Events", f"{len(filt):,}")
    m2.metric("LGAs affected", f"{filt['lga_count'].sum():,.0f}")
    m3.metric("States involved", f"{filt['state_primary'].nunique()}")
    m4.metric("Peak year", str(int(filt.groupby("year").size().idxmax())) if not filt.empty else "—")
    m5.metric("Most common hazard", filt["hazard_group"].mode().iat[0] if not filt.empty else "—")

    tab1, tab2, tab3, tab4 = st.tabs([
        "📈 Time Series", "🔥 Hazard Types", "⚡ Concurrency — Simultaneity", "🗺️ Spatial Demand",
    ])

    # ── TAB 1: TIME SERIES ────────────────────────────────────────────────────
    with tab1:
        st.subheader("External Support Demand Over Time")
        st.caption(
            "Each bar = DRFA events activated in that month. "
            "Line = 3-month rolling mean. Stacked by hazard group."
        )

        if filt.empty:
            st.info("No events match current filters.")
        else:
            monthly = (
                filt.groupby(["year_month", "hazard_group"])
                .size()
                .reset_index(name="events")
            )
            monthly["year_month"] = pd.to_datetime(monthly["year_month"])

            fig_ts = px.bar(
                monthly,
                x="year_month", y="events",
                color="hazard_group",
                title="DRFA Events per Month (stacked by hazard group)",
                labels={"year_month": "Month", "events": "DRFA Events Activated",
                        "hazard_group": "Hazard group"},
                color_discrete_sequence=px.colors.qualitative.Set2,
            )

            # Rolling 3-month mean (on total)
            monthly_total = filt.groupby("year_month").size().reset_index(name="events")
            monthly_total["year_month"] = pd.to_datetime(monthly_total["year_month"])
            monthly_total = monthly_total.sort_values("year_month")
            monthly_total["rolling_3m"] = monthly_total["events"].rolling(3, center=True).mean()

            fig_ts.add_scatter(
                x=monthly_total["year_month"],
                y=monthly_total["rolling_3m"],
                mode="lines",
                line=dict(color="black", width=2, dash="dot"),
                name="3-month rolling mean",
            )
            fig_ts.update_layout(
                xaxis_title="Month",
                yaxis_title="Number of DRFA Events",
                legend_title="Hazard group",
                barmode="stack",
            )
            # 2018 policy change: NDRRA replaced by DRFA (different activation thresholds)
            fig_ts.add_vline(
                x=pd.Timestamp("2018-11-01").timestamp() * 1000,
                line_dash="dash", line_color="red",
                annotation_text="NDRRA → DRFA (Nov 2018)",
                annotation_position="top left",
                annotation_font_color="red",
            )
            st.plotly_chart(fig_ts, width="stretch")
            st.caption(
                "⚠️ **Policy discontinuity:** DRFA replaced NDRRA in November 2018 with revised "
                "activation thresholds and category definitions. Pre- and post-2018 event counts "
                "are not directly comparable — apparent trends may partly reflect the policy change."
            )

            # Annual trend with Mann-Kendall test
            annual = filt.groupby("year").size().reset_index(name="events")
            annual["year"] = annual["year"].astype(int)

            # Mann-Kendall trend test (non-parametric, robust to non-normality)
            try:
                from scipy import stats as _stats
                _s_vals = annual["events"].values
                _mk_tau, _mk_p = _stats.kendalltau(range(len(_s_vals)), _s_vals)
                _trend_str = (
                    f"Mann-Kendall τ = {_mk_tau:+.3f}, p = {_mk_p:.3f} "
                    f"({'significant ↑' if _mk_p < 0.05 and _mk_tau > 0 else 'significant ↓' if _mk_p < 0.05 and _mk_tau < 0 else 'no significant trend'} at α = 0.05)"
                )
            except ImportError:
                _trend_str = "Install scipy for Mann-Kendall trend test."

            fig_ann = px.bar(
                annual, x="year", y="events",
                title="Annual DRFA Event Count (2006–2026)",
                labels={"year": "Year", "events": "DRFA Events"},
                color="events",
                color_continuous_scale="Blues",
            )
            fig_ann.add_vline(
                x=2018.5, line_dash="dash", line_color="red",
                annotation_text="NDRRA → DRFA",
                annotation_position="top left",
                annotation_font_color="red",
            )
            fig_ann.update_layout(coloraxis_showscale=False)
            st.plotly_chart(fig_ann, width="stretch")
            st.caption(f"📈 Trend test: {_trend_str}")

    # ── TAB 2: HAZARD TYPES ───────────────────────────────────────────────────
    with tab2:
        st.subheader("Hazard Type Analysis")

        if filt.empty:
            st.info("No events match current filters.")
        else:
            c1, c2 = st.columns(2)

            with c1:
                by_haz = (
                    filt.groupby("hazard_group")
                    .agg(events=("agrn", "count"), lgas=("lga_count", "sum"))
                    .reset_index()
                    .sort_values("events", ascending=True)
                )
                fig_haz = px.bar(
                    by_haz, x="events", y="hazard_group",
                    orientation="h",
                    title="DRFA Events by Hazard Group",
                    labels={"events": "Events", "hazard_group": "Hazard"},
                    color="events",
                    color_continuous_scale="Oranges",
                )
                fig_haz.update_layout(coloraxis_showscale=False)
                st.plotly_chart(fig_haz, width="stretch")

            with c2:
                fig_lga = px.bar(
                    by_haz, x="lgas", y="hazard_group",
                    orientation="h",
                    title="LGAs Affected by Hazard Group",
                    labels={"lgas": "LGA-event exposures", "hazard_group": "Hazard"},
                    color="lgas",
                    color_continuous_scale="Reds",
                )
                fig_lga.update_layout(coloraxis_showscale=False)
                st.plotly_chart(fig_lga, width="stretch")

            # State × hazard heatmap
            state_haz = (
                filt.groupby(["state_primary", "hazard_group"])
                .size()
                .reset_index(name="events")
            )
            fig_heat = px.density_heatmap(
                state_haz, x="hazard_group", y="state_primary", z="events",
                title="Event Count: State × Hazard Group",
                labels={"hazard_group": "Hazard", "state_primary": "State", "events": "Events"},
                color_continuous_scale="YlOrRd",
                text_auto=True,
            )
            fig_heat.update_layout(xaxis_tickangle=-20)
            st.plotly_chart(fig_heat, width="stretch")

            # DRFA categories breakdown
            cat_totals = pd.DataFrame({
                "Category": ["Cat A – Emergency relief", "Cat B – Recovery",
                             "Cat C – Reconstruction", "Cat D – Extraordinary"],
                "LGA-event pairs": [
                    filt["cat_A_lgas"].sum(), filt["cat_B_lgas"].sum(),
                    filt["cat_C_lgas"].sum(), filt["cat_D_lgas"].sum(),
                ],
            })
            fig_cat = px.bar(
                cat_totals, x="Category", y="LGA-event pairs",
                title="DRFA Category Activations (LGA-event pairs)",
                labels={"LGA-event pairs": "LGA-event pairs", "Category": ""},
                color="Category",
                color_discrete_sequence=["#4e79a7", "#f28e2b", "#e15759", "#76b7b2"],
            )
            fig_cat.update_layout(showlegend=False, xaxis_tickangle=-10)
            st.plotly_chart(fig_cat, width="stretch")

    # ── TAB 3: CONCURRENCY / SIMULTANEITY ─────────────────────────────────────
    with tab3:
        st.subheader("Concurrent Active Events — Simultaneity Problem")
        st.caption(
            "On any given day, how many DRFA events were still in their active support window? "
            "Multiple simultaneous events signal **national capacity stress** — "
            "the central research question of this PhD."
        )

        if filt.empty:
            st.info("No events match current filters.")
        else:
            duration_days = st.select_slider(
                "Assumed active support window per event (days)",
                options=[30, 60, 90, 180, 270, 365],
                value=180,
                key="ra_duration",
                help="Approximate period during which a DRFA event is consuming active support capacity. "
                     "180 days (6 months) is a conservative estimate for recovery-phase operations.",
            )

            conc = compute_concurrent_events(filt, duration_days)

            if conc.empty:
                st.info("Insufficient data for concurrency calculation.")
            else:
                # Peak stats
                peak_count = int(conc["active_events"].max())
                peak_date = conc.loc[conc["active_events"].idxmax(), "date"]

                pm1, pm2, pm3 = st.columns(3)
                pm1.metric("Peak simultaneous events", str(peak_count))
                pm2.metric("Peak date", peak_date.strftime("%d/%m/%Y"))
                pm3.metric(
                    "Days with ≥ 5 concurrent events",
                    f"{(conc['active_events'] >= 5).sum():,}",
                )

                # Add rolling 30-day mean for clarity
                conc = conc.sort_values("date")
                conc["rolling_30d"] = conc["active_events"].rolling(30, center=True).mean()

                fig_conc = px.area(
                    conc, x="date", y="active_events",
                    title=f"Concurrent Active DRFA Events per Day "
                          f"(active window = {duration_days} days)",
                    labels={"date": "Date", "active_events": "Active events"},
                    color_discrete_sequence=["#4e79a7"],
                )
                fig_conc.add_scatter(
                    x=conc["date"],
                    y=conc["rolling_30d"],
                    mode="lines",
                    line=dict(color="red", width=2),
                    name="30-day rolling mean",
                )
                # Annotate peak
                fig_conc.add_annotation(
                    x=peak_date, y=peak_count,
                    text=f"Peak: {peak_count} events<br>{peak_date.strftime('%b %Y')}",
                    showarrow=True, arrowhead=2,
                    bgcolor="white", bordercolor="red",
                    font=dict(size=13),
                )
                fig_conc.update_layout(
                    legend_title="",
                    xaxis_title="Date",
                    yaxis_title="Simultaneously active DRFA events",
                )
                st.plotly_chart(fig_conc, width="stretch")

                # Monthly heatmap: average concurrent events per year × month
                st.markdown("#### Monthly capacity pressure heatmap")
                st.caption(
                    "Average number of simultaneously active DRFA events in each calendar month. "
                    "Darker red = higher concurrent load. Read across a row to see seasonal patterns; "
                    "read down a column to see whether that month is getting worse over time."
                )
                conc["year"]  = conc["date"].dt.year
                conc["month"] = conc["date"].dt.month
                heatmap_data = (
                    conc.groupby(["year", "month"])["active_events"]
                    .mean()
                    .reset_index()
                )
                heatmap_pivot = (
                    heatmap_data.pivot(index="year", columns="month", values="active_events")
                    .rename(columns={
                        1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                        7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec",
                    })
                )
                fig_heat = go.Figure(data=go.Heatmap(
                    z=heatmap_pivot.values,
                    x=list(heatmap_pivot.columns),
                    y=[str(y) for y in heatmap_pivot.index],
                    colorscale="YlOrRd",
                    hoverongaps=False,
                    hovertemplate="Year: %{y}<br>Month: %{x}<br>Avg concurrent: %{z:.1f}<extra></extra>",
                    colorbar=dict(title="Avg concurrent<br>DRFA events"),
                ))
                fig_heat.update_layout(
                    title=f"Average concurrent active DRFA events (window = {duration_days} days)",
                    xaxis_title="Month",
                    yaxis_title="Year",
                    height=420,
                )
                st.plotly_chart(fig_heat, width="stretch")

                st.caption(
                    f"**Summary:** With a {duration_days}-day active window, the peak was "
                    f"**{peak_count} simultaneously active events** on {peak_date.strftime('%d/%m/%Y')}. "
                    f"The heatmap shows average load — the main time series above shows the day-by-day detail."
                )

    # ── TAB 4: SPATIAL DEMAND ─────────────────────────────────────────────────
    with tab4:
        st.subheader("Spatial Distribution of External Support Demand")
        st.caption(
            "Each state is shaded by the number of DRFA events where it was the **primary** "
            "affected state. Hover for details. Use the metric selector to switch between "
            "event count, LGA exposures, and Cat A (emergency relief) intensity."
        )

        if filt.empty:
            st.info("No events match current filters.")
        else:
            mc1, mc2 = st.columns(2)
            with mc1:
                metric_col = st.selectbox(
                    "Shade states by",
                    ["DRFA Events", "LGA Exposures", "Cat A – Emergency relief LGAs",
                     "Cat B – Recovery LGAs", "Cat C – Reconstruction LGAs"],
                    key="ra_metric",
                )
            with mc2:
                show_external_only = st.checkbox(
                    "Cat B+ only (recovery & reconstruction events)",
                    value=False,
                    key="ra_ext_only",
                    help="Filter to events where at least one LGA received Cat B, C, or D support.",
                )

            map_filt = filt.copy()
            if show_external_only:
                map_filt = map_filt[
                    (map_filt["cat_B_lgas"] > 0) |
                    (map_filt["cat_C_lgas"] > 0) |
                    (map_filt["cat_D_lgas"] > 0)
                ]

            # Build state-level summary
            state_summary = (
                map_filt.groupby("state_primary")
                .agg(
                    events=("agrn", "count"),
                    lga_exposures=("lga_count", "sum"),
                    cat_A=("cat_A_lgas", "sum"),
                    cat_B=("cat_B_lgas", "sum"),
                    cat_C=("cat_C_lgas", "sum"),
                    cat_D=("cat_D_lgas", "sum"),
                    dominant_hazard=("hazard_group", lambda x: x.mode().iat[0]),
                )
                .reset_index()
            )

            _metric_map = {
                "DRFA Events":                   ("events",        "DRFA Events"),
                "LGA Exposures":                 ("lga_exposures", "LGA Exposures"),
                "Cat A – Emergency relief LGAs": ("cat_A",         "Cat A LGA activations"),
                "Cat B – Recovery LGAs":         ("cat_B",         "Cat B LGA activations"),
                "Cat C – Reconstruction LGAs":   ("cat_C",         "Cat C LGA activations"),
            }
            color_col, color_label = _metric_map[metric_col]

            geojson = load_state_geojson()
            if geojson is None:
                st.warning("State boundary file (aus_states.geojson) not found in data directory.")
            else:
                fig_choro = px.choropleth_map(
                    state_summary,
                    geojson=geojson,
                    locations="state_primary",
                    featureidkey="properties.state",
                    color=color_col,
                    color_continuous_scale="YlOrRd",
                    range_color=(0, state_summary[color_col].max()),
                    map_style="open-street-map",
                    zoom=3,
                    center={"lat": -27.0, "lon": 134.0},
                    opacity=0.72,
                    labels={color_col: color_label, "state_primary": "State"},
                    hover_data={
                        "state_primary": True,
                        "events": True,
                        "lga_exposures": True,
                        "dominant_hazard": True,
                        color_col: True,
                    },
                    height=560,
                    title=f"{color_label} by State — DRFA 2006–2026",
                )
                fig_choro.update_layout(
                    margin={"r": 0, "t": 40, "l": 0, "b": 0},
                    coloraxis_colorbar=dict(title=color_label),
                )
                st.plotly_chart(fig_choro, width="stretch")

            # State-level summary table
            st.subheader("State-level Summary")
            display_summary = state_summary.sort_values("events", ascending=False).rename(columns={
                "state_primary": "State", "events": "Events",
                "lga_exposures": "LGA exposures", "cat_A": "Cat A LGAs",
                "cat_B": "Cat B LGAs", "cat_C": "Cat C LGAs", "cat_D": "Cat D LGAs",
                "dominant_hazard": "Dominant hazard",
            })
            st.dataframe(display_summary.reset_index(drop=True), width="stretch")
            csv = display_summary.to_csv(index=False).encode("utf-8")
            st.download_button("⬇ Download state summary CSV", data=csv,
                               file_name="drfa_state_summary.csv", mime="text/csv")


def render_state_cooccurrence():  # noqa: C901
    """State co-occurrence analysis: which state pairs are most often simultaneously under DRFA activation."""
    import numpy as np

    st.title("State Co-occurrence Analysis")
    st.caption(
        "How often are pairs of Australian states simultaneously under DRFA activation? "
        "Each cell = number of calendar days both states had at least one active disaster event. "
        "Active window = disaster start date + the duration you select below."
    )

    with st.expander("Methodology", expanded=False):
        st.markdown("""
**Method:** For each DRFA event the 'active window' is defined as `disaster_start_date` through
`disaster_start_date + duration_days`. For every calendar day in the dataset (2006–2026) we record
which states have at least one event active. We then count the number of days each pair of states
was *simultaneously* active.

**Why this matters:** High co-occurrence between two states signals compound or spatially widespread
disasters — events that simultaneously strain two state governments and may exhaust Commonwealth
support capacity.  The diagonal (single-state count) is the number of days that state had *any*
active event, providing the baseline for normalising pair counts as conditional probabilities.

**Data source:** DRFA activations CSV — deduplicated to one row per (STATE, AGRN) to avoid
LGA-level inflation.
        """)

    duration_days = st.slider(
        "Event active duration (days)", 30, 365, 180, step=30, key="sco_dur",
        help="Assumed number of days each DRFA event remains 'active' from its start date."
    )

    matrix_df, daily_df = compute_state_cooccurrence(duration_days)

    if matrix_df.empty:
        st.warning("No DRFA activation data found.")
        return

    states = list(matrix_df.columns)

    tab1, tab2, tab3 = st.tabs(["🗺️ Co-occurrence Matrix", "📊 Ranked Pairs", "📈 Time Trends"])

    # ── TAB 1: HEATMAP ────────────────────────────────────────────────────────
    with tab1:
        st.subheader("Days Both States Were Simultaneously Active")

        col_norm, col_anno = st.columns(2)
        with col_norm:
            norm_mode = st.radio(
                "Normalise values",
                ["Raw counts (days)", "Conditional probability (given row-state active)"],
                key="sco_norm",
                horizontal=True,
            )
        with col_anno:
            show_diag = st.checkbox("Show diagonal (single-state days)", value=False, key="sco_diag")

        # Build display matrix
        disp = matrix_df.copy().astype(float)

        if not show_diag:
            np.fill_diagonal(disp.values, np.nan)

        if norm_mode.startswith("Conditional"):
            diag_vals = np.diag(matrix_df.values).astype(float)
            diag_vals[diag_vals == 0] = np.nan
            # P(both active | row-state active) = cooc[i,j] / diag[i]
            disp = disp.div(diag_vals, axis=0)
            color_label = "Conditional probability"
            fmt_str = ".2f"
            z_title = "P(col active | row active)"
        else:
            color_label = "Co-occurrence days"
            fmt_str = ".0f"
            z_title = "Days both states active"

        # Short state abbreviations for axis labels
        _ABBREV = {
            "Queensland": "QLD", "New South Wales": "NSW", "Victoria": "VIC",
            "South Australia": "SA", "Western Australia": "WA", "Tasmania": "TAS",
            "Northern Territory": "NT", "Australian Capital Territory": "ACT",
        }
        abbrev_states = [_ABBREV.get(s, s) for s in states]

        z_vals = disp.values.tolist()
        fig_heat = go.Figure(go.Heatmap(
            z=z_vals,
            x=abbrev_states,
            y=abbrev_states,
            colorscale="YlOrRd",
            text=[[
                f"{v:{fmt_str}}" if not (isinstance(v, float) and np.isnan(v)) else ""
                for v in row
            ] for row in z_vals],
            texttemplate="%{text}",
            colorbar=dict(title=color_label),
            hoverongaps=False,
        ))
        fig_heat.update_layout(
            title=f"State Co-occurrence ({z_title}) — {duration_days}-day active window",
            xaxis_title="State",
            yaxis_title="State",
            yaxis_autorange="reversed",
            height=520,
        )
        st.plotly_chart(fig_heat, width="stretch")

        st.caption(
            "Diagonal (if shown) = days that state had any active DRFA event. "
            "Off-diagonal = days both states were simultaneously active."
        )

    # ── TAB 2: RANKED PAIRS ───────────────────────────────────────────────────
    with tab2:
        st.subheader("Top Co-occurring State Pairs")

        # Extract upper triangle
        pair_rows = []
        for i, s1 in enumerate(states):
            for j, s2 in enumerate(states):
                if j <= i:
                    continue
                days_both = int(matrix_df.iloc[i, j])
                days_s1   = int(matrix_df.iloc[i, i])
                days_s2   = int(matrix_df.iloc[j, j])
                denom = max(days_s1, days_s2) if max(days_s1, days_s2) > 0 else 1
                pair_rows.append({
                    "State A":      s1,
                    "State B":      s2,
                    "Pair":         f"{_ABBREV.get(s1, s1)} + {_ABBREV.get(s2, s2)}",
                    "Days both active": days_both,
                    "Cond. prob (%)": round(days_both / denom * 100, 1),
                })
        pair_df = pd.DataFrame(pair_rows).sort_values("Days both active", ascending=False)

        top_n = st.slider("Show top N pairs", 5, len(pair_df), min(20, len(pair_df)), key="sco_topn")
        plot_pairs = pair_df.head(top_n).sort_values("Days both active")

        fig_bar = px.bar(
            plot_pairs,
            x="Days both active",
            y="Pair",
            orientation="h",
            color="Cond. prob (%)",
            color_continuous_scale="YlOrRd",
            title=f"Top {top_n} State Pairs by Co-occurrence Days ({duration_days}-day window)",
            labels={"Days both active": "Days simultaneously active", "Pair": "State pair"},
            text="Days both active",
            height=max(350, top_n * 28),
        )
        fig_bar.update_traces(textposition="outside")
        fig_bar.update_layout(
            coloraxis_colorbar=dict(title="Cond. prob (%)"),
            yaxis_title="",
        )
        st.plotly_chart(fig_bar, width="stretch")

        # Summary table
        st.dataframe(
            pair_df.reset_index(drop=True).rename(columns={"Cond. prob (%)": "Cond. prob % (/ max solo days)"}),
            width="stretch",
        )
        csv_pairs = pair_df.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download pair table CSV", data=csv_pairs,
                           file_name="drfa_state_cooccurrence_pairs.csv", mime="text/csv")

    # ── TAB 3: TIME TRENDS ────────────────────────────────────────────────────
    with tab3:
        st.subheader("Multi-state Activation Over Time")

        if daily_df.empty:
            st.info("No daily data available.")
        else:
            daily_df["date"] = pd.to_datetime(daily_df["date"])
            daily_df["year"] = daily_df["date"].dt.year

            # ── Annual: days with 2+, 3+, 4+ states active ───────────────────
            thresholds = [2, 3, 4]
            annual_rows = []
            for yr, grp in daily_df.groupby("year"):
                for thr in thresholds:
                    n = int((grp["n_states_active"] >= thr).sum())
                    annual_rows.append({"Year": yr, "Threshold": f"≥{thr} states", "Days": n})
            annual_thr = pd.DataFrame(annual_rows)

            fig_thr = px.line(
                annual_thr,
                x="Year", y="Days",
                color="Threshold",
                markers=True,
                title=f"Days Per Year with Multiple States Simultaneously Active ({duration_days}-day window)",
                labels={"Days": "Days / year", "Threshold": ""},
                color_discrete_sequence=["#2196F3", "#FF9800", "#F44336"],
            )
            fig_thr.update_layout(legend_title="")
            st.plotly_chart(fig_thr, width="stretch")

            # ── Top pair trend over time ─────────────────────────────────────
            st.markdown("**Annual co-occurrence for the top 5 state pairs**")
            top5_pairs = pair_df.head(5)[["State A", "State B", "Pair"]].copy()

            trend_rows = []
            for yr, grp in daily_df.groupby("year"):
                grp_states = grp[states] if all(s in grp.columns for s in states) else pd.DataFrame()
                if grp_states.empty:
                    continue
                for _, prow in top5_pairs.iterrows():
                    sa, sb = prow["State A"], prow["State B"]
                    if sa in grp_states.columns and sb in grp_states.columns:
                        both = int((grp_states[sa].values & grp_states[sb].values).sum())
                        trend_rows.append({"Year": yr, "Pair": prow["Pair"], "Days": both})

            if trend_rows:
                trend_df = pd.DataFrame(trend_rows)
                fig_trend = px.line(
                    trend_df,
                    x="Year", y="Days",
                    color="Pair",
                    markers=True,
                    title="Annual Co-occurrence Days — Top 5 State Pairs",
                    labels={"Days": "Co-occurrence days / year"},
                )
                fig_trend.update_layout(legend_title="State pair")
                st.plotly_chart(fig_trend, width="stretch")
            else:
                st.info("No trend data available.")

            # ── Monthly heatmap (year × month) for any multi-state days ──────
            st.markdown("**Monthly calendar: days with ≥2 states active**")
            daily_df["month"] = daily_df["date"].dt.month
            monthly_ms = (
                daily_df[daily_df["n_states_active"] >= 2]
                .groupby(["year", "month"])
                .size()
                .reset_index(name="days")
            )
            if not monthly_ms.empty:
                pivot = monthly_ms.pivot(index="year", columns="month", values="days").fillna(0)
                month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                               "Jul","Aug","Sep","Oct","Nov","Dec"]
                pivot.columns = [month_names[m - 1] for m in pivot.columns]
                fig_cal = go.Figure(go.Heatmap(
                    z=pivot.values.tolist(),
                    x=list(pivot.columns),
                    y=[str(y) for y in pivot.index],
                    colorscale="Blues",
                    colorbar=dict(title="Days"),
                ))
                fig_cal.update_layout(
                    title="Days per Month with ≥2 States Simultaneously Active",
                    xaxis_title="Month",
                    yaxis_title="Year",
                    height=420,
                )
                st.plotly_chart(fig_cal, width="stretch")


def render_oni():  # noqa: C901
    """ENSO / Oceanic Niño Index explorer page."""
    st.title("ENSO — Oceanic Niño Index")
    st.caption(
        "ONI = 3-month running mean of SST anomaly in the Niño 3.4 region "
        "(5°N–5°S, 120°–170°W). Threshold: ±0.5 °C for 5 consecutive overlapping "
        "seasons = El Niño / La Niña episode.  \n"
        "Source: [**NOAA Climate Prediction Center**](https://www.cpc.ncep.noaa.gov/products/analysis_monitoring/ensostuff/ONI_v5.php) · "
        "[raw data](https://www.cpc.ncep.noaa.gov/data/indices/oni.ascii.txt) · data cached for 24 h."
    )

    try:
        oni = fetch_oni_data()
    except Exception as exc:
        st.error(f"Could not load ONI data: {exc}")
        st.info(
            "The app tried to fetch from NOAA CPC and found no local cache. "
            "Check your internet connection and reload."
        )
        return

    # ── Summary metrics ───────────────────────────────────────────────────────
    latest     = oni.iloc[-1]
    _PHASE_ICON = {"El Niño": "🔴", "La Niña": "🔵", "Neutral": "⚪"}
    _PHASE_CLR  = {"El Niño": "#d62728", "La Niña": "#1f77b4", "Neutral": "#7f7f7f"}

    m1, m2, m3, m4 = st.columns(4)
    m1.metric(
        f"Latest ONI ({latest['date'].strftime('%b %Y')})",
        f"{latest['oni']:+.2f} °C",
    )
    m2.metric(
        "Current phase",
        f"{_PHASE_ICON[latest['enso_phase']]} {latest['enso_phase']}",
    )
    m3.metric(
        "El Niño months on record",
        f"{(oni['enso_phase'] == 'El Niño').sum()}",
        help="Number of months (since 1950) with ONI ≥ +0.5 °C",
    )
    m4.metric(
        "La Niña months on record",
        f"{(oni['enso_phase'] == 'La Niña').sum()}",
        help="Number of months (since 1950) with ONI ≤ −0.5 °C",
    )

    tab1, tab3 = st.tabs(["📈 ONI Time Series", "📋 Data"])

    # ── TAB 1: ONI TIME SERIES ────────────────────────────────────────────────
    with tab1:
        st.subheader("ONI Time Series (1950 – present)")

        yr_min = int(oni["year"].min())
        yr_max = int(oni["year"].max())
        y_start, y_end = st.slider(
            "Year range", yr_min, yr_max, (max(yr_min, 2000), yr_max),
            key="oni_yr_range",
        )
        view = oni[(oni["year"] >= y_start) & (oni["year"] <= y_end)].copy()

        view["_roll12"] = view["oni"].rolling(12, center=True, min_periods=6).mean()

        fig_oni = go.Figure()
        # Red fill above zero (El Niño)
        fig_oni.add_trace(go.Scatter(
            x=view["date"], y=view["oni"].clip(lower=0),
            fill="tozeroy", fillcolor="rgba(180, 30, 30, 0.55)",
            line=dict(width=0), showlegend=True,
            name="Warm / El Niño", hoverinfo="skip",
        ))
        # Blue fill below zero (La Niña)
        fig_oni.add_trace(go.Scatter(
            x=view["date"], y=view["oni"].clip(upper=0),
            fill="tozeroy", fillcolor="rgba(30, 90, 180, 0.55)",
            line=dict(width=0), showlegend=True,
            name="Cool / La Niña", hoverinfo="skip",
        ))
        # Thin neutral monthly line
        fig_oni.add_trace(go.Scatter(
            x=view["date"], y=view["oni"],
            mode="lines", line=dict(color="rgba(40,40,40,0.35)", width=0.8),
            showlegend=False, hovertemplate="%{x|%b %Y}<br>ONI: %{y:+.2f} °C<extra></extra>",
        ))
        # 12-month rolling mean
        fig_oni.add_trace(go.Scatter(
            x=view["date"], y=view["_roll12"],
            mode="lines", line=dict(color="#222222", width=2),
            name="12-month mean",
            hovertemplate="%{x|%b %Y}<br>12-mo mean: %{y:+.2f} °C<extra></extra>",
        ))
        # Zero baseline + ±0.5 thresholds
        fig_oni.add_hline(y=0,   line=dict(color="black", width=1.2))
        fig_oni.add_hline(y= 0.5, line=dict(color="#b41c1c", dash="dash", width=1),
                          annotation_text="+0.5 °C (El Niño)", annotation_position="top right")
        fig_oni.add_hline(y=-0.5, line=dict(color="#1c5ab4", dash="dash", width=1),
                          annotation_text="−0.5 °C (La Niña)", annotation_position="bottom right")
        fig_oni.update_layout(
            xaxis_title="Date", yaxis_title="ONI (°C anomaly)",
            legend=dict(orientation="h", y=1.08),
            height=420, hovermode="x unified",
        )
        st.plotly_chart(fig_oni, width="stretch")
        st.caption(
            "Red fill = warm anomaly (El Niño tendency); blue fill = cool anomaly (La Niña tendency). "
            "Black line = 12-month centred rolling mean. Dashed lines = ±0.5 °C ENSO classification threshold. "
            "Format follows NOAA CPC / Met Office convention."
        )
        st.info(
            "**Literature synthesis:** ENSO explains ~25% of eastern Australian rainfall variability in spring. "
            "La Niña produces stronger, more widespread impacts than El Niño (78.6% vs 56.6% land area affected in SON). "
            "Impacts are amplified when ENSO aligns with IOD and SAM — see **Climate Science** in the sidebar.",
            icon="📚",
        )

    # ── TAB 3: DATA TABLE ─────────────────────────────────────────────────────
    with tab3:
        st.subheader("Raw ONI Data")

        col_filter = st.columns(3)
        with col_filter[0]:
            phase_filter = st.multiselect(
                "Filter by phase",
                ["El Niño", "Neutral", "La Niña"],
                default=["El Niño", "Neutral", "La Niña"],
                key="oni_phase_filter",
            )
        with col_filter[1]:
            yr_f1, yr_f2 = st.slider(
                "Year range",
                int(oni["year"].min()), int(oni["year"].max()),
                (int(oni["year"].min()), int(oni["year"].max())),
                key="oni_tbl_range",
            )

        tbl = oni[
            (oni["enso_phase"].isin(phase_filter)) &
            (oni["year"] >= yr_f1) &
            (oni["year"] <= yr_f2)
        ].copy()
        tbl["date"] = tbl["date"].dt.strftime("%b %Y")
        tbl = tbl.rename(columns={
            "date": "Centre Month", "year": "Year", "month": "Mo.",
            "season": "Season (3-month)", "oni": "ONI (°C)", "enso_phase": "ENSO Phase",
        })[["Centre Month", "Year", "Season (3-month)", "ONI (°C)", "ENSO Phase"]]

        def _colour_phase(val: str) -> str:
            return {
                "El Niño": "background-color:#ffd7d7",
                "La Niña": "background-color:#d0e8ff",
            }.get(val, "")

        st.dataframe(
            tbl.reset_index(drop=True).style.map(_colour_phase, subset=["ENSO Phase"]),
            width="stretch",
            height=420,
        )
        csv_oni = tbl.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇ Download ONI CSV",
            data=csv_oni,
            file_name="oni_data.csv",
            mime="text/csv",
        )


def render_sam():  # noqa: C901
    """Southern Annular Mode (Marshall index) explorer page."""
    st.title("SAM — Southern Annular Mode")
    st.caption(
        "Monthly Marshall SAM index, 1957 – present. "
        "Positive SAM = poleward-shifted westerlies → reduced rainfall in southern Australia → elevated fire and drought risk.  \n"
        "Negative SAM = equatorward-shifted westerlies → increased mid-latitude precipitation → elevated flood and cyclone risk.  \n"
        "Index standardised against 1971–2000 baseline.  \n"
        "Sources: [**BAS Marshall SAM**](https://legacy.bas.ac.uk/met/gjma/sam.html) (1957–1978, station-based) · "
        "[**NOAA CPC AAO**](https://www.cpc.ncep.noaa.gov/products/precip/CWlink/daily_ao_index/aao/aao.shtml) "
        "(1979–present, reanalysis-based) · refreshed hourly when data is stale."
    )

    try:
        sam = fetch_sam_data()
    except Exception as exc:
        st.error(f"Could not load SAM data: {exc}")
        st.info(
            "The app tried BAS (historical) and NOAA CPC (live) and found no local cache. "
            "Check your internet connection and reload."
        )
        return

    _SAM_PHASE_CLR = {
        "Positive SAM": "#d62728",
        "Negative SAM": "#1f77b4",
        "Neutral":       "#7f7f7f",
    }
    _SAM_PHASE_ICON = {
        "Positive SAM": "🔴",
        "Negative SAM": "🔵",
        "Neutral":       "⚪",
    }

    latest = sam.iloc[-1]
    m1, m2, m3, m4 = st.columns(4)
    m1.metric(
        f"Latest SAM ({latest['date'].strftime('%b %Y')})",
        f"{latest['sam']:+.2f}",
    )
    m2.metric(
        "Current phase",
        f"{_SAM_PHASE_ICON[latest['sam_phase']]} {latest['sam_phase']}",
    )
    m3.metric(
        "Positive SAM months on record",
        f"{(sam['sam_phase'] == 'Positive SAM').sum()}",
        help="Months (since 1957) with SAM ≥ +1.0",
    )
    m4.metric(
        "Negative SAM months on record",
        f"{(sam['sam_phase'] == 'Negative SAM').sum()}",
        help="Months (since 1957) with SAM ≤ −1.0",
    )

    tab1, tab3 = st.tabs(["📈 SAM Time Series", "📋 Data"])

    # ── TAB 1: SAM TIME SERIES ────────────────────────────────────────────────
    with tab1:
        st.subheader("SAM Time Series (1957 – present)")

        yr_min = int(sam["year"].min())
        yr_max = int(sam["year"].max())
        y_start, y_end = st.slider(
            "Year range", yr_min, yr_max, (max(yr_min, 2000), yr_max),
            key="sam_yr_range",
        )
        view = sam[(sam["year"] >= y_start) & (sam["year"] <= y_end)].copy()

        view["_roll12"] = view["sam"].rolling(12, center=True, min_periods=6).mean()

        fig_sam = go.Figure()
        # Red fill above zero (Positive SAM — fire/drought risk)
        fig_sam.add_trace(go.Scatter(
            x=view["date"], y=view["sam"].clip(lower=0),
            fill="tozeroy", fillcolor="rgba(180, 30, 30, 0.55)",
            line=dict(width=0), showlegend=True,
            name="Positive SAM", hoverinfo="skip",
        ))
        # Blue fill below zero (Negative SAM — flood/rain risk)
        fig_sam.add_trace(go.Scatter(
            x=view["date"], y=view["sam"].clip(upper=0),
            fill="tozeroy", fillcolor="rgba(30, 90, 180, 0.55)",
            line=dict(width=0), showlegend=True,
            name="Negative SAM", hoverinfo="skip",
        ))
        # Thin neutral monthly line
        fig_sam.add_trace(go.Scatter(
            x=view["date"], y=view["sam"],
            mode="lines", line=dict(color="rgba(40,40,40,0.35)", width=0.8),
            showlegend=False, hovertemplate="%{x|%b %Y}<br>SAM: %{y:+.2f}<extra></extra>",
        ))
        # 12-month rolling mean
        fig_sam.add_trace(go.Scatter(
            x=view["date"], y=view["_roll12"],
            mode="lines", line=dict(color="#222222", width=2),
            name="12-month mean",
            hovertemplate="%{x|%b %Y}<br>12-mo mean: %{y:+.2f}<extra></extra>",
        ))
        # Zero baseline + ±1.0 thresholds
        fig_sam.add_hline(y=0,   line=dict(color="black", width=1.2))
        fig_sam.add_hline(y= 1.0, line=dict(color="#b41c1c", dash="dash", width=1),
                          annotation_text="+1.0 (Positive SAM)", annotation_position="top right")
        fig_sam.add_hline(y=-1.0, line=dict(color="#1c5ab4", dash="dash", width=1),
                          annotation_text="−1.0 (Negative SAM)", annotation_position="bottom right")
        fig_sam.update_layout(
            xaxis_title="Date", yaxis_title="SAM index (standardised)",
            legend=dict(orientation="h", y=1.08),
            height=420, hovermode="x unified",
        )
        st.plotly_chart(fig_sam, width="stretch")
        st.caption(
            "Red fill = Positive SAM (enhanced westerlies → reduced southern Australian rainfall / elevated fire risk); "
            "blue fill = Negative SAM (weakened westerlies → elevated rainfall / flood risk). "
            "Black line = 12-month centred rolling mean. Dashed lines = ±1.0 phase threshold. "
            "Format follows BoM / BAS convention."
        )
        st.info(
            "**Literature synthesis:** SAM explains 10–15% of weekly rainfall variance in southern Australia — "
            "comparable to ENSO. Positive SAM in winter reduces SW/SE Australia rainfall; in summer it increases "
            "east coast rainfall. SAM has trended positive since the 1970s due to ozone depletion and GHG forcing. "
            "See **Climate Science** in the sidebar for compound effects with ENSO and IOD.",
            icon="📚",
        )

    # ── TAB 3: DATA TABLE ────────────────────────────────────────────────────
    with tab3:
        st.subheader("Raw SAM Data")

        col_filter = st.columns(3)
        with col_filter[0]:
            phase_filter = st.multiselect(
                "Filter by phase",
                ["Positive SAM", "Neutral", "Negative SAM"],
                default=["Positive SAM", "Neutral", "Negative SAM"],
                key="sam_phase_filter",
            )
        with col_filter[1]:
            yr_f1, yr_f2 = st.slider(
                "Year range",
                int(sam["year"].min()), int(sam["year"].max()),
                (int(sam["year"].min()), int(sam["year"].max())),
                key="sam_tbl_range",
            )

        tbl = sam[
            (sam["sam_phase"].isin(phase_filter)) &
            (sam["year"] >= yr_f1) &
            (sam["year"] <= yr_f2)
        ].copy()
        tbl["date"] = tbl["date"].dt.strftime("%b %Y")
        tbl = tbl.rename(columns={
            "date": "Month", "year": "Year", "month": "Mo.",
            "sam": "SAM Index", "sam_phase": "SAM Phase",
        })[["Month", "Year", "Mo.", "SAM Index", "SAM Phase"]]

        def _colour_sam_phase(val: str) -> str:
            return {
                "Positive SAM": "background-color:#ffd7d7",
                "Negative SAM": "background-color:#d0e8ff",
            }.get(val, "")

        st.dataframe(
            tbl.reset_index(drop=True).style.map(_colour_sam_phase, subset=["SAM Phase"]),
            width="stretch",
            height=420,
        )
        csv_sam = tbl.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇ Download SAM CSV",
            data=csv_sam,
            file_name="sam_data.csv",
            mime="text/csv",
        )


def render_iod():  # noqa: C901
    """Indian Ocean Dipole (Dipole Mode Index) explorer page."""
    st.title("IOD — Indian Ocean Dipole")
    st.caption(
        "Monthly Dipole Mode Index (DMI), January 1870 – present. "
        "DMI = SST anomaly in the western Indian Ocean (50°E–70°E, 10°S–10°N) minus the eastern box (90°E–110°E, 10°S–0°).  \n"
        "**Positive IOD** (DMI ≥ +0.4 °C): warm west / cool east → reduced rainfall, elevated fire and drought risk in southern and eastern Australia.  \n"
        "**Negative IOD** (DMI ≤ −0.4 °C): cool west / warm east → increased rainfall, elevated flood and cyclone risk.  \n"
        "IOD is most active during austral winter–spring (June–November).  \n"
        "Sources: [**NOAA PSL / HadISST1.1**](https://psl.noaa.gov/data/timeseries/month/DMI/) (Jan 1870 – ~12 months ago) "
        "extended with [**BoM GAMSSA**](http://www.bom.gov.au/climate/iod/) weekly data (Jul 2008 – present) for the most recent months.  \n"
        "Saji & Yamagata (2003) · threshold: ±0.4 °C (BoM standard) · data cached for 24 h."
    )

    try:
        iod = fetch_iod_data()
    except Exception as exc:
        st.error(f"Could not load IOD data: {exc}")
        st.info(
            "The app tried to fetch from NOAA PSL and found no local cache. "
            "Check your internet connection and reload."
        )
        return

    _IOD_PHASE_CLR = {
        "Positive IOD": "#d62728",
        "Negative IOD": "#1f77b4",
        "Neutral":       "#7f7f7f",
    }
    _IOD_PHASE_ICON = {
        "Positive IOD": "🔴",
        "Negative IOD": "🔵",
        "Neutral":       "⚪",
    }

    latest = iod.iloc[-1]
    m1, m2, m3, m4 = st.columns(4)
    m1.metric(
        f"Latest DMI ({latest['date'].strftime('%b %Y')})",
        f"{latest['dmi']:+.3f} °C",
    )
    m2.metric(
        "Current phase",
        f"{_IOD_PHASE_ICON[latest['iod_phase']]} {latest['iod_phase']}",
    )
    m3.metric(
        "Positive IOD months on record",
        f"{(iod['iod_phase'] == 'Positive IOD').sum()}",
        help="Months (since 1870) with DMI ≥ +0.4 °C",
    )
    m4.metric(
        "Negative IOD months on record",
        f"{(iod['iod_phase'] == 'Negative IOD').sum()}",
        help="Months (since 1870) with DMI ≤ −0.4 °C",
    )

    tab1, tab3 = st.tabs(["📈 DMI Time Series", "📋 Data"])

    # ── TAB 1: DMI TIME SERIES ────────────────────────────────────────────────
    with tab1:
        st.subheader("DMI Time Series (1870 – present)")

        yr_min = int(iod["year"].min())
        yr_max = int(iod["year"].max())
        y_start, y_end = st.slider(
            "Year range", yr_min, yr_max, (max(yr_min, 1960), yr_max),
            key="iod_yr_range",
        )
        view = iod[(iod["year"] >= y_start) & (iod["year"] <= y_end)].copy()

        view["_roll12"] = view["dmi"].rolling(12, center=True, min_periods=6).mean()

        fig_iod = go.Figure()
        # Red fill above zero (Positive IOD — drought/fire risk)
        fig_iod.add_trace(go.Scatter(
            x=view["date"], y=view["dmi"].clip(lower=0),
            fill="tozeroy", fillcolor="rgba(180, 30, 30, 0.55)",
            line=dict(width=0), showlegend=True,
            name="Positive IOD", hoverinfo="skip",
        ))
        # Blue fill below zero (Negative IOD — flood/cyclone risk)
        fig_iod.add_trace(go.Scatter(
            x=view["date"], y=view["dmi"].clip(upper=0),
            fill="tozeroy", fillcolor="rgba(30, 90, 180, 0.55)",
            line=dict(width=0), showlegend=True,
            name="Negative IOD", hoverinfo="skip",
        ))
        # Thin neutral monthly line
        fig_iod.add_trace(go.Scatter(
            x=view["date"], y=view["dmi"],
            mode="lines", line=dict(color="rgba(40,40,40,0.35)", width=0.8),
            showlegend=False, hovertemplate="%{x|%b %Y}<br>DMI: %{y:+.3f} °C<extra></extra>",
        ))
        # 12-month rolling mean
        fig_iod.add_trace(go.Scatter(
            x=view["date"], y=view["_roll12"],
            mode="lines", line=dict(color="#222222", width=2),
            name="12-month mean",
            hovertemplate="%{x|%b %Y}<br>12-mo mean: %{y:+.3f} °C<extra></extra>",
        ))
        # Zero baseline + ±0.4 thresholds
        fig_iod.add_hline(y=0,    line=dict(color="black", width=1.2))
        fig_iod.add_hline(y= 0.4, line=dict(color="#b41c1c", dash="dash", width=1),
                          annotation_text="+0.4 °C (pIOD)", annotation_position="top right")
        fig_iod.add_hline(y=-0.4, line=dict(color="#1c5ab4", dash="dash", width=1),
                          annotation_text="−0.4 °C (nIOD)", annotation_position="bottom right")
        fig_iod.update_layout(
            xaxis_title="Date", yaxis_title="DMI (°C anomaly)",
            legend=dict(orientation="h", y=1.08),
            height=420, hovermode="x unified",
        )
        st.plotly_chart(fig_iod, width="stretch")
        st.caption(
            "Red fill = Positive IOD (warm west / cool east Indian Ocean → drought and fire risk in southern/eastern Australia); "
            "blue fill = Negative IOD (cool west / warm east → flood and cyclone risk). "
            "Black line = 12-month centred rolling mean. Dashed lines = ±0.4 °C BoM classification threshold. "
            "Format follows NOAA CPC / BoM convention. "
            "Note: IOD is most active June–November; the 12-month mean smooths the seasonal signal."
        )
        st.info(
            "**Literature synthesis:** Every major southern Australian drought since 1889 coincided with "
            "positive or neutral IOD. Extreme positive IOD events are projected to occur ~3× more frequently "
            "this century. IOD variability is partly energised by ENSO — robust attribution requires care. "
            "See **Climate Science** in the sidebar for detailed findings and compound effects.",
            icon="📚",
        )

    # ── TAB 2: IOD × DRFA ────────────────────────────────────────────────────
    # ── TAB 3: DATA TABLE ────────────────────────────────────────────────────
    with tab3:
        st.subheader("Raw DMI Data")

        col_filter = st.columns(3)
        with col_filter[0]:
            phase_filter = st.multiselect(
                "Filter by phase",
                ["Positive IOD", "Neutral", "Negative IOD"],
                default=["Positive IOD", "Neutral", "Negative IOD"],
                key="iod_phase_filter",
            )
        with col_filter[1]:
            yr_f1, yr_f2 = st.slider(
                "Year range",
                int(iod["year"].min()), int(iod["year"].max()),
                (int(iod["year"].min()), int(iod["year"].max())),
                key="iod_tbl_range",
            )

        tbl = iod[
            (iod["iod_phase"].isin(phase_filter)) &
            (iod["year"] >= yr_f1) &
            (iod["year"] <= yr_f2)
        ].copy()
        tbl["date"] = tbl["date"].dt.strftime("%b %Y")
        tbl = tbl.rename(columns={
            "date": "Month", "year": "Year", "month": "Mo.",
            "dmi": "DMI (°C)", "iod_phase": "IOD Phase", "source": "Source",
        })[["Month", "Year", "Mo.", "DMI (°C)", "IOD Phase", "Source"]]

        def _colour_iod_phase(val: str) -> str:
            return {
                "Positive IOD": "background-color:#ffd7d7",
                "Negative IOD": "background-color:#d0e8ff",
            }.get(val, "")

        st.dataframe(
            tbl.reset_index(drop=True).style.map(_colour_iod_phase, subset=["IOD Phase"]),
            width="stretch",
            height=420,
        )
        csv_iod = tbl.to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇ Download DMI CSV",
            data=csv_iod,
            file_name="iod_dmi_data.csv",
            mime="text/csv",
        )


def render_mjo():  # noqa: C901
    """Madden-Julian Oscillation (RMM index) explorer page."""
    import numpy as _np

    # ── Phase metadata ────────────────────────────────────────────────────────
    # Phase sectors: counterclockwise in RMM1-RMM2 space starting at 180°.
    # Boundaries at 45° intervals; center angles and geographic labels below.
    _PHASE_GEO = {
        1: "W. Hemisphere\n& Africa",
        2: "Indian Ocean\n(SE Aus warming↑)",
        3: "Indian Ocean\n(SE Aus heat risk↑)",
        4: "Maritime\nContinent (rain↑ N. Aus)",
        5: "Maritime\nContinent (heavy rain↑ SE Qld)",
        6: "W. Pacific\n(rain↑ NE Aus)",
        7: "W. Pacific\n(cold nights↑ Qld)",
        8: "W. Hemisphere\n(SE Aus warming↑)",
    }
    # Centre angle of each phase sector (degrees, counterclockwise from +x axis)
    _PHASE_ANGLE = {1: 202.5, 2: 247.5, 3: 292.5, 4: 337.5,
                    5: 22.5,  6: 67.5,  7: 112.5, 8: 157.5}
    # 8 visually distinct colours — matches BoM colour convention roughly
    _PHASE_CLR = {
        1: "#7B4F9E", 2: "#C466AA", 3: "#D63B3B", 4: "#E07830",
        5: "#C8B400", 6: "#4BA84B", 7: "#2A9AD6", 8: "#3B5FC0",
    }

    st.title("MJO — Madden-Julian Oscillation")
    st.markdown(
        "The **Real-time Multivariate MJO (RMM) index** tracks the dominant mode of tropical "
        "intraseasonal variability: a slow-moving pulse of organised convection that circles the globe "
        "every **30–90 days**. RMM1 and RMM2 are the leading principal components of combined outgoing "
        "longwave radiation (OLR), 850 hPa and 200 hPa wind anomalies. "
        "**Amplitude ≥ 1.0** = active MJO event; **Phase 1–8** tracks the longitude of peak convection.  \n\n"
        "Recent research documents significant Australian climate impacts across all phases and seasons "
        "([Marshall et al. 2023](https://doi.org/10.1175/JCLI-D-22-0413.1); "
        "[Dao et al. 2025](https://doi.org/10.1002/qj.4995)). "
        "Key signals include: **Phase 3** spring — extreme heat risk doubles over Victoria/Tasmania; "
        "**Phase 7** winter — cold nights double over northeast Queensland; "
        "**Phases 4–6** — enhanced rainfall and heavy convective events over NE Australia. "
        "See the **Climate Science** page for detailed phase-by-phase literature findings."
    )
    st.caption(
        "RMM index: [Wheeler & Hendon (2004)](https://doi.org/10.1175/1520-0493(2004)132%3C1917:AARMMI%3E2.0.CO;2) · "
        "[Australian Bureau of Meteorology](http://www.bom.gov.au/climate/mjo/) (product IDCKGEM000) · "
        "[raw data](http://www.bom.gov.au/clim_data/IDCKGEM000/rmm.74toRealtime.txt) · updated daily."
    )

    try:
        mjo = fetch_mjo_data()
    except Exception as exc:
        st.error(f"Could not load MJO data: {exc}")
        st.info("The app tried to fetch from BoM and found no local cache. "
                "Check your internet connection and reload.")
        return

    latest = mjo.iloc[-1]
    active = latest["amplitude"] >= 1.0
    phase_lbl = f"Phase {int(latest['phase'])} — {_PHASE_GEO.get(int(latest['phase']), '').replace(chr(10), ' ')}"

    m1, m2, m3, m4 = st.columns(4)
    m1.metric(f"Latest ({latest['date'].strftime('%d %b %Y')})",
              f"Amp {latest['amplitude']:.2f}",
              delta="Active" if active else "Weak (<1.0)",
              delta_color="normal" if active else "off")
    m2.metric("Current phase", phase_lbl if active else "Weak / indeterminate")
    m3.metric("Active MJO days on record",
              f"{(mjo['amplitude'] >= 1.0).sum():,}",
              help="Days with RMM amplitude ≥ 1.0 (active MJO event)")
    m4.metric("Record span", f"{mjo['year'].min()}–{mjo['year'].max()}")

    tab_phase, tab_amp, tab_drfa, tab_data = st.tabs(
        ["🌀 Phase Diagram", "📊 Amplitude & Phase", "🌿 MJO × DRFA", "📋 Data"]
    )

    # ── TAB 1: PHASE-SPACE DIAGRAM ────────────────────────────────────────────
    with tab_phase:
        st.subheader("RMM Phase-Space Diagram (Wheeler-Hendon)")
        st.markdown(
            "RMM1 (x-axis) and RMM2 (y-axis) define an 8-sector phase space — the standard visualisation "
            "used by BoM, NOAA CPC, and the ECMWF. The **unit circle** marks the amplitude = 1.0 "
            "active-MJO threshold; points inside the circle indicate a weak or absent MJO. "
            "Phases increase **counterclockwise**, tracking the eastward propagation of the convective envelope."
        )

        days_back = st.slider("Trajectory length (days)", 14, 90, 40, step=7,
                              key="mjo_days_back")
        recent = mjo.tail(days_back).copy().reset_index(drop=True)
        n = len(recent)

        fig_phase = go.Figure()

        # Boundary lines at 45° intervals (8 spokes)
        R_spoke = 3.2
        for ang_deg in range(0, 360, 45):
            rad = _np.radians(ang_deg)
            fig_phase.add_trace(go.Scatter(
                x=[0, R_spoke * _np.cos(rad)],
                y=[0, R_spoke * _np.sin(rad)],
                mode="lines",
                line=dict(color="lightgrey", width=0.8, dash="dot"),
                showlegend=False, hoverinfo="skip",
            ))

        # Unit circle
        theta = _np.linspace(0, 2 * _np.pi, 300)
        fig_phase.add_trace(go.Scatter(
            x=_np.cos(theta), y=_np.sin(theta),
            mode="lines",
            line=dict(color="black", width=1.5),
            name="Amplitude = 1.0",
            hoverinfo="skip",
        ))

        # Phase sector labels (number + geography)
        for ph, ang_deg in _PHASE_ANGLE.items():
            rad = _np.radians(ang_deg)
            r_lbl = 2.7
            geo = _PHASE_GEO[ph].replace("\n", "<br>")
            fig_phase.add_annotation(
                x=r_lbl * _np.cos(rad), y=r_lbl * _np.sin(rad),
                text=f"<b>{ph}</b><br><span style='font-size:11px;color:{_PHASE_CLR[ph]}'>{geo}</span>",
                showarrow=False, font=dict(size=13), align="center",
            )

        # Trajectory — colour gradient from light (old) to dark (recent)
        if n > 1:
            for i in range(n - 1):
                frac = i / (n - 1)
                # Grey (#aaaaaa) → deep charcoal (#222222)
                v = int(170 - frac * 138)
                colour = f"rgb({v},{v},{v})"
                lw = 0.8 + frac * 1.6          # line grows thicker toward present
                fig_phase.add_trace(go.Scatter(
                    x=recent["rmm1"].iloc[i:i + 2].tolist(),
                    y=recent["rmm2"].iloc[i:i + 2].tolist(),
                    mode="lines",
                    line=dict(color=colour, width=lw),
                    showlegend=False,
                    hovertemplate=(
                        f"{recent['date'].iloc[i].strftime('%d %b %Y')}<br>"
                        "RMM1: %{x:.2f}<br>RMM2: %{y:.2f}<extra></extra>"
                    ),
                ))

        # Phase-coloured dots along trajectory
        for ph in range(1, 9):
            seg = recent[recent["phase"] == ph]
            if seg.empty:
                continue
            fig_phase.add_trace(go.Scatter(
                x=seg["rmm1"], y=seg["rmm2"],
                mode="markers",
                marker=dict(size=5, color=_PHASE_CLR[ph], opacity=0.7),
                name=f"Ph {ph}",
                hovertemplate=(
                    "%{customdata}<br>RMM1: %{x:.2f}<br>RMM2: %{y:.2f}"
                    f"<br>Phase {ph}<extra></extra>"
                ),
                customdata=seg["date"].dt.strftime("%d %b %Y"),
            ))

        # Latest point as star
        fig_phase.add_trace(go.Scatter(
            x=[latest["rmm1"]], y=[latest["rmm2"]],
            mode="markers",
            marker=dict(size=15, color="crimson", symbol="star",
                        line=dict(color="darkred", width=1)),
            name=f"Latest ({latest['date'].strftime('%d %b %Y')})",
            hovertemplate=(
                f"{latest['date'].strftime('%d %b %Y')}<br>"
                f"RMM1: {latest['rmm1']:.2f}<br>RMM2: {latest['rmm2']:.2f}<br>"
                f"Phase {int(latest['phase'])}<br>Amplitude: {latest['amplitude']:.2f}"
                "<extra></extra>"
            ),
        ))

        fig_phase.update_layout(
            xaxis=dict(title="RMM1", range=[-3.5, 3.5],
                       zeroline=True, zerolinewidth=1, zerolinecolor="grey"),
            yaxis=dict(title="RMM2", range=[-3.5, 3.5],
                       zeroline=True, zerolinewidth=1, zerolinecolor="grey",
                       scaleanchor="x"),
            plot_bgcolor="white",
            height=580,
            legend=dict(orientation="h", y=-0.12),
        )
        st.plotly_chart(fig_phase, width="stretch")
        st.caption(
            "Trajectory fades from light grey (oldest) to dark (most recent). "
            "Coloured dots show phase classification; ★ = latest observation. "
            "Phases increase counterclockwise. Format follows BoM / NOAA CPC convention."
        )

    # ── TAB 2: AMPLITUDE & PHASE TIME SERIES ─────────────────────────────────
    with tab_amp:
        st.subheader("RMM Amplitude and Phase Over Time")
        st.markdown(
            "Bar height = daily RMM amplitude; bar colour = MJO phase (1–8). "
            "The dashed line marks the **amplitude 1.0 active-MJO threshold**. "
            "Switch to **Monthly mean** to reduce visual noise over longer periods."
        )

        yr_min, yr_max = int(mjo["year"].min()), int(mjo["year"].max())
        a_start, a_end = st.slider(
            "Year range", yr_min, yr_max, (max(yr_min, 2015), yr_max),
            key="mjo_amp_yr",
        )
        agg_mode = st.radio("Aggregation", ["Daily", "Monthly mean"],
                            horizontal=True, key="mjo_agg")

        view_amp = mjo[(mjo["year"] >= a_start) & (mjo["year"] <= a_end)].copy()

        if agg_mode == "Monthly mean":
            view_amp["_ym"] = view_amp["date"].dt.to_period("M").dt.to_timestamp()
            view_amp = (
                view_amp.groupby("_ym")
                .agg(amplitude=("amplitude", "mean"),
                     phase=("phase", lambda x: x.mode().iat[0]))
                .reset_index()
                .rename(columns={"_ym": "date"})
            )

        fig_amp = go.Figure()
        for ph in range(1, 9):
            seg = view_amp[view_amp["phase"] == ph]
            if seg.empty:
                continue
            fig_amp.add_trace(go.Bar(
                x=seg["date"], y=seg["amplitude"],
                marker_color=_PHASE_CLR[ph],
                name=f"Phase {ph} — {_PHASE_GEO[ph].replace(chr(10), ' ')}",
                hovertemplate=(
                    "%{x|%b %Y}<br>Amplitude: %{y:.2f}"
                    f"<br>Phase {ph}<extra></extra>"
                ),
            ))

        fig_amp.add_hline(y=1.0, line=dict(color="black", dash="dash", width=1.5),
                          annotation_text="Active MJO (≥ 1.0)",
                          annotation_position="top right")
        fig_amp.update_layout(
            barmode="overlay",
            xaxis_title="Date",
            yaxis_title="RMM Amplitude",
            legend=dict(orientation="h", y=-0.25),
            height=420,
            hovermode="x unified",
            bargap=0,
        )
        st.plotly_chart(fig_amp, width="stretch")

    # ── TAB 3: AUSTRALIAN IMPACTS ─────────────────────────────────────────────
    # ── TAB 3: MJO × DRFA ────────────────────────────────────────────────────
    with tab_drfa:
        st.subheader("DRFA Event Activations by MJO Phase")
        st.markdown(
            "Each DRFA activation is matched to the **dominant MJO phase** and **mean amplitude** "
            "for the calendar month of its start date. "
            "Only post-2006 DRFA events are included (MJO data extends back to 1974)."
        )
        st.info(
            "**Interpretation note:** The MJO–disaster link is indirect and operates on a 30–90 day "
            "timescale, making it weaker than ENSO, IOD, or SAM for explaining annual or seasonal "
            "disaster patterns. The strongest Australian signals are: **phases 4–6** → enhanced monsoon "
            "rainfall and tropical cyclone activity in northern Australia; **phases 7–8** → suppressed "
            "convection and elevated heat risk in the southeast. "
            "Monthly-resolution matching introduces timing noise — treat these distributions as "
            "exploratory rather than causal.",
            icon="ℹ️",
        )

        # Aggregate MJO to monthly: dominant phase, mean amplitude
        mjo_monthly = (
            mjo.groupby(mjo["date"].dt.to_period("M").dt.to_timestamp())
            .agg(mean_amplitude=("amplitude", "mean"),
                 dominant_phase=("phase", lambda x: x.mode().iat[0]))
            .reset_index()
            .rename(columns={"date": "date"})
        )

        try:
            drfa = load_drfa_events()
        except Exception as exc:
            st.error(f"Could not load DRFA events: {exc}")
        else:
            drfa_m = drfa.dropna(subset=["start"]).copy()
            drfa_m["date"] = drfa_m["start"].dt.to_period("M").dt.to_timestamp()
            drfa_m = drfa_m.merge(mjo_monthly, on="date", how="left")

            d1, d2 = st.columns(2)
            with d1:
                phase_counts = (
                    drfa_m["dominant_phase"]
                    .value_counts()
                    .reindex(range(1, 9), fill_value=0)
                    .reset_index()
                )
                phase_counts.columns = ["Phase", "DRFA Events"]
                phase_counts["colour"] = phase_counts["Phase"].map(_PHASE_CLR)
                phase_counts["label"] = phase_counts["Phase"].apply(
                    lambda p: f"Ph {p}: {_PHASE_GEO[p].replace(chr(10), ' ')}"
                )
                fig_ph = go.Figure(go.Bar(
                    x=phase_counts["label"], y=phase_counts["DRFA Events"],
                    marker_color=phase_counts["colour"],
                    hovertemplate="%{x}<br>Events: %{y}<extra></extra>",
                ))
                fig_ph.update_layout(
                    title="DRFA Events by MJO Phase",
                    xaxis_tickangle=-30, showlegend=False, height=380,
                )
                st.plotly_chart(fig_ph, width="stretch")

            with d2:
                active_split = drfa_m.copy()
                active_split["MJO State"] = active_split["mean_amplitude"].apply(
                    lambda a: "Active MJO (≥1.0)" if a >= 1.0 else "Weak MJO (<1.0)"
                )
                active_counts = (
                    active_split["MJO State"]
                    .value_counts()
                    .reindex(["Active MJO (≥1.0)", "Weak MJO (<1.0)"], fill_value=0)
                    .reset_index()
                )
                active_counts.columns = ["MJO State", "DRFA Events"]
                fig_active = px.pie(
                    active_counts, names="MJO State", values="DRFA Events",
                    title="DRFA Events: Active vs Weak MJO months",
                    color="MJO State",
                    color_discrete_map={
                        "Active MJO (≥1.0)": "#d62728",
                        "Weak MJO (<1.0)":   "#aaaaaa",
                    },
                )
                fig_active.update_layout(height=380)
                st.plotly_chart(fig_active, width="stretch")

            # Hazard type by MJO phase
            haz_phase = (
                drfa_m.dropna(subset=["dominant_phase"])
                .groupby(["dominant_phase", "hazard_group"])
                .size()
                .reset_index(name="events")
            )
            haz_phase["phase_label"] = haz_phase["dominant_phase"].apply(
                lambda p: f"Ph {p}"
            )
            fig_haz = px.bar(
                haz_phase, x="phase_label", y="events",
                color="hazard_group", barmode="stack",
                title="Hazard Type by MJO Phase",
                labels={"phase_label": "MJO Phase", "events": "DRFA Events",
                        "hazard_group": "Hazard"},
                color_discrete_sequence=px.colors.qualitative.Set2,
                category_orders={"phase_label": [f"Ph {p}" for p in range(1, 9)]},
            )
            fig_haz.update_layout(height=380)
            st.plotly_chart(fig_haz, width="stretch")

    # ── TAB 4: DATA ───────────────────────────────────────────────────────────
    with tab_data:
        st.subheader("Raw RMM Data")

        dc1, dc2, dc3 = st.columns(3)
        with dc1:
            ph_filter = st.multiselect(
                "Filter by phase", list(range(1, 9)),
                default=list(range(1, 9)), key="mjo_ph_filter",
            )
        with dc2:
            active_only = st.checkbox("Active MJO only (amplitude ≥ 1.0)",
                                      key="mjo_active_only")
        with dc3:
            yr_d1, yr_d2 = st.slider(
                "Year range",
                int(mjo["year"].min()), int(mjo["year"].max()),
                (max(int(mjo["year"].min()), 2010), int(mjo["year"].max())),
                key="mjo_data_yr",
            )

        tbl = mjo[
            mjo["phase"].isin(ph_filter) &
            (mjo["year"] >= yr_d1) &
            (mjo["year"] <= yr_d2)
        ].copy()
        if active_only:
            tbl = tbl[tbl["amplitude"] >= 1.0]

        tbl_disp = tbl[["date", "rmm1", "rmm2", "phase", "amplitude"]].copy()
        tbl_disp["date"]      = tbl_disp["date"].dt.strftime("%d %b %Y")
        tbl_disp["rmm1"]      = tbl_disp["rmm1"].round(3)
        tbl_disp["rmm2"]      = tbl_disp["rmm2"].round(3)
        tbl_disp["amplitude"] = tbl_disp["amplitude"].round(3)
        tbl_disp["geo"]       = tbl_disp["phase"].map(
            lambda p: _PHASE_GEO.get(p, "").replace("\n", " ")
        )
        tbl_disp = tbl_disp.rename(columns={
            "date": "Date", "rmm1": "RMM1", "rmm2": "RMM2",
            "phase": "Phase", "amplitude": "Amplitude", "geo": "Geography",
        })

        st.dataframe(tbl_disp.reset_index(drop=True),
                     width="stretch", height=420)
        csv_mjo = tbl_disp.to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download RMM CSV", data=csv_mjo,
                           file_name="mjo_rmm_data.csv", mime="text/csv")
# ── home page ─────────────────────────────────────────────────────────────────

def render_home():
    st.title("Australian Disaster Data Explorer")
    st.markdown(
        "**PhD Research Tool** · Samuel Marcus · Monash University  \n"
        "*Building Adaptive Capacity: Quantifying Emergency Resource Demands "
        "under Intensifying Weather Extremes*"
    )
    st.markdown("---")

    st.markdown(
        "This tool provides interactive access to five source datasets covering "
        "Australian natural disasters from 1727 to the present, integrated and "
        "analysed through the lens of compound weather extremes and emergency "
        "management capacity constraints."
    )

    # ── Dataset coverage cards ────────────────────────────────────────────────
    st.subheader("Dataset Coverage")

    def _card(col, icon, title, provider, coverage, records, colour, page=None):
        with col:
            st.markdown(
                f"""
<div style="border-left: 4px solid {colour}; padding: 10px 14px; border-radius: 4px;
            background: #f9f9f9; margin-bottom: 4px;">
<span style="font-size:1.4rem">{icon}</span>
<strong style="font-size:1.0rem">&nbsp;{title}</strong><br>
<span style="color:#555; font-size:0.82rem">{provider}</span><br>
<span style="font-size:0.85rem">📅 {coverage} &nbsp;·&nbsp; {records}</span>
</div>""",
                unsafe_allow_html=True,
            )
            if page is not None:
                if st.button("Open →", key=f"nav__{title}", width="stretch"):
                    st.switch_page(page)

    st.markdown("##### Source Datasets")
    r1c1, r1c2, r1c3 = st.columns(3)
    r2c1, r2c2, r2c3 = st.columns(3)
    _card(r1c1, "📋", "AIDR Event Catalogue",
          "AIDR / Attorney-General's Dept",
          "1727 – 2023", "946 events", "#4e79a7", _PAGE_AIDR)
    _card(r1c2, "💰", "ICA Catastrophes",
          "Insurance Council of Australia",
          "1967 – 2026", "744 events · A$169B NL", "#f28e2b", _PAGE_ICA)
    _card(r1c3, "🗂", "DRFA Activations",
          "NEMA (National Emergency Management Agency)",
          "2006 – 2026", "5,967 LGA–event rows · 809 events", "#e15759", _PAGE_DRFA_ACT)
    _card(r2c1, "💵", "DRFA Payments",
          "Services Australia / NEMA",
          "2009 – 2026", "A$6.95B paid across 44 events", "#76b7b2", _PAGE_DRFA_PAY)
    _card(r2c2, "🌍", "EM-DAT",
          "CRED, UCLouvain",
          "1939 – present", "224 Australian events", "#59a14f", _PAGE_EMDAT)

    st.markdown("##### Climate Indices")
    c1, c2, c3, c4 = st.columns(4)
    _card(c1, "🌊", "ENSO / ONI",
          "NOAA Climate Prediction Centre",
          "1950 – present", "Monthly Niño 3.4 index", "#9c755f", _PAGE_ONI)
    _card(c2, "🌬️", "SAM Index",
          "British Antarctic Survey",
          "1957 – present", "Monthly SAM index", "#b07aa1", _PAGE_SAM)
    _card(c3, "🌡️", "IOD / DMI",
          "NOAA PSL / HadISST1.1",
          "1870 – present", "Monthly DMI index", "#ff9da7", _PAGE_IOD)
    _card(c4, "🌀", "MJO / RMM",
          "Bureau of Meteorology",
          "1974 – present", "Daily RMM1 / RMM2 indices", "#9edae5", _PAGE_MJO)

    st.markdown("##### EM Capacity")
    e1, e2 = st.columns(2)
    _card(e1, "🛡️", "National Capability (AFAC)",
          "AFAC / 2023 National Capability Statement",
          "2023 snapshot", "State resource allocations & deployable teams", "#17becf", _PAGE_AFAC)
    _card(e2, "📍", "State Capability Profiles",
          "AFAC · DRFA",
          "2006 – 2026", "DRFA burden vs AFAC capability share per state", "#aec7e8", _PAGE_STATE_CAP)

    st.markdown("---")

    # ── Section descriptions ──────────────────────────────────────────────────
    st.subheader("Tool Sections")

    sc1, sc2 = st.columns(2)
    with sc1:
        st.markdown("""
**📂 Source Datasets**
Browse and filter the five source datasets in their original form.
Useful for data validation, exploratory queries, and export.

**🌦 Climate Data**
Four live climate indices — ENSO/ONI, SAM, IOD/DMI, and MJO/RMM —
fetched and cached from authoritative sources, plus a Climate Science
synthesis of compound driver interactions and historical case studies.
        """)
    with sc2:
        st.markdown("""
**🔗 Integrated Data**
Datasets derived by joining the source data:
DRFA activations merged with payment records,
and a spatial event map.

**📊 Research Analysis**
Compound disaster detection (Gissing et al. 2022 methodology)
and concurrent EM resource demand analysis.

**🛡️ EM Capacity**
AFAC 2023 National Capability Statement: national resource overview,
per-state DRFA burden vs capacity profiles, and a sweep-line stress
analysis of concurrent DRFA demand against the deployable pool.
        """)

    st.markdown("---")

    # ── Attribution ───────────────────────────────────────────────────────────
    st.markdown(
        "**Contact:** [samuel.marcus@monash.edu](mailto:samuel.marcus@monash.edu)  \n"
    )
    st.caption(
        "[AIDR/AGD](https://knowledge.aidr.org.au/resources/disaster-mapper/) · "
        "[Insurance Council of Australia](https://insurancecouncil.com.au/industry-members/data-hub/) · "
        "[NEMA / Services Australia](https://data.gov.au/data/dataset/drfa-activation-history-by-lga) · "
        "[EM-DAT (CRED, UCLouvain)](https://public.emdat.be/) · "
        "[NOAA CPC](https://www.cpc.ncep.noaa.gov/products/analysis_monitoring/ensostuff/ONI_v5.php) · "
        "[BAS SAM](https://legacy.bas.ac.uk/met/gjma/sam.html) · "
        "[NOAA PSL / HadISST1.1](https://psl.noaa.gov/data/timeseries/month/DMI/) · "
        "[BoM MJO](http://www.bom.gov.au/climate/mjo/) · "
        "[AFAC](https://www.afac.com.au/)"
    )


# ── entry point ───────────────────────────────────────────────────────────────

st.sidebar.markdown(
    "**Australian Disaster Data Explorer**  \n"
    "Samuel Marcus · Monash University"
)
st.sidebar.caption("PhD Research Tool · ARC 21st Century Weather")
def render_climate_science():
    """Climate Science — synthesis of Australian climate driver impacts and interactions."""
    st.title("Climate Science — Australian Driver Impacts")
    st.markdown(
        "Peer-reviewed synthesis of how ENSO, IOD, SAM, and MJO affect Australian weather extremes, "
        "and — critically — how they **interact** to amplify or suppress hazard risk. "
        "This page draws on the individual index data available in the Climate Data section."
    )

    tab_enso, tab_iod, tab_sam, tab_mjo, tab_compound, tab_cases, tab_refs = st.tabs([
        "🌊 ENSO", "🌡️ IOD", "🌬️ SAM", "🌀 MJO",
        "⚡ Compound Interactions", "📖 Historical Case Studies", "📚 References",
    ])

    # ── ENSO ─────────────────────────────────────────────────────────────────
    with tab_enso:
        st.subheader("ENSO — El Niño / Southern Oscillation")
        st.markdown(
            "ENSO is the dominant driver of interannual Australian climate variability. "
            "La Niña produces **stronger and more spatially extensive** impacts than El Niño — "
            "an asymmetry confirmed across rainfall, floods, droughts, cyclones, and fire weather "
            "([Taschetto et al. 2026](https://doi.org/10.1038/s43017-025-00747-x))."
        )

        st.markdown("#### Rainfall")
        st.markdown("""
| Finding | El Niño | La Niña | Source |
|---------|---------|---------|--------|
| Land area significantly affected (SON) | 56.6% | **78.6%** | Taschetto et al. 2026 |
| Eastern Australian rainfall variability explained | ~25% | ~25% | Taschetto et al. 2026 |
| Murray-Darling Basin winter–spring rainfall | −28% avg | — | Taschetto et al. 2026 |
| Eastern Australia spring–summer streamflow | reduced | **~2× above average** | Taschetto et al. 2026 |
| NE Australia rainfall anomaly (2010–11 event) | — | +84% | Taschetto et al. 2026 |
""")

        st.markdown("#### Hazard Risk")
        st.markdown("""
| Hazard | El Niño | La Niña | Source |
|--------|---------|---------|--------|
| Multi-year drought risk (QLD/NSW) | Precedes **~80%** of events; ~2× risk | reduced | Taschetto et al. 2026 |
| Major floods (eastern seaboard) | fewer | **~2× more frequent** | Taschetto et al. 2026 |
| Tropical cyclones (annual average) | **~8/yr** | **~12/yr** | Taschetto et al. 2026 |
| Area burned (Central Pacific El Niño) | **~519,000 km²** | — | Taschetto et al. 2026 |
| Fire weather days >FFDI 90th pctile (SON) | **>14 days** (with +IOD, −SAM) | — | Taschetto et al. 2026 |
""")

        st.info(
            "**CP vs EP El Niño:** Central Pacific (Modoki) events produce stronger and more spatially extensive "
            "rainfall deficits across eastern Australia than canonical Eastern Pacific events. "
            "CP events have become more frequent under modern climate conditions "
            "(Taschetto et al. 2026; Rensch et al. 2025).",
            icon="ℹ️",
        )

    # ── IOD ──────────────────────────────────────────────────────────────────
    with tab_iod:
        st.subheader("IOD — Indian Ocean Dipole")
        st.markdown(
            "The IOD is most active **June–November** (austral winter–spring). "
            "Its effects on Australian rainfall are tightly linked to drought and fire in positive phase "
            "and floods in negative phase — but its variability is partly driven by ENSO, "
            "which complicates attribution (Liguori et al. 2021)."
        )

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("##### Positive IOD (DMI ≥ +0.4 °C)")
            st.markdown("""
- **Every major southern Australian drought since 1889** coincided with positive or neutral IOD (Ummenhofer et al. 2009)
- Absence of negative IOD events during 1995–2009 "Big Dry" deprived SE Australia of normal rainfall quota
- Three consecutive positive IOD events 2006–2008 preceded Black Saturday
- 2019 extreme positive IOD amplified Black Summer 2019–20 fire conditions
- Extreme positive IOD events projected **~3× more frequent** this century (once per 6 yr vs once per 18 yr)
""")
        with col_b:
            st.markdown("##### Negative IOD (DMI ≤ −0.4 °C)")
            st.markdown("""
- Anomalously **wet conditions** in southern and eastern Australia
- 2010–11: Strong negative IOD + La Niña → Queensland's wettest month on record (Dec 2010: **209.45 mm**)
- Twice as many severe floods on eastern seaboard during La Niña + negative IOD years
- Victorian floods 2011 associated with strong negative IOD co-occurrence with La Niña
""")

        st.warning(
            "**Statistical caution — ENSO–IOD attribution** (Liguori et al. 2021):  \n"
            "Standard partial regression (statistically 'removing' IOD) **underestimates ENSO's true role** because IOD "
            "variability is partly energised by ENSO — removing ENSO reduces DMI variance by ~30%. "
            "Physical (model-based) ENSO removal produces substantially stronger ENSO precipitation patterns over SE Australia. "
            "Peak co-variability window: **June–October**. Robust attribution requires a multivariate framework.",
            icon="⚠️",
        )

    # ── SAM ──────────────────────────────────────────────────────────────────
    with tab_sam:
        st.subheader("SAM — Southern Annular Mode")
        st.markdown(
            "SAM explains **10–15% of weekly rainfall variance** in southern Australia — comparable to ENSO. "
            "Its effects are strongly seasonal and differ by region, with a long-term positive trend "
            "driven by stratospheric ozone depletion and greenhouse gas forcing "
            "([Hendon et al. 2007](https://doi.org/10.1175/JCLI4134.1); "
            "[Raut et al. 2014](https://doi.org/10.1175/JCLI-D-13-00773.1))."
        )

        st.markdown("#### Rainfall by Region and Season")
        st.markdown("""
| Region | Season | High (positive) SAM | Low (negative) SAM | Source |
|--------|--------|---------------------|---------------------|--------|
| SW Australia coastal | Winter (JJA) | **−20% July rainfall**; fewer westerly fronts | wetter | Raut et al. 2014 |
| SW Australia inland / semi-arid | Summer (DJF) | **+40–50% rainfall** via easterly troughs | — | Raut et al. 2014 |
| SE Australia | Winter | **Decreased** daily rainfall | increased | Hendon et al. 2007 |
| Southern east coast + Tasmania | Spring–Summer | **Increased** (upslope Tasman flow) | decreased | Hendon et al. 2007 |
| Southern Australia (broad) | Spring–Summer | Significant events **up to 2× more likely** | — | Hendon et al. 2007 |
""")

        st.markdown("#### Temperature Extremes")
        st.markdown("""
| Location | Season | High SAM | Low SAM | Source |
|----------|--------|----------|---------|--------|
| South-central Australia | Summer | Heat >90th pctile **~half as likely** | Heat **~2× more likely** | Hendon et al. 2007 |
| Kalgoorlie max temperature | DJF | **−2.9 °C** anomaly vs low SAM | — | Hendon et al. 2007 |
| Sydney top-decile temperature days | DJF | 22 of 333 days | **53 of 348 days** | Hendon et al. 2007 |
""")

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("##### ENSO × SAM interaction (SW Australia)")
            st.markdown("""
- **La Niña + Positive SAM:** coastal winter rainfall **doubles** vs neutral
- **El Niño + Negative SAM:** compound drying in winter; least favourable for coastal rainfall
""")
        with col_b:
            st.markdown("##### Long-term SAM trend")
            st.markdown("""
- Positive trend: **+0.54 std devs** in DJF 1979–2005
- Driven by ozone depletion + GHG forcing; expected to continue
- Contributed to SE Australia summer rainfall increase and SW Australia winter drying
""")

    # ── MJO ──────────────────────────────────────────────────────────────────
    with tab_mjo:
        st.subheader("MJO — Madden-Julian Oscillation")
        st.markdown(
            "The MJO operates on **30–90 day** timescales and provides the best sub-seasonal forecast skill "
            "for Australian weather extremes. Phase and season determine the signal direction; "
            "ENSO can substantially amplify or dampen MJO teleconnections."
        )

        st.markdown("#### Temperature Impacts (Marshall et al. 2023, *J. Climate*)")
        st.markdown("""
| Phase | Season | Key Signal | Magnitude | Mechanism |
|-------|--------|------------|-----------|-----------|
| **2** | Spring (SON) | Daytime warming SE Australia | up to **+1.4 °C** | Anticyclonic flow, suppressed rainfall |
| **3** | Spring (SON) | **Strongest heat signal** — Vic, Tas, SA | up to **+1.8 °C**; extreme heat **2× more likely** | Rossby wave train → anticyclonic flow; warm northerlies |
| **3** | Spring (SON) | *El Niño compound:* SE Australia | extreme heat **>3× more likely** | ENSO amplifies anticyclonic warming; peak 3.5 °C anomaly |
| **5–6** | Summer (DJF) | Cooling, N. Australia | **−0.5 °C** | Enhanced convection → increased rainfall → radiative cooling |
| **7** | Winter (JJA) | **Cold nights**, NE Queensland | down to **−2 °C**; extremes **>2× more likely** | Suppressed convection over Indian Ocean → westerly/SW cool air |
| **8** | Summer (DJF) | Warming, SE Australia | up to **+1 °C** | Broad anticyclonic anomaly; warm interior air advected SE |
""")

        st.markdown("#### NE Australia Rainfall (Dao et al. 2025, *QJRMS*)")
        st.markdown("""
| Phase | Brisbane | Townsville | Notes |
|-------|----------|------------|-------|
| **1–3, 8** | Suppressed | Suppressed | Suppressed convection; reduced moisture |
| **3** | Suppressed | **Enhanced +16%** | Orographic enhancement despite suppressed large-scale phase |
| **4–6** | Enhanced | Enhanced | Active convection; increased moisture transport |
| **5** | Heavy rain **2× more likely** (ocean) | Moderate | Highest CAPE (330 vs 255 J/kg); deep northerly moisture |
| **6** | Frequency up **~10%** | Enhanced | ~43% of days exceed upper tercile threshold |
| **7** | Decreased | Heavy rain **+4–6% above baseline** | Convective rainfall strongest over Townsville land |
""")

        st.info(
            "**Flood risk:** Phase 3 at Townsville is notable — large-scale models predict suppression but "
            "topographic enhancement can still produce heavy rainfall. Relevant to the 2019 Townsville "
            "and 2023 Cairns floods. MJO–ENSO compound (Phase 3 + El Niño) produces the most extreme "
            "spring heat events across SE Australia.",
            icon="🌊",
        )

    # ── COMPOUND INTERACTIONS ────────────────────────────────────────────────
    with tab_compound:
        st.subheader("Compound Climate Driver Interactions")
        st.markdown(
            "Single-driver analyses understate risk. When multiple drivers align, hazard impacts are "
            "**multiplicative, not additive** — the interaction of modes is more important than any single mode. "
            "The most extreme Australian fire, drought, and flood events on record all involved two or more "
            "co-occurring drivers."
        )

        st.markdown("#### Driver Interaction Matrix")
        st.markdown("""
| ENSO | IOD | SAM | Hazard outcome | Example |
|------|-----|-----|----------------|---------|
| **El Niño** | Positive | Negative | **Maximum fire/drought risk** — >14 FFDI days SON, ~519,000 km² burned | Black Summer 2019–20 |
| **El Niño** | Positive | Neutral | Severe drying SE/SW Australia; drought amplified | 2006, 2015 |
| **El Niño** | Neutral | Negative | Drying; moderate-to-severe fire risk | 2002 |
| **La Niña** | Negative | Positive | **Maximum flood risk** — widespread rainfall, record streamflow | 2010–11, 2021–22 |
| **La Niña** | Negative | Neutral | Above-average rainfall; elevated flood risk | Most La Niña events |
| **La Niña** | Positive | Positive | **Opposing** — positive IOD suppresses La Niña wetting in SE Australia | 2007–09 |
| **Neutral** | Positive | Negative | Drought/fire risk without ENSO forcing | Possible in JJA–SON |
""")

        st.markdown("#### Key Quantitative Compound Findings")
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("""
**Fire and drought:**
- El Niño + positive IOD + negative SAM = only compound producing >14 fire weather days in SON (FFDI >90th pctile) — Taschetto et al. 2026
- Compound drought-heatwave (CDHW) season: ~1 in 2 seasons during strong El Niño + positive IOD over NE Australia
- 2018–19: only second occurrence since 1911 of simultaneous CP El Niño + extreme positive IOD — directly preconditioned Black Summer
- Black Summer 2019–20 FFDI: 200-year return period event (Abram et al. 2021)
""")
        with col_b:
            st.markdown("""
**Floods and wet extremes:**
- La Niña + negative IOD + positive SAM: Australia's second wettest spring in 123 years of records (2010)
- 2010–11 two La Niña events combined with strong negative IOD: wettest 24-month period on record
- Multi-year La Niña: increased flood risk as rainfall falls on already-saturated catchments
- ENSO and blocking together explain ~50% of rainfall variance in eastern Australia in spring

**Attribution caution:**
- Suppressing ENSO reduces DMI variance by ~30% — standard IOD partial regression underestimates ENSO contribution
- Robust attribution requires multivariate framework (Liguori et al. 2021)
""")

        st.info(
            "**Implication for emergency management:** The simultaneity problem — multiple states facing "
            "concurrent crises — is most acute during large-scale compound driver events (e.g., a synoptic "
            "pattern that simultaneously forces El Niño + positive IOD + negative SAM conditions across "
            "multiple regions). These events are expected to increase in frequency under GHG forcing.",
            icon="🚨",
        )

    # ── HISTORICAL CASE STUDIES ───────────────────────────────────────────────
    with tab_cases:
        st.subheader("Historical Compound Event Case Studies")

        with st.expander("🔥 Black Summer 2019–20 — El Niño + extreme positive IOD + record negative SAM", expanded=True):
            st.markdown("""
**Climate preconditions:**
- 2018–19: Central Pacific El Niño (CP ENSO) — only second co-occurrence with extreme positive IOD since 1911
- 2019: Record-strong positive IOD — contributed to rainfall deficit across southern/eastern Australia
- Spring 2019: Record negative SAM anomaly — amplified drying and heat; reduced frontal rainfall

**Meteorological outcomes:**
- Spring 2019: 4th-lowest September rainfall on record nationally; Tasmania's driest spring on record
- October–December 2019: Highest mean temperatures on record nationally
- FFDI: record-high values across all states and territories; Murrurundi Gap (NSW Hunter): FFDI 174

**Rarity:** FFDI conditions in 2019–20 represent a ~200-year return period event (Abram et al. 2021).
Antarctic ice-core fire-weather reconstruction confirms 2019–20 patterns are extremely rare over the past 2,000 years.

**EM implications:** Victoria, NSW, QLD, SA, WA, Tasmania — simultaneous fire crises. AFAC managed three
concurrent state deployments in the January–March 2024 follow-on season (interview: Emma Conway, AFAC 2026).
""")

        with st.expander("🌊 2010–11 Queensland Floods — La Niña + strong negative IOD + positive SAM"):
            st.markdown("""
**Climate preconditions:**
- 2010: One of the strongest La Niña events on record (strongest since 1973)
- October 2010: Strong negative IOD developed, coupling with La Niña
- Positive SAM during spring/summer amplified moisture delivery to east coast

**Meteorological outcomes:**
- December 2010: Queensland's wettest month on record — **209.45 mm** (previous record: 200.1 mm in 1975)
- 2010 and 2011: Australia's wettest 24-month period on record
- NE Australia rainfall anomaly: +84% above average
- Flooding: widespread across SE Queensland, N/W Victoria, NSW, NW WA, E Tasmania (Sep 2010–Mar 2011)

**EM implications:** Largest insured natural disaster losses in Australian history (at the time).
Demonstrated that negative IOD + La Niña + positive SAM creates maximum east coast flood risk.
""")

        with st.expander("🌵 The 'Big Dry' 1995–2009 — Absence of negative IOD + prolonged positive phase"):
            st.markdown("""
**Climate preconditions:**
- 1995–2009: Complete absence of negative IOD events — unprecedented in the modern record
- Multiple El Niño events within this window (1997–98, 2002–03, 2006)
- SAM trending positive (reduced frontal systems reaching SE Australia)

**Meteorological outcomes:**
- Persistent rainfall deficits across southeastern Australia (Victoria, SA, SW WA)
- Murray-Darling Basin: extended drought with severe water security impacts
- Three consecutive positive IOD events 2006–2008 severely amplified conditions before Black Saturday

**Research finding (Ummenhofer et al. 2009):** The absence of negative IOD events — not just the
presence of positive ones — deprived SE Australia of its normal annual rainfall quota across 14 years.
This is the key insight: it's the *absence of the wetting phase*, not just the presence of the drying phase.
""")

    # ── REFERENCES ────────────────────────────────────────────────────────────
    with tab_refs:
        st.subheader("Key References")
        st.markdown("""
| Authors | Year | Title | Journal | DOI |
|---------|------|-------|---------|-----|
| Taschetto et al. | 2026 | Climate impacts of the El Niño–Southern Oscillation on Australia | *Nature Reviews Earth & Environment* | [10.1038/s43017-025-00747-x](https://doi.org/10.1038/s43017-025-00747-x) |
| Liguori et al. | 2021 | Revisiting ENSO and IOD contributions to Australian precipitation | *Geophysical Research Letters* | verify DOI |
| Hendon et al. | 2007 | Australian rainfall and surface temperature variations associated with the Southern Annular Mode | *Journal of Climate* | [10.1175/JCLI4134.1](https://doi.org/10.1175/JCLI4134.1) |
| Raut et al. | 2014 | A new perspective on the western Australian rainfall decline | *Journal of Climate* | [10.1175/JCLI-D-13-00773.1](https://doi.org/10.1175/JCLI-D-13-00773.1) |
| Marshall et al. | 2023 | MJO impacts on Australian temperatures and extremes | *Journal of Climate* | [10.1175/JCLI-D-22-0413.1](https://doi.org/10.1175/JCLI-D-22-0413.1) |
| Dao et al. | 2025 | Modulations of local rainfall in NE Australia associated with the MJO | *QJRMS* | [10.1002/qj.4995](https://doi.org/10.1002/qj.4995) |
| Ummenhofer et al. | 2009 | What causes southeast Australia's worst droughts? | *Geophysical Research Letters* | verify DOI |
| Cai et al. | 2009 | IOD response to increased CO₂ | *Geophysical Research Letters* | verify DOI |
| Rensch et al. | 2025 | El Niño and La Niña influence on Australian rainfall | *Geophysical Research Letters* | verify DOI |
| Abram et al. | 2021 | Climate change and variability links to large forest fires in SE Australia | *Communications Earth & Environment* | verify DOI |
""")
        st.caption(
            "See also: Zscheischler et al. (2020) *Nature Reviews Earth & Environment* for the compound hazard typology "
            "framework used throughout this project (verify DOI before citing)."
        )


# ═══════════════════════════════════════════════════════════════════════════════
# AFAC EM Capacity pages
# ═══════════════════════════════════════════════════════════════════════════════

def render_em_capability():  # noqa: C901
    """AFAC 2023 National Capability Statement — interactive reference."""
    st.title("National EM Capability (AFAC 2023)")
    st.caption(
        "AFAC National Statement of Capability for Fire and Emergency Services — 2023 Edition. "
        "Resource inventory and interstate-deployable team counts (48-hour pool) across 8 domains "
        "and all Australian jurisdictions."
    )

    with st.expander("About this data", expanded=False):
        st.markdown("""
**Source:** AFAC (Australasian Fire and Emergency Service Authorities Council), October 2023.
**Coverage:** All Australian state/territory fire and emergency service agencies.
**Domains:** Firefighting (Bushfire & Other), Search & Rescue, Severe Weather Response,
HAZMAT, Damage & Impact Assessment, Incident Management, Aviation.
**Deployable capability:** Teams agencies indicate could be deployed interstate within 48 hours under
normal operational conditions. Actual availability depends on real-time commitments.
**Limitations:** 2023 snapshot only — does not capture inter-year capability changes or real-time status.
Data sourced from the accompanying Excel spreadsheet; Aviation is inventory only (no deployable table).
        """)

    df = load_afac_capability()
    if df.empty:
        st.error("AFAC capability data file not found. Expected: 2023-national-capability-statement-data.xlsx")
        return

    STATES = _AFAC_STATES

    # ── Filters ───────────────────────────────────────────────────────────────
    fc1, fc2 = st.columns([3, 2])
    with fc1:
        domains    = sorted(df["domain"].unique())
        sel_domain = st.multiselect("Capability domain", domains,
                                    placeholder="All domains", key="emc_domain")
    with fc2:
        show_dep = st.radio("Records", ["All", "Deployable teams only"],
                            horizontal=True, key="emc_dep")

    filt = df.copy()
    if sel_domain:
        filt = filt[filt["domain"].isin(sel_domain)]
    if show_dep == "Deployable teams only":
        filt = filt[filt["deployable"]]

    # ── Summary metrics ───────────────────────────────────────────────────────
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Domains", len(df["domain"].unique()))
    m2.metric("Resource types (inventory)", int((~df["deployable"]).sum()))
    m3.metric("Deployable team types", int(df["deployable"].sum()))
    m4.metric("Total deployable units (national)", int(df[df["deployable"]]["national"].sum()))

    tab1, tab2, tab3 = st.tabs(["📋 Resource Table", "🗺️ State Allocation", "🚀 Deployable Teams"])

    # ── TAB 1: Resource table ─────────────────────────────────────────────────
    with tab1:
        display = filt[["domain", "function", "resource", "national"] + STATES].rename(columns={
            "domain": "Domain", "function": "Function",
            "resource": "Resource / capability", "national": "National",
        })
        st.dataframe(display, width="stretch", hide_index=True)
        st.caption(f"{len(filt):,} records shown")

    # ── TAB 2: State allocation heatmap ──────────────────────────────────────
    with tab2:
        st.subheader("Average % of national capability held per state")

        # Exclude Aviation State Contracted/Owned sub-tables to avoid mixing reference pools
        inv_cols = "section" in df.columns
        inv = df[
            ~df["deployable"] & (df["national"] > 0) &
            (~df["domain"].eq("Aviation") | (df.get("section", "") == "National Contracted") if inv_cols
             else ~df["deployable"] & (df["national"] > 0))
        ].copy()
        for s in STATES:
            inv[f"_p{s}"] = inv[s] / inv["national"] * 100

        pct = inv.groupby("domain")[[f"_p{s}" for s in STATES]].mean()
        pct.columns = STATES
        pct.index   = [_AFAC_DOM_SHORT.get(d, d) for d in pct.index]

        fig_hm = px.imshow(
            pct.T,
            text_auto=".1f",
            color_continuous_scale="Blues",
            labels={"x": "Domain", "y": "State", "color": "% of national"},
            title="Mean state share of national capability inventory (%)",
        )
        fig_hm.update_layout(xaxis_tickangle=-20, height=420)
        st.plotly_chart(fig_hm, width="stretch")
        st.caption(
            "Each cell = state's mean share (%) of national capability averaged across all resource types "
            "in that domain. Population-proportional share per state (ABS 2022–23) shown on the State "
            "Capability Profiles page as a more defensible benchmark than a uniform 1/8 split."
        )

    # ── TAB 3: Deployable teams ───────────────────────────────────────────────
    with tab3:
        dep = df[df["deployable"]].copy()
        if dep.empty:
            st.info("No deployable team data found.")
            return

        agg = dep.groupby("domain")["national"].sum().reset_index()
        agg["domain_short"] = agg["domain"].map(_AFAC_DOM_SHORT).fillna(agg["domain"])
        fig_dep = px.bar(
            agg.sort_values("national"),
            x="national", y="domain_short", orientation="h",
            title="Total interstate-deployable team units by domain (national pool, 48-hr)",
            labels={"national": "Deployable units", "domain_short": "Domain"},
            color="national", color_continuous_scale="Reds",
        )
        fig_dep.update_coloraxes(showscale=False)
        fig_dep.update_layout(height=320)
        st.plotly_chart(fig_dep, width="stretch")

        st.subheader("Deployable contribution by state and domain")
        dep_melt = dep.melt(
            id_vars=["domain", "resource", "national"],
            value_vars=STATES, var_name="state", value_name="teams",
        )
        dep_melt = dep_melt[dep_melt["teams"] > 0]
        state_dom = dep_melt.groupby(["state", "domain"])["teams"].sum().reset_index()
        fig_sd = px.bar(
            state_dom, x="state", y="teams", color="domain",
            title="Deployable teams per state, stacked by domain",
            labels={"state": "State", "teams": "Teams", "domain": "Domain"},
            color_discrete_sequence=px.colors.qualitative.Set2,
        )
        st.plotly_chart(fig_sd, width="stretch")

        with st.expander("Full deployable teams table"):
            st.dataframe(
                dep[["domain", "resource", "national"] + STATES].rename(columns={
                    "domain": "Domain", "resource": "Resource / team type",
                    "national": "National total",
                }),
                width="stretch", hide_index=True,
            )


def render_state_capability_profile():  # noqa: C901
    """Per-jurisdiction DRFA activation history and AFAC capability allocation (separate views)."""

    st.title("State Capability Profiles")
    st.caption(
        "Two independent views of each jurisdiction: its DRFA activation history (2006–2026) "
        "and its share of national EM capability (AFAC 2023). These datasets measure different "
        "things — administrative funding activations vs static capability headcount — and are "
        "presented separately rather than combined."
    )

    with st.expander("Data sources and limitations", expanded=False):
        st.markdown("""
**DRFA activation history:** Counts unique AGRN activations where the jurisdiction is the primary
state. Multi-state events are attributed to the most-represented state.
DRFA is a Commonwealth funding mechanism — activations reflect policy eligibility thresholds,
not physical EM demand directly.

**AFAC capability (2023):** For each resource type, the state's count as a percentage of the
national total. The per-domain figure averages those percentages across all resource types
in the domain. A perfectly uniform distribution = 12.5% per jurisdiction.
This is a single static snapshot (2023) and cannot be compared against the 20-year DRFA trend.

**Deployable contribution:** Teams this state contributes to the national mutual aid pool
(48-hr readiness, AFAC 2023).

**Why these datasets are not combined:** DRFA activations are an administrative signal
(funding threshold, political eligibility criteria). AFAC headcounts are a 2023 static snapshot
of declared resources. Linking them as demand vs supply would require assumptions that have
no empirical grounding — the same reason the Capacity Stress Analysis was removed.
        """)

    df_afac = load_afac_capability()
    df_drfa = load_drfa_events()

    if df_afac.empty:
        st.error("AFAC capability data not found.")
        return

    STATES = _AFAC_STATES
    FULL   = _AFAC_STATE_FULL

    sel_state  = st.selectbox(
        "Select jurisdiction",
        STATES,
        format_func=lambda s: f"{s} — {FULL[s]}",
        key="scp_state",
    )
    state_full = FULL[sel_state]

    drfa_state = df_drfa[df_drfa["state_primary"] == state_full]
    inv_data   = df_afac[~df_afac["deployable"] & (df_afac["national"] > 0)].copy()
    dep_data   = df_afac[df_afac["deployable"]].copy()

    for s in STATES:
        inv_data[f"_p{s}"] = inv_data[s] / inv_data["national"] * 100

    # ── Top metrics ───────────────────────────────────────────────────────────
    total_ev = len(drfa_state)
    dep_sum  = int(dep_data[sel_state].sum())
    nat_dep  = int(dep_data["national"].sum())
    dep_pct  = dep_sum / nat_dep * 100 if nat_dep else 0
    inv_pct     = float(inv_data[f"_p{sel_state}"].mean())
    pop_share   = _AUS_STATE_POP_SHARE.get(sel_state, 12.5)

    # ── DRFA metrics ──────────────────────────────────────────────────────────
    st.subheader("DRFA Activation History — " + state_full)
    d1, d2, d3 = st.columns(3)
    d1.metric("DRFA activations (2006–2026)", f"{total_ev}")
    if not drfa_state.empty:
        pk = drfa_state.groupby("year").size().idxmax()
        d2.metric("Peak activation year", str(int(pk)))
        most_common_haz = drfa_state["hazard_group"].mode().iloc[0]
        d3.metric("Most common hazard", most_common_haz)
    else:
        d2.metric("Peak activation year", "—")
        d3.metric("Most common hazard", "—")

    # ── AFAC metrics ──────────────────────────────────────────────────────────
    st.subheader("AFAC Capability — " + state_full + " (2023 snapshot)")
    a1, a2 = st.columns(2)
    a1.metric("Mean capability share", f"{inv_pct:.1f}%",
              delta=f"{inv_pct - pop_share:+.1f} pp vs pop. share ({pop_share}%)",
              delta_color="normal")
    a2.metric("Deployable contribution", f"{dep_sum} teams",
              delta=f"{dep_pct:.1f}% of national pool", delta_color="off")

    tab1, tab2 = st.tabs(["📊 DRFA Activation History", "🔧 AFAC Capability Allocation"])

    # ── TAB 1: DRFA history ───────────────────────────────────────────────────
    with tab1:
        if drfa_state.empty:
            st.info("No DRFA activation events recorded for this jurisdiction.")
        else:
            yearly = (
                drfa_state.groupby(["year", "hazard_group"])
                .size().reset_index(name="events")
            )
            fig_ts = px.bar(
                yearly, x="year", y="events", color="hazard_group",
                title=f"{sel_state} — DRFA activations by year and hazard type",
                labels={"year": "Year", "events": "Events", "hazard_group": "Hazard"},
                color_discrete_sequence=px.colors.qualitative.Set2,
            )
            st.plotly_chart(fig_ts, width="stretch")

            col_a, col_b = st.columns(2)
            haz_sum = drfa_state["hazard_group"].value_counts().reset_index()
            haz_sum.columns = ["Hazard", "Events"]
            with col_a:
                st.dataframe(haz_sum, width="stretch", hide_index=True)
            with col_b:
                fig_pie = px.pie(haz_sum, values="Events", names="Hazard",
                                 title="Hazard breakdown")
                st.plotly_chart(fig_pie, width="stretch")

    # ── TAB 2: Capability allocation ──────────────────────────────────────────
    with tab2:
        dom_pct = inv_data.groupby("domain")[f"_p{sel_state}"].mean().reset_index()
        dom_pct.columns = ["domain", "pct"]
        dom_pct["domain_short"] = dom_pct["domain"].map(_AFAC_DOM_SHORT).fillna(dom_pct["domain"])
        dom_pct = dom_pct.sort_values("pct")

        fig_cap = px.bar(
            dom_pct, x="pct", y="domain_short", orientation="h",
            title=f"{sel_state} — mean share of national capability by domain (%)",
            labels={"pct": "% of national", "domain_short": "Domain"},
            color="pct",
            color_continuous_scale=["#d62728", "#ff7f0e", "#2ca02c"],
            range_color=[0, 25],
        )
        pop_exp = _AUS_STATE_POP_SHARE.get(sel_state, 100 / 8)
        fig_cap.add_vline(x=pop_exp, line_dash="dash", line_color="#1f77b4",
                          annotation_text=f"Pop. share: {pop_exp}% (ABS 2022–23)",
                          annotation_position="top right")
        fig_cap.add_vline(x=12.5, line_dash="dot", line_color="grey",
                          annotation_text="Uniform: 12.5%",
                          annotation_position="bottom right")
        fig_cap.update_coloraxes(showscale=False)
        fig_cap.update_layout(height=380)
        st.plotly_chart(fig_cap, width="stretch")
        st.caption(
            "Blue dashed = population-proportional expected share (ABS 2022–23). "
            "Grey dotted = naive uniform share (1/8 = 12.5%) — shown for reference only; "
            "population share is the more defensible benchmark."
        )

        st.subheader(f"Deployable team contribution — {sel_state}")
        dep_tbl = dep_data[["domain", "resource", "national", sel_state]].copy()
        dep_tbl["% of national"] = (
            dep_tbl[sel_state] / dep_tbl["national"].replace(0, float("nan")) * 100
        ).round(1)
        st.dataframe(
            dep_tbl.rename(columns={
                "domain": "Domain", "resource": "Resource / team type",
                "national": "National total", sel_state: f"{sel_state} teams",
            }),
            width="stretch", hide_index=True,
        )






def render_climate_linkage():  # noqa: C901
    """Climate-phase vs compound disaster season linkage analysis."""
    import numpy as np
    from plotly.subplots import make_subplots as _msp

    st.title("Climate–Disaster Linkage")
    st.caption(
        "Tests whether Australian compound disaster seasons are statistically associated "
        "with ENSO, SAM, and IOD phases. ICA dataset, FY1967–present."
    )

    with st.expander("Methodology & limitations", expanded=False):
        st.markdown("""
**Financial year (FY):** 1 July – 30 June. FY label = start year (FY2000 = Jul 2000 – Jun 2001).

**Compound season:** A FY is *compound* if Gissing et al. (2022) chain-link clustering
finds ≥ 1 cluster of ≥ 2 ICA events above A\\$100M normalised loss within a 91-day window.

**Phase classification per FY:**
| Index | Positive phase | Negative phase | Neutral |
|---|---|---|---|
| ENSO / ONI | El Niño (FY mean ONI ≥ +0.5 °C) | La Niña (≤ −0.5 °C) | otherwise |
| SAM / AAO | Positive SAM (FY mean ≥ +1.0) | Negative SAM (≤ −1.0) | otherwise |
| IOD / DMI | Positive IOD (≥ 3 of 7 active-season months with DMI ≥ +0.4 °C) | Negative IOD (≥ 3 months ≤ −0.4 °C) | otherwise |

IOD uses a sustained-event criterion (≥ 3 months above BoM's ±0.4 °C threshold during the
May–November active season) rather than a seasonal mean. Averaging DMI over the full season
dilutes short events to near-zero and produces near-universal Neutral classification.

**Statistical tests:** Chi-squared test of independence (3-phase × 2, df = 2) followed by
pairwise Fisher's exact tests (each phase vs all others). Requires scipy.

**⚠️ Important limitations — read before interpreting results:**
- **These are exploratory associations, not causal claims.** A statistically significant
  association between climate phase and compound season frequency does not establish causation.
- **ENSO and IOD are not independent.** La Niña tends to co-occur with negative IOD (wet
  conditions); El Niño tends to co-occur with positive IOD (drought/fire). Treating them as
  independent predictors is not supported by the climate literature.
- **Direction of influence is hazard-specific.** La Niña elevates flood/cyclone risk but
  reduces fire risk. Positive IOD elevates fire and drought risk but reduces flood risk.
  No single climate phase is uniformly "adverse" across all disaster types.
- **ICA covers insured losses only.** High insurance penetration in urban NSW/VIC
  systematically inflates ICA footprints relative to rural events, biasing the climate
  signal toward storm/hail perils.
- **Sample size is modest.** With ~58 financial years, individual cells in the contingency
  tables may have counts < 5, at which point chi-squared assumptions are strained.
        """)

    # ── Load data ─────────────────────────────────────────────────────────────
    with st.spinner("Computing compound seasons and climate phases…"):
        try:
            _, cl_ica = load_compound_disasters(100.0, 91)
        except Exception as e:
            st.error(f"Could not load ICA compound data: {e}")
            return
        try:
            fy_clim = load_climate_fy_phases()
        except Exception as e:
            st.error(f"Could not load climate phases: {e}")
            return

    if fy_clim.empty:
        st.error("ONI data unavailable — climate linkage analysis requires the ONI index.")
        return

    # Compound flag per FY (True if ≥1 compound cluster that year)
    fy_compound = (
        cl_ica.groupby("fy")["_is_compound"]
        .any()
        .reset_index()
        .rename(columns={"_is_compound": "is_compound"})
    )
    ica_fy_min = int(cl_ica["fy"].min())
    ica_fy_max = int(cl_ica["fy"].max())

    fy_df = (
        pd.DataFrame({"fy": range(ica_fy_min, ica_fy_max + 1)})
        .merge(fy_compound, on="fy", how="left")
        .merge(fy_clim,     on="fy", how="left")
    )
    fy_df["is_compound"] = fy_df["is_compound"].astype("boolean").fillna(False).astype(bool)

    # Adverse-driver counter: La Niña + Negative SAM + Positive IOD
    fy_df["n_adverse"] = (
        (fy_df["enso_phase"] == "La Niña").astype(int) +
        (fy_df["sam_phase"]  == "Negative SAM").astype(int) +
        (fy_df["iod_phase"]  == "Positive IOD").astype(int)
    )

    total_fy    = len(fy_df)
    compound_fy = int(fy_df["is_compound"].sum())
    base_rate   = compound_fy / total_fy if total_fy > 0 else 0.0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Financial years analysed", total_fy)
    m2.metric("Compound seasons", compound_fy)
    m3.metric("Base compound rate", f"{base_rate:.0%}")
    m4.metric("ICA coverage", f"FY{ica_fy_min}–FY{ica_fy_max}")

    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Contingency Tables", "📈 Timeline", "🔀 Phase Co-occurrence", "📋 Data Table",
    ])

    # ── helpers ───────────────────────────────────────────────────────────────
    _PC = {"El Niño": "#d62728", "La Niña": "#1f77b4", "Neutral": "#7f7f7f",
           "Positive SAM": "#2ca02c", "Negative SAM": "#d62728",
           "Positive IOD": "#ff7f0e", "Negative IOD": "#1f77b4"}

    def _contingency(phase_col: str, phases: list) -> pd.DataFrame:
        rows = []
        for ph in phases:
            mask = fy_df[phase_col] == ph
            n_tot = int(mask.sum())
            n_c   = int((fy_df["is_compound"] & mask).sum())
            rows.append({
                "Phase": ph,
                "Compound FYs":     n_c,
                "Non-compound FYs": n_tot - n_c,
                "Total FYs":        n_tot,
                "P(compound)":      n_c / n_tot if n_tot > 0 else float("nan"),
            })
        return pd.DataFrame(rows)

    def _wilson_ci(k: int, n: int, z: float = 1.96) -> tuple:
        if n == 0:
            return 0.0, 1.0
        p      = k / n
        denom  = 1 + z ** 2 / n
        centre = (p + z ** 2 / (2 * n)) / denom
        margin = z * (p * (1 - p) / n + z ** 2 / (4 * n ** 2)) ** 0.5 / denom
        return max(0.0, centre - margin), min(1.0, centre + margin)

    def _run_tests(cont_df: pd.DataFrame) -> str:
        try:
            from scipy.stats import chi2_contingency, fisher_exact
        except ImportError:
            return "*(Install scipy for statistical tests.)*"
        obs = cont_df[["Compound FYs", "Non-compound FYs"]].values.astype(float)
        if obs.sum() == 0 or (obs == 0).all():
            return "*(Insufficient data.)*"
        try:
            chi2, p_chi2, dof, _ = chi2_contingency(obs)
        except Exception:
            return "*(Chi-squared test failed.)*"
        sig = p_chi2 < 0.05
        out = (
            f"**Chi-squared test:** χ²({dof}) = {chi2:.2f}, p = {p_chi2:.3f} "
            f"{'✓ *significant at α = 0.05*' if sig else '— no significant association at α = 0.05'}"
            "\n\n**Pairwise Fisher's exact tests (phase vs all other phases):**"
        )
        for _, row in cont_df.iterrows():
            a = int(row["Compound FYs"]);     b = int(row["Non-compound FYs"])
            c = int(cont_df["Compound FYs"].sum()) - a
            d = int(cont_df["Non-compound FYs"].sum()) - b
            if a + b + c + d == 0:
                continue
            _, p_f = fisher_exact([[a, b], [c, d]], alternative="two-sided")
            or_v = (a * d) / (b * c) if b * c > 0 else float("nan")
            out += f"\n- **{row['Phase']}:** OR = {or_v:.2f}, p = {p_f:.3f}"
            if p_f < 0.05:
                out += " ✓"
        return out

    # ── TAB 1: CONTINGENCY TABLES ─────────────────────────────────────────────
    with tab1:
        st.subheader("Compound Season Frequency by Climate Phase")
        st.caption(
            f"Base rate: {compound_fy} of {total_fy} financial years ({base_rate:.0%}) "
            "were compound seasons. Each panel tests whether this rate varies by climate phase."
        )

        index_specs = [
            ("ENSO / ONI",  "enso_phase", ["El Niño", "Neutral", "La Niña"],
             "La Niña drives above-average rainfall over eastern Australia — expected to elevate "
             "compound flood and cyclone risk."),
            ("SAM / AAO",   "sam_phase",  ["Positive SAM", "Neutral", "Negative SAM"],
             "Negative SAM intensifies mid-latitude storm tracks and east-coast rainfall. "
             "Positive SAM is associated with drier, warmer conditions."),
            ("IOD / DMI",   "iod_phase",  ["Positive IOD", "Neutral", "Negative IOD"],
             "Positive IOD suppresses Australian rainfall and elevates drought and fire risk. "
             "Negative IOD is associated with increased rainfall. Note: ENSO and IOD are not "
             "independent — La Niña tends to co-occur with negative IOD. "
             "Classified by sustained-event criterion: ≥ 3 of 7 May–Nov active-season months above ±0.4 °C."),
        ]

        for idx_label, phase_col, phases, note in index_specs:
            n_valid = int(fy_df[phase_col].notna().sum())
            with st.expander(f"**{idx_label}** — {n_valid} FYs with phase data", expanded=True):
                cont_df = _contingency(phase_col, phases)
                st.caption(note)

                # Conditional probability bar chart + Wilson CIs
                fig_ct = go.Figure()
                for _, row in cont_df.iterrows():
                    ph = row["Phase"]
                    p  = row["P(compound)"]
                    lo, hi = _wilson_ci(int(row["Compound FYs"]), int(row["Total FYs"]))
                    fig_ct.add_bar(
                        x=[ph], y=[p],
                        error_y=dict(type="data", symmetric=False,
                                     array=[hi - p], arrayminus=[p - lo]),
                        name=ph, marker_color=_PC.get(ph, "#aec7e8"),
                        text=[f"{p:.0%} ({int(row['Compound FYs'])}/{int(row['Total FYs'])})"],
                        textposition="outside",
                    )
                fig_ct.add_hline(
                    y=base_rate, line_dash="dash", line_color="black",
                    annotation_text=f"Base rate {base_rate:.0%}",
                    annotation_position="top right",
                )
                fig_ct.update_layout(
                    showlegend=False,
                    yaxis=dict(title="P(compound season)", tickformat=".0%", range=[0, 1.1]),
                    xaxis_title=idx_label, height=320, margin=dict(t=20, b=10),
                )
                st.plotly_chart(fig_ct, width="stretch")

                # Contingency table
                disp = cont_df.copy()
                disp["P(compound)"] = disp["P(compound)"].apply(
                    lambda v: f"{v:.0%}" if pd.notna(v) else "—"
                )
                st.dataframe(disp, hide_index=True, width="stretch")
                st.markdown(_run_tests(cont_df))

    # ── TAB 2: TIMELINE ───────────────────────────────────────────────────────
    with tab2:
        st.subheader("Compound Seasons and Climate Phase Timeline")
        st.caption(
            "Monthly ONI (top) and FY-mean SAM (bottom). "
            "Gold stars mark compound disaster seasons (ICA, A\\$100M threshold, 91-day window)."
        )

        try:
            oni_monthly = fetch_oni_data()
            oni_monthly["fy"] = oni_monthly["date"].apply(
                lambda d: d.year if d.month >= 7 else d.year - 1
            )
            oni_monthly = oni_monthly[oni_monthly["fy"] >= ica_fy_min].copy()
        except Exception:
            st.warning("Could not load ONI data for timeline.")
            oni_monthly = pd.DataFrame()

        fig_tl = _msp(
            rows=2, cols=1,
            row_heights=[0.72, 0.28],
            shared_xaxes=True,
            vertical_spacing=0.05,
        )

        if not oni_monthly.empty:
            for phase, colour in [("El Niño", "#d62728"), ("Neutral", "#aec7e8"), ("La Niña", "#1f77b4")]:
                mask = oni_monthly["enso_phase"] == phase
                fig_tl.add_scatter(
                    x=oni_monthly.loc[mask, "date"],
                    y=oni_monthly.loc[mask, "oni"],
                    mode="markers", marker=dict(color=colour, size=4, opacity=0.7),
                    name=phase, legendgroup=phase, row=1, col=1,
                )
            fig_tl.add_scatter(
                x=oni_monthly["date"], y=oni_monthly["oni"],
                mode="lines", line=dict(color="#888", width=1),
                name="ONI", showlegend=False, row=1, col=1,
            )
            for y_val, colour in [(0.5, "#d62728"), (-0.5, "#1f77b4"), (0, "#ccc")]:
                fig_tl.add_hline(
                    y=y_val,
                    line_dash="dot" if y_val != 0 else "solid",
                    line_color=colour, line_width=1, row=1, col=1,
                )

            # Gold star markers for compound seasons
            comp_fys = fy_df[fy_df["is_compound"]]["fy"].tolist()
            for i, cfy in enumerate(comp_fys):
                marker_date = pd.Timestamp(year=int(cfy), month=10, day=1)
                oni_val = fy_df.loc[fy_df["fy"] == cfy, "oni_mean"]
                y_star  = float(oni_val.iloc[0]) if len(oni_val) and pd.notna(oni_val.iloc[0]) else 0.0
                fig_tl.add_scatter(
                    x=[marker_date], y=[y_star],
                    mode="markers",
                    marker=dict(symbol="star", size=14, color="gold",
                                line=dict(color="darkred", width=1)),
                    name="Compound season" if i == 0 else "",
                    showlegend=(i == 0),
                    legendgroup="compound",
                    hovertemplate=f"FY{cfy}: compound season<br>ONI mean: {y_star:.2f}<extra></extra>",
                    row=1, col=1,
                )

        # SAM FY-mean line in lower panel
        sam_plot = fy_df[["fy", "sam_mean", "sam_phase"]].dropna(subset=["sam_mean"])
        if not sam_plot.empty:
            sam_dates = sam_plot["fy"].apply(lambda y: pd.Timestamp(year=int(y), month=10, day=1))
            for phase, colour in [("Positive SAM", "#2ca02c"), ("Negative SAM", "#d62728"), ("Neutral", "#aec7e8")]:
                mask = sam_plot["sam_phase"] == phase
                if mask.any():
                    fig_tl.add_scatter(
                        x=sam_dates[mask], y=sam_plot.loc[mask, "sam_mean"],
                        mode="markers", marker=dict(color=colour, size=6, opacity=0.8),
                        name=phase + " (SAM)", legendgroup=phase + "_sam", row=2, col=1,
                    )
            fig_tl.add_scatter(
                x=sam_dates, y=sam_plot["sam_mean"],
                mode="lines", line=dict(color="#555", width=1),
                name="SAM (FY mean)", showlegend=False, row=2, col=1,
            )
            fig_tl.add_hline(y=0, line_dash="solid", line_color="#ccc", line_width=1, row=2, col=1)

        fig_tl.update_yaxes(title_text="ONI (°C)", row=1, col=1)
        fig_tl.update_yaxes(title_text="SAM (FY mean)", row=2, col=1)
        fig_tl.update_xaxes(title_text="Date", row=2, col=1)
        fig_tl.update_layout(
            height=540,
            legend=dict(orientation="h", y=-0.14, font=dict(size=12)),
        )
        st.plotly_chart(fig_tl, width="stretch")

    # ── TAB 3: PHASE CO-OCCURRENCE ────────────────────────────────────────────
    with tab3:
        st.subheader("Climate Phase Co-occurrence — Descriptive Summary")
        st.caption(
            "Counts of financial years where each pair of climate phases was simultaneously active. "
            "This is purely descriptive — it illustrates the known ENSO–IOD teleconnection "
            "(La Niña tends to co-occur with Negative IOD) and should inform how the "
            "contingency table results in Tab 1 are interpreted."
        )
        st.info(
            "**Note on interpretation:** ENSO and IOD are physically coupled. La Niña years "
            "tend to produce negative IOD conditions (wet Indian Ocean dipole), while El Niño "
            "years tend to co-occur with positive IOD. Treating these as independent predictors "
            "of compound disaster risk is not supported by the climate literature.",
            icon="⚠️",
        )

        valid = fy_df[fy_df[["enso_phase", "sam_phase", "iod_phase"]].notna().all(axis=1)].copy()
        if valid.empty:
            st.info("Insufficient data (requires ONI, SAM, and IOD for all FYs).")
        else:
            # Cross-tabulation: ENSO × IOD
            st.markdown("#### ENSO × IOD phase co-occurrence (count of FYs)")
            enso_iod = pd.crosstab(valid["enso_phase"], valid["iod_phase"])
            st.dataframe(enso_iod, width="stretch")

            # Cross-tabulation: ENSO × SAM
            st.markdown("#### ENSO × SAM phase co-occurrence (count of FYs)")
            enso_sam = pd.crosstab(valid["enso_phase"], valid["sam_phase"])
            st.dataframe(enso_sam, width="stretch")

            # Full per-FY phase table for compound vs non-compound seasons
            st.markdown("#### All-index phase summary for compound seasons only")
            comp_only = valid[valid["is_compound"]][["fy", "enso_phase", "sam_phase", "iod_phase"]].copy()
            comp_only["FY"] = "FY" + comp_only["fy"].astype(str)
            st.dataframe(
                comp_only[["FY", "enso_phase", "sam_phase", "iod_phase"]].rename(
                    columns={"enso_phase": "ENSO", "sam_phase": "SAM", "iod_phase": "IOD"}
                ),
                hide_index=True, width="stretch",
            )

    # ── TAB 4: DATA TABLE ─────────────────────────────────────────────────────
    with tab4:
        st.subheader("Per-Financial-Year Climate and Compound Season Data")
        tbl = fy_df[["fy", "is_compound", "oni_mean", "enso_phase",
                     "sam_mean", "sam_phase", "dmi_mean", "iod_phase"]].copy()
        tbl.columns = ["FY", "Compound?", "ONI mean", "ENSO phase",
                       "SAM mean", "SAM phase", "DMI mean (May–Nov)", "IOD phase"]
        tbl["FY"] = "FY" + tbl["FY"].astype(str)
        tbl["Compound?"] = tbl["Compound?"].map({True: "Yes ★", False: "No"})
        for col in ["ONI mean", "SAM mean", "DMI mean (May–Nov)"]:
            tbl[col] = tbl[col].round(3)
        st.dataframe(tbl, hide_index=True, width="stretch")
        download_button(
            fy_df[["fy", "is_compound", "oni_mean", "enso_phase", "sam_mean", "sam_phase", "dmi_mean", "iod_phase"]],
            "climate–compound FY data", "climate_compound_fy.csv",
        )


# ── page objects (pre-defined so render_home can call st.switch_page) ─────────
_PAGE_HOME          = st.Page(render_home,                    title="Home",                        icon="🏠", default=True)
_PAGE_AIDR          = st.Page(render_knowledge_hub,           title="AIDR Event Catalogue",        icon="📋")
_PAGE_ICA           = st.Page(render_ica,                     title="ICA Catastrophes",            icon="💰")
_PAGE_DRFA_ACT      = st.Page(render_drfa_activations,        title="DRFA Activations",            icon="🗂")
_PAGE_DRFA_PAY      = st.Page(render_drfa_payments,           title="DRFA Payments",               icon="💵")
_PAGE_EMDAT         = st.Page(render_emdat,                   title="EM-DAT (CRED)",               icon="🌍")
_PAGE_ONI           = st.Page(render_oni,                     title="ENSO / ONI",                  icon="🌊")
_PAGE_SAM           = st.Page(render_sam,                     title="SAM Index",                   icon="🌬️")
_PAGE_IOD           = st.Page(render_iod,                     title="IOD / DMI",                   icon="🌡️")
_PAGE_MJO           = st.Page(render_mjo,                     title="MJO / RMM",                   icon="🌀")
_PAGE_CLIMATE_SCI   = st.Page(render_climate_science,         title="Climate Science",             icon="📚")
_PAGE_DRFA_MERGED   = st.Page(render_drfa_merged,             title="DRFA Activations + Payments", icon="🔀")
_PAGE_MAP           = st.Page(render_map,                     title="Event Map",                   icon="🗺️")
_PAGE_COMPOUND_ICA  = st.Page(render_compound_disasters,      title="Compound Disasters (ICA)",    icon="⚡")
_PAGE_COMPOUND_DRFA = st.Page(render_compound_disasters_drfa, title="Compound Disasters (DRFA)",   icon="🏛️")
_PAGE_CLIMATE_LINK  = st.Page(render_climate_linkage,         title="Climate–Disaster Linkage",    icon="🌦️")
_PAGE_EM_CONC       = st.Page(render_research_analysis,       title="EM Concurrency Analysis",     icon="📊")
_PAGE_STATE_CO      = st.Page(render_state_cooccurrence,      title="State Co-occurrence",         icon="🔁")
_PAGE_AFAC          = st.Page(render_em_capability,           title="National Capability (AFAC)",  icon="🛡️")
_PAGE_STATE_CAP     = st.Page(render_state_capability_profile,title="State Capability Profiles",   icon="📍")

st.sidebar.markdown("---")

_pg = st.navigation(
    {
        "": [_PAGE_HOME],
        "Source Datasets": [
            _PAGE_AIDR, _PAGE_ICA, _PAGE_DRFA_ACT, _PAGE_DRFA_PAY, _PAGE_EMDAT,
        ],
        "Climate Data": [
            _PAGE_ONI, _PAGE_SAM, _PAGE_IOD, _PAGE_MJO, _PAGE_CLIMATE_SCI,
        ],
        "Integrated Data": [
            _PAGE_DRFA_MERGED, _PAGE_MAP,
        ],
        "Research Analysis": [
            _PAGE_COMPOUND_ICA, _PAGE_COMPOUND_DRFA, _PAGE_CLIMATE_LINK,
            _PAGE_EM_CONC, _PAGE_STATE_CO,
        ],
        "EM Capacity": [
            _PAGE_AFAC, _PAGE_STATE_CAP,
        ],
    }
)

st.sidebar.markdown("---")
st.sidebar.caption("Data: AIDR · ICA · NEMA · EM-DAT · NOAA CPC · BAS · HadISST1.1 · BoM · AFAC")

_pg.run()
